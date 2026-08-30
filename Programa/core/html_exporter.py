"""
core/html_exporter.py
Función pura de construcción de HTML editorial.
No depende de tkinter, customtkinter ni de ningún framework de UI.

Uso:
    from core.html_exporter import build_html

    html_str = build_html(
        bloques           = lista_de_dicts,   # [{"contenido":…, "clasificacion":…, "italic":…}, …]
        referencias_externas = [...],         # list[str] — vacío si no hay .txt cargado
        autores_orcid     = [...],            # list[dict] — vacío si no se llenó la pestaña
        afiliaciones_txt  = "...",            # str — vacío si no se cargó
        figuras           = [...],            # list[dict] — figuras_manuales
        tablas            = [...],            # list[dict] — tablas_manuales
    )
"""

from __future__ import annotations

import re
from collections import defaultdict

from core.constans import HTML_CSS
from core.utils import (
    esc,
    esc_con_etiquetas_editoriales as _esc_con_etiquetas_editoriales,
    insertar_orcid as _insertar_orcid,
    img_to_base64 as _img_to_base64,
    split_afiliaciones_linea as _split_afiliaciones_linea,
    es_encabezado_resumen as _es_encabezado_resumen,
)


# ─── Helpers internos ────────────────────────────────────────────────────────

def _email_duplicado_en_afiliaciones(bloque_texto: str, afiliaciones_txt: str) -> bool:
    """True si el correo que trae bloque_texto (el detectado en el PDF) ya
    aparece, literal, dentro del texto de afiliaciones cargado a mano."""
    m = re.search(r"[\w.\-]+@[\w\-.]+\.\w{2,}", bloque_texto or "")
    if not m:
        return False
    return m.group(0).lower() in (afiliaciones_txt or "").lower()


def _afil_a_html(txt: str) -> str:
    """Convierte líneas de afiliaciones a HTML con superíndices."""
    html_lineas = []
    for linea in txt.splitlines():
        linea = linea.strip()
        if not linea:
            continue
        # Email: * correo@dominio → link mailto
        if re.match(r"^\*\s*[\w\.\-]+@[\w\-\.]+\.\w{2,}", linea):
            email = re.search(r"[\w\.\-]+@[\w\-\.]+\.\w{2,}", linea)
            if email:
                e = email.group(0)
                html_lineas.append(
                    f'<p class="email sin-sangria">'
                    f'* <a href="mailto:{e}">{esc(e)}</a></p>'
                )
            continue

        segs = _split_afiliaciones_linea(linea)
        if segs:
            for marca, resto in segs:
                html_lineas.append(
                    f'<p class="filiaciones sin-sangria">'
                    f'<sup>{esc(marca)}</sup> {esc(resto)}</p>'
                )
        else:
            html_lineas.append(
                f'<p class="filiaciones sin-sangria">{esc(linea)}</p>'
            )
    return "\n".join(html_lineas)


def _render_parrafo_o_lista(texto: str, es_cuerpo: bool,
                             abstract_mode: bool) -> list[str]:
    """Renderiza un bloque de texto como párrafo(s) o lista con bullets '•'."""
    t = texto.strip()
    if not t:
        return []

    def _render_p(s: str) -> str:
        s = s.strip()
        if not s:
            return ""
        s_html = _esc_con_etiquetas_editoriales(s)
        if abstract_mode:
            return f'<p class="abstract sin-sangria">{s_html}</p>'
        if es_cuerpo:
            return f'<p class="cuerpo">{s_html}</p>'
        return f'<p>{s_html}</p>'

    if "•" not in t:
        p = _render_p(t)
        return [p] if p else []

    empieza_con_bullet = t.lstrip().startswith("•")
    partes = [s.strip() for s in re.split(r"\s*•\s*", t) if s.strip()]
    if not partes:
        p = _render_p(t)
        return [p] if p else []

    intro_actual = "" if empieza_con_bullet else partes[0]
    items_actual: list[str] = []
    bloques: list[tuple[str, list[str]]] = []

    idx_ini = 0 if empieza_con_bullet else 1
    for seg in partes[idx_ini:]:
        m = re.search(r"^(.*?[.;])\s+([A-ZÁÉÍÓÚÑ][^:]{4,}:)\s*$", seg)
        if m:
            item_prev = m.group(1).strip()
            if item_prev:
                items_actual.append(item_prev)
            if intro_actual or items_actual:
                bloques.append((intro_actual, items_actual))
            intro_actual = m.group(2).strip()
            items_actual = []
        else:
            items_actual.append(seg)

    if intro_actual or items_actual:
        bloques.append((intro_actual, items_actual))

    out: list[str] = []
    ul_class = "lista-bullets"
    if es_cuerpo:
        ul_class += " cuerpo-list"
    if abstract_mode:
        ul_class += " abstract-list"

    for intro, items in bloques:
        p_intro = _render_p(intro)
        if p_intro:
            out.append(p_intro)
        if items:
            out.append(f'<ul class="{ul_class}">')
            for it in items:
                out.append(f'  <li>{_esc_con_etiquetas_editoriales(it)}</li>')
            out.append("</ul>")

    return out


def _excel_a_html_tabla(ruta: str, hoja: str | None = None) -> str:
    """Convierte una hoja de un .xlsx en tabla HTML con estilos PM."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            return "<p><em>[Tabla vacía]</em></p>"

        html = ['<table class="pm-tabla">']
        html.append("<thead><tr>")
        for celda in filas[0]:
            val = "" if celda is None else str(celda)
            html.append(f"<th>{esc(val)}</th>")
        html.append("</tr></thead><tbody>")
        for fila in filas[1:]:
            if all(c is None for c in fila):
                continue
            html.append("<tr>")
            for celda in fila:
                val = "" if celda is None else str(celda)
                html.append(f"<td>{esc(val)}</td>")
            html.append("</tr>")
        html.append("</tbody></table>")
        return "\n".join(html)
    except ImportError:
        return "<p><em>[Instala openpyxl: pip install openpyxl]</em></p>"
    except Exception as e:
        return f"<p><em>[Error al leer tabla: {esc(str(e))}]</em></p>"


def _buscar_pos_ancla(ancla: str, html: str) -> int:
    """Busca la posición (después del </p>) donde insertar un bloque tras el ancla.
    Devuelve -1 si no se encuentra."""
    if not ancla:
        return -1
    ancla_norm = re.sub(r"[\u00ad\ufffc\ufffe]", "", ancla)
    ancla_norm = re.sub(r"-\s+", "", ancla_norm)
    ancla_norm = re.sub(r"\s+", " ", ancla_norm).strip()
    muestra = ancla_norm[-80:].strip()
    escaped = re.sub(r"([.+*?()\[\]{}\\|^$])", r"\\\1", muestra)
    spacer  = "(?:[\N{SOFT HYPHEN}\ufffc]?\\s*(?:<[^>]+>)?\\s*)+"
    pattern = escaped.replace("\\ ", spacer)
    m = re.search(pattern, html, re.IGNORECASE | re.DOTALL)
    if m:
        cierre = html.find("</p>", m.end())
        return cierre + 4 if cierre != -1 else -1
    return -1


# ─── Función pública principal ────────────────────────────────────────────────

def build_html(
    bloques:              list[dict],
    referencias_externas: list[str],
    autores_orcid:        list[dict],
    afiliaciones_txt:     str,
    figuras:              list[dict],
    tablas:               list[dict],
) -> str:
    """Construye y devuelve el HTML completo del artículo como string.

    Parámetros
    ----------
    bloques:
        Lista de dicts con al menos las claves:
        ``contenido`` (str), ``clasificacion`` (str), ``italic`` (bool).
        La clasificación debe ser la elegida por el usuario (b["menu"].get()
        ya resuelto antes de llamar a esta función).
    referencias_externas:
        Lista de strings con las referencias cargadas desde .txt.
        Si está vacía, se usan los bloques de clase "Referencia" del PDF.
    autores_orcid:
        Lista de dicts ``{"nombre": …, "orcid": …}``.
    afiliaciones_txt:
        Texto crudo de afiliaciones (puede ser cadena vacía).
    figuras:
        Lista de dicts ``{"ruta": …, "pie": …, "ancla": …}``.
    tablas:
        Lista de dicts ``{"ruta": …, "titulo": …, "hoja": …, "ancla": …}``.

    Retorna
    -------
    str — HTML completo listo para guardar en disco.
    """

    # ── Detectar zona de autores del PDF (para ignorarla) ──────────────────
    idx_tit_pri = next(
        (i for i, b in enumerate(bloques) if b["clasificacion"] == "Título principal"),
        None,
    )
    idx_tit_sec = next(
        (i for i, b in enumerate(bloques) if b["clasificacion"] == "Título secundario"),
        None,
    )
    idx_tit_base = idx_tit_sec if idx_tit_sec is not None else idx_tit_pri
    idx_resumen = next(
        (
            i for i, b in enumerate(bloques)
            if b["clasificacion"] == "Encabezado sección"
            and _es_encabezado_resumen(b["contenido"])
            and (idx_tit_base is None or i > idx_tit_base)
        ),
        None,
    )
    zona_autores_pdf: set[int] = set()
    if idx_tit_base is not None and idx_resumen is not None:
        zona_autores_pdf = set(range(idx_tit_base + 1, idx_resumen))

    # ── Afiliaciones ────────────────────────────────────────────────────────
    afil_html       = _afil_a_html(afiliaciones_txt) if afiliaciones_txt else ""
    afil_inyectado  = False

    # ── Autores manuales ────────────────────────────────────────────────────
    autores_html_manual = ""
    autores_inyectado   = False
    if autores_orcid:
        autores_html_manual = (
            f'<p class="autores sin-sangria">'
            f'{_insertar_orcid("", autores_orcid)}</p>'
        )

    hay_titulo_sec = any(b["clasificacion"] == "Título secundario" for b in bloques)

    # ── Contadores de estado para resumen/abstract ──────────────────────────
    primer_nivel1_emitido = False
    contador_resumenes    = 0
    en_bloque_resumen     = False
    bloque_resumen_gris   = False

    # ── Decidir fuente de referencias ───────────────────────────────────────
    usar_refs_externas = bool(referencias_externas)
    refs_a_usar        = referencias_externas if usar_refs_externas else []

    # ── Índice del encabezado "Referencias" ─────────────────────────────────
    idx_refs_start = next(
        (
            i for i, b in enumerate(bloques)
            if b["clasificacion"] == "Encabezado sección"
            and re.search(r"referencia|reference", b["contenido"], re.I)
        ),
        None,
    )

    # ── Separar bloques por tipo ─────────────────────────────────────────────
    cuerpo_bloques: list[dict] = []
    como_citar_lst: list[dict] = []
    fechas_mss_lst: list[dict] = []
    pies_pendientes: list[str] = []

    for i, b in enumerate(bloques):
        cls = b["clasificacion"]
        if i in zona_autores_pdf:
            # Esta zona (entre el título y "Resumen") solo se salta si hay
            # un reemplazo manual cargado para lo que había ahí. Antes se
            # saltaba TODO el rango sin importar la clase, así que si no
            # se había cargado nada a mano, la Filiación detectada
            # automáticamente (que ahora sí sale limpia) desaparecía sin
            # dejar nada en su lugar.
            if cls in ("Cuerpo", "Normal", "Autores") and autores_orcid:
                continue
            if cls == "Filiación" and afiliaciones_txt:
                continue
            if (
                cls == "Email / Metadatos" and afiliaciones_txt
                and _email_duplicado_en_afiliaciones(b["contenido"], afiliaciones_txt)
            ):
                # Evita el correo duplicado cuando ya lo escribiste a mano
                # dentro del texto de Afiliaciones (caso típico: "* correo@
                # ..." pegado debajo de las filiaciones). Si el correo
                # detectado del PDF NO está en tu texto manual, se sigue
                # mostrando — así nunca desaparece un correo real.
                continue
        dentro_de_refs = idx_refs_start is not None and i > idx_refs_start

        if cls == "Cómo citar":
            if not dentro_de_refs:
                como_citar_lst.append(b)
        elif cls == "Fecha manuscrito":
            if not dentro_de_refs:
                fechas_mss_lst.append(b)
        elif cls == "Imagen":
            pass
        elif cls == "Referencia" and usar_refs_externas:
            pass
        else:
            cuerpo_bloques.append(b)

    def _dedup(lst: list[dict]) -> list[dict]:
        seen: set[str] = set()
        out: list[dict] = []
        for b in lst:
            key = b["contenido"].strip()[:80]
            if key not in seen:
                seen.add(key)
                out.append(b)
        return out

    como_citar_lst = _dedup(como_citar_lst)
    fechas_mss_lst = _dedup(fechas_mss_lst)

    # ── Secciones de encabezado con comportamiento especial ─────────────────
    _SECCIONES_CON_LINEA = {
        "resumen", "abstract", "resumen no técnico", "non-technical abstract",
        "referencias", "references", "contribuciones de los autores",
        "contribución de autores", "contribucción de autores",
        "author contribution", "author contributions", "authors' contribution",
        "agradecimientos", "acknowledgements", "acknowledgments",
        "conflicto de intereses", "conflict of interest",
        "conflicts of interest", "competing interests",
    }
    _SECCIONES_GRISES = {"keywords"}
    _SECCIONES_RESUMEN = {
        "resumen", "abstract", "resumen no técnico", "non-technical abstract"
    }

    # ── Construir líneas del HTML ────────────────────────────────────────────
    lineas = [
        "<!DOCTYPE html>",
        '<html lang="es">',
        "<head>",
        '  <meta charset="UTF-8">',
        '  <meta name="viewport" content="width=device-width, initial-scale=1.0">',
        "  <title>Artículo</title>",
        HTML_CSS,
        "</head>",
        "<body><article>",
    ]

    en_refs = False

    for b in cuerpo_bloques:
        cls   = b["clasificacion"]
        texto = esc(b["contenido"])
        ital  = b.get("italic", False)

        if cls == "Ignorar":
            continue
        if usar_refs_externas and en_refs and cls == "Referencia":
            continue

        if cls == "Título principal":
            lineas.append(f'<h1 class="titulo-principal">{texto}</h1>')
            if not hay_titulo_sec:
                if autores_html_manual and not autores_inyectado:
                    lineas.append(autores_html_manual)
                    autores_inyectado = True
                if afil_html and not afil_inyectado:
                    lineas.append(afil_html)
                    afil_inyectado = True

        elif cls == "Título secundario":
            lineas.append(f'<h2 class="titulo-secundario">{texto}</h2>')
            if autores_html_manual and not autores_inyectado:
                lineas.append(autores_html_manual)
                autores_inyectado = True
            if afil_html and not afil_inyectado:
                lineas.append(afil_html)
                afil_inyectado = True

        elif cls == "Autores":
            if not autores_html_manual:
                lineas.append(
                    f'<p class="autores sin-sangria">'
                    f'{_insertar_orcid(b["contenido"], None)}</p>'
                )

        elif cls == "Filiación":
            segs_pdf = _split_afiliaciones_linea(b["contenido"])
            if segs_pdf:
                for marca, resto in segs_pdf:
                    lineas.append(
                        f'<p class="filiaciones sin-sangria">'
                        f'<sup>{esc(marca)}</sup> {esc(resto)}</p>'
                    )
            else:
                lineas.append(f'<p class="filiaciones sin-sangria">{texto}</p>')

        elif cls == "Email / Metadatos":
            txt_link = re.sub(
                r"([\w\.\-]+@[\w\-\.]+\.\w{2,})",
                r'<a href="mailto:\1">\1</a>',
                texto,
            )
            lineas.append(f'<p class="email sin-sangria">{txt_link}</p>')

        elif cls == "Encabezado sección":
            en_refs = bool(re.search(r"referencia|reference", texto, re.I))
            txt_low = texto.strip().lower()

            if txt_low in _SECCIONES_RESUMEN:
                contador_resumenes += 1
                en_bloque_resumen  = True
                bloque_resumen_gris = contador_resumenes >= 2
            elif txt_low not in ("palabras clave", "keywords"):
                en_bloque_resumen  = False
                bloque_resumen_gris = False

            _es_meta = bool(re.search(
                r"issn|volumen\s+\d|vol\.\s*\d|núm\.\s*\d|p\.\s*\d{2,}"
                r"|enero|febrero|marzo|abril|mayo|junio|julio|agosto"
                r"|septiembre|octubre|noviembre|diciembre"
                r"|january|february|march|april|june|july|august"
                r"|september|october|november|december"
                r"|\(\d{4}\)",
                texto, re.I,
            )) or re.match(r"^paleontolog[íi]a mexicana$", texto.strip(), re.I)

            es_gris  = txt_low in _SECCIONES_GRISES
            if txt_low in _SECCIONES_RESUMEN:
                es_gris = bloque_resumen_gris
            con_linea = txt_low in _SECCIONES_CON_LINEA

            if _es_meta:
                clase_h2 = "seccion meta"
            elif con_linea and es_gris:
                clase_h2 = "seccion con-linea gris"
            elif con_linea:
                clase_h2 = "seccion con-linea"
            elif es_gris:
                clase_h2 = "seccion gris"
            else:
                clase_h2 = "seccion"

            lineas.append(f'<h2 class="{clase_h2}">{texto}</h2>')
            if en_refs and refs_a_usar:
                lineas.append('<ol class="referencias">')
                for ref in refs_a_usar:
                    # Mismo motivo que en el bloque "Referencia" del PDF: si
                    # el usuario ya numeró a mano ("1. Autor, A. ..."), se le
                    # saca el número acá — el <ol> ya numera solo, y así no
                    # queda "1. 1. Autor...".
                    ref_limpio = re.sub(r"^\d+[\.\)]\s*", "", esc(ref))
                    lineas.append(f"  <li>{ref_limpio}</li>")
                lineas.append("</ol>")

        elif cls == "Subencabezado":
            en_bloque_resumen  = False
            bloque_resumen_gris = False
            if not primer_nivel1_emitido:
                lineas.append(f'<h3 class="subseccion primer-nivel1">{texto}</h3>')
                primer_nivel1_emitido = True
            else:
                lineas.append(f'<h3 class="subseccion">{texto}</h3>')

        elif cls == "Subencabezado-bajo":
            lineas.append(f'<h3 class="subseccion-bajo">{texto}</h3>')

        elif cls == "Resumen / Abstract":
            tag = "abstract" if ital else "resumen"
            lineas.append(f'<p class="{tag} sin-sangria">{texto}</p>')

        elif cls == "Palabras clave":
            t_kw = esc(b["contenido"])
            t_kw = re.sub(
                r"^(Palabras\s+clave|Keywords)\s*[:\.]?\s*",
                lambda m: f"<strong>{m.group(0).rstrip()}</strong> ",
                t_kw, count=1, flags=re.IGNORECASE,
            )
            if en_bloque_resumen and bloque_resumen_gris:
                lineas.append(
                    f'<p class="keywords sin-sangria" '
                    f'style="color:#666;font-style:italic;">{t_kw}</p>'
                )
            else:
                lineas.append(f'<p class="keywords sin-sangria">{t_kw}</p>')

        elif cls == "Cuerpo del abstract":
            lineas.extend(
                _render_parrafo_o_lista(b["contenido"], es_cuerpo=False, abstract_mode=True)
            )

        elif cls in ("Normal", "Cuerpo"):
            partes_raw   = b["contenido"].split("\n\n")
            partes_unidas: list[str] = []
            for parte in partes_raw:
                parte = parte.strip()
                if not parte:
                    continue
                if parte.startswith("§SUB§"):
                    partes_unidas.append(parte)
                elif partes_unidas and not partes_unidas[-1].startswith("§SUB§"):
                    prev = partes_unidas[-1].rstrip()
                    if prev and prev[-1] not in ".?!:":
                        if re.search(
                            r"[\u2010\u2011\u2012\u2013\u2014-]\s*$", prev
                        ) and re.match(r"^[a-záéíóúñü]", parte):
                            partes_unidas[-1] = re.sub(
                                r"[\u2010\u2011\u2012\u2013\u2014-]\s*$", "", prev
                            ) + parte
                        else:
                            partes_unidas[-1] = prev + " " + parte
                    else:
                        partes_unidas.append(parte)
                else:
                    partes_unidas.append(parte)

            for parte in partes_unidas:
                if parte.startswith("§SUB§"):
                    lineas.append(
                        f'<h3 class="subseccion-bajo">{esc(parte[5:])}</h3>'
                    )
                else:
                    lineas.extend(
                        _render_parrafo_o_lista(parte, es_cuerpo=True, abstract_mode=False)
                    )

        elif cls == "Referencia":
            # Antes cada bloque "Referencia" era un fragmento crudo de
            # PyMuPDF y por eso un <p> por bloque se veía bien. Ahora que
            # pdf_processor.py fusiona todo el texto de referencias en un
            # solo bloque, hay que partirlo acá para no volcar todo junto
            # en un único <p>: cada línea no vacía del bloque se convierte
            # en su propio <li>. Si el usuario ya numeró manualmente las
            # referencias ("1. Autor, A. ..."), se le saca el número al
            # renderizar porque el <ol> ya numera solo (si no, quedaría
            # "1. 1. Autor...").
            entradas = [l.strip() for l in texto.split("\n") if l.strip()]
            if not entradas and texto.strip():
                entradas = [texto.strip()]
            lineas.append('<ol class="referencias">')
            for entrada in entradas:
                entrada = re.sub(r"^\d+[\.\)]\s*", "", entrada)
                lineas.append(f"  <li>{entrada}</li>")
            lineas.append("</ol>")

        elif cls == "Título tabla":
            pass  # las tablas se inyectan por ancla más abajo

        elif cls == "Pie de figura":
            pies_pendientes.append(texto)

    # ── Post-referencias: Cómo citar + Fechas manuscrito ────────────────────
    if como_citar_lst or fechas_mss_lst:
        lineas.append('<div class="post-referencias">')
        for b in como_citar_lst:
            lineas.append(f'<p class="como-citar">{esc(b["contenido"])}</p>')
        if fechas_mss_lst:
            lineas.append('<div class="fechas-manuscrito"><ul>')
            doi_items: list[str] = []
            doi_seen:  set[str]  = set()
            for b in fechas_mss_lst:
                t      = b["contenido"].strip()
                doi_m  = re.search(r"(https?://doi\.org/\S+)", t)
                if doi_m:
                    doi_clean = re.sub(r"\s+", "", doi_m.group(1))
                    if doi_clean not in doi_seen:
                        doi_seen.add(doi_clean)
                        doi_items.append(doi_clean)
                    t = t.replace(doi_m.group(1), "").strip(" .")
                for parte in re.split(r'(?<=[.!?])\s+(?=Manuscrito|Manuscript)', t):
                    parte = parte.strip()
                    if parte:
                        lineas.append(f"  <li>{esc(parte)}</li>")
            for doi in doi_items:
                lineas.append(
                    f'  <li><a href="{doi}" target="_blank">{esc(doi)}</a></li>'
                )
            lineas.append("</ul></div>")
        lineas.append("</div>")

    # ── Cerrar HTML base ─────────────────────────────────────────────────────
    lineas.append("</article></body></html>")
    html_body = "\n".join(lineas)

    # ── Inyectar tablas ──────────────────────────────────────────────────────
    if tablas:
        tablas_ordenadas: list[tuple[int, int, str]] = []
        for idx_t, t_item in enumerate(tablas, 1):
            ancla  = t_item.get("ancla", "").strip()
            # [label] = rótulo (ej. "Tabla 1"); [caption] = descripción.
            rotulo = (t_item.get("rotulo", "") or "").strip() or f"Tabla {idx_t}"
            descripcion = ((t_item.get("descripcion", "") or "").strip()
                           or (t_item.get("titulo", "") or "").strip())   # compat
            thtml  = _excel_a_html_tabla(t_item["ruta"], t_item.get("hoja"))
            bloque = (
                f'\n<div class="tabla-wrapper">\n'
                f'<p class="tabla-titulo"><strong>{esc(rotulo)}.</strong> {esc(descripcion)}</p>\n'
                f'{thtml}\n</div>\n'
            )
            pos = _buscar_pos_ancla(ancla, html_body)
            tablas_ordenadas.append((pos, idx_t, bloque))

        tablas_ordenadas.sort(key=lambda x: (x[0] == -1, x[0], x[1]))
        tablas_inline   = [(p, b) for p, _, b in tablas_ordenadas if p != -1]
        tablas_al_final = [b for p, _, b in tablas_ordenadas if p == -1]

        for pos, bloque in sorted(tablas_inline, key=lambda x: -x[0]):
            html_body = html_body[:pos] + bloque + html_body[pos:]

        if tablas_al_final:
            seccion = (
                '\n<div class="figuras-finales">\n'
                '<h2 style="text-align:center;font-size:10pt;font-weight:700;'
                'margin-bottom:14px;">Tablas</h2>\n'
                + "".join(tablas_al_final)
                + "</div>\n"
            )
            html_body = html_body.replace("</article>", seccion + "</article>", 1)

    # ── Inyectar figuras ─────────────────────────────────────────────────────
    if figuras:
        def _fig_html(i: int, fig: dict) -> str:
            pie_txt = fig.get("pie", "")
            try:
                src = _img_to_base64(fig["ruta"])
            except Exception:
                src = "imagen.jpg"
            cap = (
                f"<strong>Figura {i}.</strong> {esc(pie_txt)}"
                if pie_txt
                else f"<strong>Figura {i}.</strong>"
            )
            return (
                f'<figure id="fig{i}" style="margin:18px auto;text-align:center;">\n'
                f'  <img src="{src}" alt="Figura {i}" '
                f'style="max-width:60%;max-height:420px;border:1px solid #bbb;">\n'
                f'  <figcaption style="font-size:9pt;color:#1a1a1a;'
                f'margin-top:5px;text-align:left;">{cap}</figcaption>\n'
                f'</figure>'
            )

        figs_inline:   list[tuple[int, int, str]] = []
        figs_al_final: list[tuple[int, dict]]     = []

        for i, fig in enumerate(figuras, 1):
            ancla  = fig.get("ancla", "").strip()
            bloque = "\n" + _fig_html(i, fig) + "\n"
            pos    = _buscar_pos_ancla(ancla, html_body)
            if pos != -1:
                figs_inline.append((pos, i, bloque))
            else:
                figs_al_final.append((i, fig))

        grupos: dict[int, list[str]] = defaultdict(list)
        for pos, idx, bloque in sorted(figs_inline, key=lambda x: (x[0], x[1])):
            grupos[pos].append(bloque)

        for pos in sorted(grupos.keys(), reverse=True):
            bloque_conjunto = "".join(grupos[pos])
            html_body = html_body[:pos] + bloque_conjunto + html_body[pos:]

        if figs_al_final:
            figs_html = "\n"
            for i, fig in figs_al_final:
                figs_html += _fig_html(i, fig) + "\n"
            html_body = html_body.replace("</article>", figs_html + "</article>", 1)

    return html_body