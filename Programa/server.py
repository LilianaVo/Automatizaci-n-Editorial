"""
server.py
Backend FastAPI — Editor Semántico Paleontología Mexicana
Reemplaza app_window.py. Expone la lógica de core/ como endpoints REST.
"""

from __future__ import annotations

import os
import re
import sys
import uuid
import shutil
import tempfile
import subprocess
import base64
from pathlib import Path
from typing import Any

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

# ── core ──────────────────────────────────────────────────────────────────────
from core.pdf_processor    import procesar_pdf
from core.jats_exporterv2  import build_jats_xml
from core.html_exporter    import build_html
from core.epub_exporter    import build_epub
from core.xml_validator    import validar_jats
from core import proyecto
from core.constans import (
    OPCIONES,
    CLASE_COMPAT,
    COLORES_UI,
    COLOR_POR_CLASE,
)

# ─────────────────────────────────────────────────────────────────────────────
app = FastAPI(title="Editor Semántico — Paleontología Mexicana")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Directorio de static (HTML/CSS/JS) ───────────────────────────────────────
BASE_DIR = Path(__file__).parent
STATIC_DIR = BASE_DIR / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# ── Estado de sesión en memoria (una sesión a la vez — uso de escritorio) ────
_estado: dict[str, Any] = {
    "bloques":             [],   # list[dict] con clasificacion y contenido
    "referencias_externas":[],
    "figuras_manuales":    [],
    "tablas_manuales":     [],
    "autores_orcid":       [],
    "afiliaciones_txt":    "",
    "fig_dir":             None, # carpeta temporal de figuras extraídas
    "pdf_info":            {},   # nombre, páginas, tamaño
    "metadatos":           {},   # volumen, número, año, páginas, DOI, ISSN, fechas mss.
}


# ═════════════════════════════════════════════════════════════════════════════
# Modelos Pydantic
# ═════════════════════════════════════════════════════════════════════════════

class BloqueUpdate(BaseModel):
    idx: int
    contenido: str | None = None
    clasificacion: str | None = None

class BloquesBulkUpdate(BaseModel):
    bloques: list[dict]   # [{idx, contenido?, clasificacion?}]

class BloqueDividir(BaseModel):
    idx: int
    texto_nuevo: str
    contenido_restante: str
    clasificacion_nuevo: str | None = None

class BloqueNuevo(BaseModel):
    contenido: str
    clasificacion: str | None = None
    insertar_despues: int | None = None   # idx del bloque tras el cual insertar; None = al final

class UnirBloques(BaseModel):
    idx_a: int   # bloque que absorbe (queda)
    idx_b: int   # bloque que se une (desaparece)

class AutoresPayload(BaseModel):
    autores: list[dict]   # [{nombre, orcid}]

class AfiliacionesPayload(BaseModel):
    texto: str

class ReferenciasPayload(BaseModel):
    referencias: list[str]

class FigurasPayload(BaseModel):
    figuras: list[dict]   # [{pie, ruta?}]

class TablasPayload(BaseModel):
    tablas: list[dict]    # [{titulo, contenido}]

class ExportPayload(BaseModel):
    formato: str          # "html" | "xml" | "epub"
    ruta_destino: str     # ruta absoluta elegida por el usuario (desde pywebview dialog)

class MetadatosPayload(BaseModel):
    """Metadatos editoriales del artículo, confirmados/corregidos por el usuario.
    Todos los campos son opcionales: solo se actualizan los que vengan presentes.
    """
    volumen:             str | None = None
    numero:              str | None = None
    anio:                str | None = None
    pagina_inicio:       str | None = None
    pagina_fin:          str | None = None
    doi:                 str | None = None
    issn:                str | None = None
    fecha_recibido:      str | None = None
    fecha_corregido:     str | None = None
    fecha_aceptado:      str | None = None
    fecha_recibido_iso:  str | None = None
    fecha_corregido_iso: str | None = None
    fecha_aceptado_iso:  str | None = None


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Meta / Config
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/")
def root():
    """Sirve el index.html de la app."""
    return FileResponse(str(STATIC_DIR / "index.html"))


@app.get("/api/config")
def get_config():
    """Devuelve constantes de UI que el frontend necesita."""
    return {
        "opciones":      OPCIONES,
        "colores":       {c: col for c, col, _ in COLORES_UI},
        "clase_compat":  CLASE_COMPAT,
    }


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — PDF
# ═════════════════════════════════════════════════════════════════════════════

@app.post("/api/pdf/cargar")
async def cargar_pdf(file: UploadFile = File(...)):
    """
    Recibe el PDF, lo procesa con core.pdf_processor y devuelve los bloques
    clasificados, figuras y tablas detectadas.
    """
    # Guardar en temporal
    sufijo = Path(file.filename).suffix or ".pdf"
    with tempfile.NamedTemporaryFile(delete=False, suffix=sufijo) as tmp:
        contenido = await file.read()
        tmp.write(contenido)
        ruta_tmp = tmp.name

    try:
        resultado = procesar_pdf(ruta_tmp)
    except Exception as e:
        os.unlink(ruta_tmp)
        raise HTTPException(status_code=422, detail=f"Error procesando PDF: {e}")

    # Limpiar fig_dir anterior si existía
    if _estado["fig_dir"] and os.path.isdir(_estado["fig_dir"]):
        shutil.rmtree(_estado["fig_dir"], ignore_errors=True)

    # Guardar en estado de sesión
    # Convertimos los bloques al formato que usa el frontend:
    # {id, contenido, clasificacion, italic, size, bold}
    bloques_ui = []
    for i, b in enumerate(resultado["bloques"]):
        cls = CLASE_COMPAT.get(b.get("clasificacion", "Cuerpo"), b.get("clasificacion", "Cuerpo"))
        bloques_ui.append({
            "id":            i,
            "contenido":     b.get("contenido", ""),
            "clasificacion": cls,
            "italic":        b.get("italic", False),
            "bold":          b.get("bold", False),
            "size":          b.get("size", 12),
        })

    _estado["bloques"]              = bloques_ui
    _estado["figuras_manuales"]     = resultado.get("figuras", [])
    _estado["tablas_manuales"]      = resultado.get("tablas", [])
    _estado["referencias_externas"] = []
    _estado["fig_dir"]              = resultado.get("fig_dir")
    _estado["metadatos"]            = resultado.get("metadatos_detectados", {})
    _estado["pdf_info"]             = {
        "nombre":  file.filename,
        "paginas": resultado.get("body_size", "?"),   # reutilizamos para info
        "resumen": resultado.get("resumen", ""),
        "tamanio": f"{len(contenido) / 1024:.1f} KB",
    }
    _reubicar_tablas_del_pdf()   # .xlsx de tablas a ruta estable (no temporal)

    os.unlink(ruta_tmp)

    return {
        "ok":        True,
        "info":      _estado["pdf_info"],
        "bloques":   bloques_ui,
        "figuras":   _estado["figuras_manuales"],
        "tablas":    _estado["tablas_manuales"],
        "metadatos": _estado["metadatos"],
        "resumen":   resultado.get("resumen", ""),
    }


@app.get("/api/pdf/info")
def get_pdf_info():
    return _estado["pdf_info"]


@app.post("/api/pdf/limpiar")
def limpiar_pdf():
    """Borra solo los bloques y la info del PDF. Conserva autores, afiliaciones y referencias."""
    if _estado["fig_dir"] and os.path.isdir(_estado["fig_dir"]):
        shutil.rmtree(_estado["fig_dir"], ignore_errors=True)
    shutil.rmtree(_RECURSOS_TABLAS, ignore_errors=True)   # .xlsx generados
    _estado["bloques"]          = []
    _estado["figuras_manuales"] = []
    _estado["tablas_manuales"]  = []
    _estado["fig_dir"]          = None
    _estado["pdf_info"]         = {}
    _estado["metadatos"]        = {}
    return {"ok": True}


class RutaPDFPayload(BaseModel):
    ruta: str

@app.post("/api/pdf/cargar-ruta")
def cargar_pdf_por_ruta(payload: RutaPDFPayload):
    """
    Versión para PyWebView: recibe la ruta absoluta del PDF en disco
    (ya que PyWebView no puede hacer drag-and-drop con File API).
    """
    ruta = payload.ruta
    if not os.path.isfile(ruta):
        raise HTTPException(status_code=404, detail=f"Archivo no encontrado: {ruta}")

    try:
        resultado = procesar_pdf(ruta)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Error procesando PDF: {e}")

    if _estado["fig_dir"] and os.path.isdir(_estado["fig_dir"]):
        shutil.rmtree(_estado["fig_dir"], ignore_errors=True)

    bloques_ui = []
    for i, b in enumerate(resultado["bloques"]):
        cls = CLASE_COMPAT.get(b.get("clasificacion", "Cuerpo"), b.get("clasificacion", "Cuerpo"))
        bloques_ui.append({
            "id":            i,
            "contenido":     b.get("contenido", ""),
            "clasificacion": cls,
            "italic":        b.get("italic", False),
            "bold":          b.get("bold", False),
            "size":          b.get("size", 12),
        })

    tam = os.path.getsize(ruta)
    _estado["bloques"]              = bloques_ui
    _estado["figuras_manuales"]     = resultado.get("figuras", [])
    _estado["tablas_manuales"]      = resultado.get("tablas", [])
    _estado["referencias_externas"] = []
    _estado["fig_dir"]              = resultado.get("fig_dir")
    _estado["metadatos"]            = resultado.get("metadatos_detectados", {})
    _estado["pdf_info"]             = {
        "nombre":  Path(ruta).name,
        "paginas": resultado.get("body_size", "?"),
        "resumen": resultado.get("resumen", ""),
        "tamanio": f"{tam / 1024:.1f} KB",
    }
    _reubicar_tablas_del_pdf()   # .xlsx de tablas a ruta estable (no temporal)

    return {
        "ok":        True,
        "info":      _estado["pdf_info"],
        "bloques":   bloques_ui,
        "figuras":   _estado["figuras_manuales"],
        "tablas":    _estado["tablas_manuales"],
        "metadatos": _estado["metadatos"],
    }


# ── Exportar preview (para desarrollo en browser sin PyWebView) ───────────────

from fastapi.responses import Response as FastAPIResponse

@app.post("/api/exportar/html/preview")
def exportar_html_preview():
    if not _estado["bloques"]:
        raise HTTPException(status_code=400, detail="No hay bloques cargados.")
    html_str = build_html(
        bloques=_bloques_snapshot(),
        referencias_externas=_estado["referencias_externas"],
        autores_orcid=_estado["autores_orcid"],
        afiliaciones_txt=_estado["afiliaciones_txt"],
        figuras=_estado["figuras_manuales"],
        tablas=_estado["tablas_manuales"],
    )
    return FastAPIResponse(content=html_str, media_type="text/html",
        headers={"Content-Disposition": "attachment; filename=articulo.html"})


@app.get("/api/exportar/html/vista-previa")
def vista_previa_html():
    """Devuelve el HTML generado sin header de descarga — para mostrarlo en iframe."""
    if not _estado["bloques"]:
        raise HTTPException(status_code=400, detail="No hay bloques cargados.")
    html_str = build_html(
        bloques=_bloques_snapshot(),
        referencias_externas=_estado["referencias_externas"],
        autores_orcid=_estado["autores_orcid"],
        afiliaciones_txt=_estado["afiliaciones_txt"],
        figuras=_estado["figuras_manuales"],
        tablas=_estado["tablas_manuales"],
    )
    return FastAPIResponse(content=html_str, media_type="text/html")

@app.post("/api/exportar/xml/preview")
def exportar_xml_preview():
    if not _estado["bloques"]:
        raise HTTPException(status_code=400, detail="No hay bloques cargados.")
    xml_str = build_jats_xml(
        bloques=_bloques_snapshot(),
        referencias_externas=_estado["referencias_externas"],
        autores_orcid=_estado["autores_orcid"],
        afiliaciones_txt=_estado["afiliaciones_txt"],
        figuras=_estado["figuras_manuales"],
        tablas=_estado["tablas_manuales"],
    )
    return FastAPIResponse(content=xml_str, media_type="application/xml",
        headers={"Content-Disposition": "attachment; filename=articulo.xml"})


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Bloques
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/api/bloques")
def get_bloques():
    return {"bloques": _estado["bloques"]}


@app.patch("/api/bloques/{idx}")
def actualizar_bloque(idx: int, datos: BloqueUpdate):
    if idx < 0 or idx >= len(_estado["bloques"]):
        raise HTTPException(status_code=404, detail="Bloque no encontrado")
    b = _estado["bloques"][idx]
    if datos.contenido is not None:
        b["contenido"] = datos.contenido
    if datos.clasificacion is not None:
        b["clasificacion"] = datos.clasificacion
    return {"ok": True, "bloque": b}


@app.delete("/api/bloques/{idx}")
def eliminar_bloque(idx: int):
    """Elimina un bloque por completo (acción irreversible, distinta de
    'Ignorar', que solo lo excluye del export pero lo conserva en la lista).
    """
    if idx < 0 or idx >= len(_estado["bloques"]):
        raise HTTPException(status_code=404, detail="Bloque no encontrado")

    _estado["bloques"].pop(idx)
    for nuevo_idx, b in enumerate(_estado["bloques"]):
        b["id"] = nuevo_idx

    return {"ok": True, "bloques": _estado["bloques"]}


@app.post("/api/bloques/dividir")
def dividir_bloque(datos: BloqueDividir):
    """Divide un bloque en dos a partir de una selección de texto."""
    idx = datos.idx
    if idx < 0 or idx >= len(_estado["bloques"]):
        raise HTTPException(status_code=404, detail="Bloque no encontrado")

    if not datos.texto_nuevo.strip():
        raise HTTPException(status_code=400, detail="El texto seleccionado está vacío.")

    original = _estado["bloques"][idx]
    clasificacion_nuevo = datos.clasificacion_nuevo or original.get("clasificacion", "Cuerpo")

    bloque_nuevo = {
        "id":            idx + 1,
        "contenido":     datos.texto_nuevo.strip(),
        "clasificacion": clasificacion_nuevo,
        "italic":        original.get("italic", False),
        "bold":          original.get("bold", False),
        "size":          original.get("size", 12),
    }

    original["contenido"] = datos.contenido_restante.strip()

    _estado["bloques"].insert(idx + 1, bloque_nuevo)
    for nuevo_idx, b in enumerate(_estado["bloques"]):
        b["id"] = nuevo_idx

    return {"ok": True, "bloques": _estado["bloques"]}




@app.post("/api/bloques/agregar")
def agregar_bloque(datos: BloqueNuevo):
    """Crea un bloque nuevo escrito manualmente por el usuario.

    Si se provee insertar_despues (idx de un bloque existente), el nuevo
    bloque se inserta justo debajo de ese. Si no se provee, se agrega al final.
    """
    if not datos.contenido.strip():
        raise HTTPException(status_code=400, detail="El contenido no puede estar vacio.")

    clasificacion = datos.clasificacion or "Cuerpo"

    bloque_nuevo = {
        "contenido":     datos.contenido.strip(),
        "clasificacion": clasificacion,
        "italic":        False,
        "bold":          False,
        "size":          12,
    }

    n = len(_estado["bloques"])
    idx_insertar = datos.insertar_despues

    if idx_insertar is not None and 0 <= idx_insertar < n:
        _estado["bloques"].insert(idx_insertar + 1, bloque_nuevo)
    else:
        _estado["bloques"].append(bloque_nuevo)

    for nuevo_idx, b in enumerate(_estado["bloques"]):
        b["id"] = nuevo_idx

    return {"ok": True, "bloques": _estado["bloques"]}

@app.post("/api/bloques/unir")
def unir_bloques(datos: UnirBloques):
    """Une dos bloques: el contenido de idx_b se append al de idx_a, separado
    por un espacio. idx_b se elimina. Siempre se reindexan los ids."""
    n = len(_estado["bloques"])
    a, b = datos.idx_a, datos.idx_b
    if not (0 <= a < n and 0 <= b < n):
        raise HTTPException(status_code=404, detail="Bloque no encontrado")
    if a == b:
        raise HTTPException(status_code=400, detail="No puedes unir un bloque consigo mismo")

    bloque_a = _estado["bloques"][a]
    bloque_b = _estado["bloques"][b]

    bloque_a["contenido"] = (bloque_a["contenido"].rstrip() + " " + bloque_b["contenido"].lstrip()).strip()

    _estado["bloques"].pop(b)
    for nuevo_idx, bl in enumerate(_estado["bloques"]):
        bl["id"] = nuevo_idx

    return {"ok": True, "bloques": _estado["bloques"]}


@app.put("/api/bloques")
def actualizar_bloques_bulk(payload: BloquesBulkUpdate):
    """Actualiza múltiples bloques de una vez (sync desde frontend)."""
    for item in payload.bloques:
        idx = item.get("idx") or item.get("id")
        if idx is None:
            continue
        if 0 <= idx < len(_estado["bloques"]):
            b = _estado["bloques"][idx]
            if "contenido" in item:
                b["contenido"] = item["contenido"]
            if "clasificacion" in item:
                b["clasificacion"] = item["clasificacion"]
    return {"ok": True}


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Autores
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/api/autores")
def get_autores():
    return {"autores": _estado["autores_orcid"]}


@app.put("/api/autores")
def set_autores(payload: AutoresPayload):
    _estado["autores_orcid"] = payload.autores
    return {"ok": True, "total": len(payload.autores)}


class RutaExcelPayload(BaseModel):
    ruta: str

def _parsear_excel_autores(ruta: str) -> tuple[list[dict], int]:
    """Lee un Excel de autores y devuelve (lista_autores, num_importados)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
        wb.close()
    except ImportError:
        raise HTTPException(status_code=500, detail="Instala openpyxl: pip install openpyxl")
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Error leyendo Excel: {e}")

    nuevos = []
    for fila in filas:
        if not fila or all(c is None for c in fila):
            continue
        nombre = str(fila[0]).strip() if fila[0] else ""
        orcid  = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""
        if nombre.lower() in ("autor", "nombre", "author", "name"):
            continue
        if not nombre:
            continue
        m = re.search(r"(\d{4}-\d{4}-\d{4}-\d{3}[\dX])", orcid, re.IGNORECASE)
        orcid_limpio = m.group(1) if m else orcid
        nuevos.append({"nombre": nombre, "orcid": orcid_limpio})
    return nuevos, len(nuevos)


@app.post("/api/autores/excel")
def importar_autores_excel_ruta(payload: RutaExcelPayload):
    """Para PyWebView — recibe ruta absoluta del Excel."""
    if not os.path.isfile(payload.ruta):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    nuevos, total = _parsear_excel_autores(payload.ruta)
    _estado["autores_orcid"].extend(nuevos)
    return {"ok": True, "autores": _estado["autores_orcid"], "importados": total}


@app.post("/api/autores/excel-upload")
async def importar_autores_excel_upload(file: UploadFile = File(...)):
    """Para desarrollo en browser — recibe el archivo directamente."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        ruta_tmp = tmp.name
    try:
        nuevos, total = _parsear_excel_autores(ruta_tmp)
    finally:
        os.unlink(ruta_tmp)
    _estado["autores_orcid"].extend(nuevos)
    return {"ok": True, "autores": _estado["autores_orcid"], "importados": total}


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Afiliaciones
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/api/afiliaciones")
def get_afiliaciones():
    return {"texto": _estado["afiliaciones_txt"]}


@app.put("/api/afiliaciones")
def set_afiliaciones(payload: AfiliacionesPayload):
    _estado["afiliaciones_txt"] = payload.texto
    return {"ok": True}


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Afiliaciones (carga desde .txt)
# ═════════════════════════════════════════════════════════════════════════════

class RutaTxtPayload(BaseModel):
    ruta: str

@app.post("/api/afiliaciones/txt")
def cargar_afiliaciones_txt(payload: RutaTxtPayload):
    """PyWebView: carga afiliaciones desde ruta de .txt en disco."""
    if not os.path.isfile(payload.ruta):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    with open(payload.ruta, encoding="utf-8", errors="replace") as f:
        texto = f.read()
    _estado["afiliaciones_txt"] = texto
    return {"ok": True, "texto": texto}

@app.post("/api/afiliaciones/txt-upload")
async def cargar_afiliaciones_txt_upload(file: UploadFile = File(...)):
    """Browser fallback: recibe el .txt directamente."""
    contenido = await file.read()
    texto = contenido.decode("utf-8", errors="replace")
    _estado["afiliaciones_txt"] = texto
    return {"ok": True, "texto": texto}


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Metadatos editoriales (volumen, número, año, páginas, DOI, ISSN,
# fechas de manuscrito)
# ═════════════════════════════════════════════════════════════════════════════
#
# NOTA IMPORTANTE: estos metadatos (detectados automáticamente y luego
# confirmados/corregidos por el usuario aquí) todavía NO se pasan a
# build_jats_xml() ni a build_html(). Esos exportadores (core/jats_exporterv2.py
# y core/html_exporter.py) tienen su propia extracción interna independiente
# (ver _extract_doi, _extract_year_and_pages, _parse_manuscript_dates en
# jats_exporterv2.py), que vuelve a leer directamente de los bloques del PDF.
# Conectar ambas fuentes es trabajo pendiente (Paso 4): hay que decidir si el
# dato corregido manualmente por el usuario debe tener prioridad sobre el
# extraído automáticamente antes de inyectarlo en el XML/HTML.

@app.get("/api/metadatos")
def get_metadatos():
    return {"metadatos": _estado["metadatos"]}


@app.put("/api/metadatos")
def set_metadatos(payload: MetadatosPayload):
    """Actualiza (parcialmente) los metadatos editoriales del artículo.
    Solo se sobrescriben los campos presentes en el payload; el resto
    conserva el valor ya detectado/guardado previamente.
    """
    actualizados = payload.model_dump(exclude_unset=True)
    _estado["metadatos"].update(actualizados)
    return {"ok": True, "metadatos": _estado["metadatos"]}


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Referencias
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/api/referencias")
def get_referencias():
    return {"referencias": _estado["referencias_externas"]}

@app.put("/api/referencias")
def set_referencias(payload: ReferenciasPayload):
    _estado["referencias_externas"] = payload.referencias
    return {"ok": True, "total": len(payload.referencias)}

@app.post("/api/referencias/txt")
def cargar_referencias_txt(payload: RutaTxtPayload):
    """PyWebView: carga referencias desde ruta de .txt en disco."""
    if not os.path.isfile(payload.ruta):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    try:
        from core.utils import parsear_referencias
        with open(payload.ruta, encoding="utf-8", errors="replace") as f:
            contenido = f.read()
        refs = parsear_referencias(contenido)
        _estado["referencias_externas"] = refs
        return {"ok": True, "referencias": refs, "total": len(refs)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/referencias/txt-upload")
async def cargar_referencias_txt_upload(file: UploadFile = File(...)):
    """Browser fallback."""
    try:
        from core.utils import parsear_referencias
        contenido = (await file.read()).decode("utf-8", errors="replace")
        refs = parsear_referencias(contenido)
        _estado["referencias_externas"] = refs
        return {"ok": True, "referencias": refs, "total": len(refs)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Figuras
# ═════════════════════════════════════════════════════════════════════════════

def _figura_a_dict(fig: dict) -> dict:
    """Convierte una figura a dict con imagen en base64 si existe."""
    f = {k: v for k, v in fig.items() if not k.startswith("_")}
    ruta = f.get("ruta") or ""
    if ruta and os.path.isfile(ruta):
        try:
            with open(ruta, "rb") as fh:
                f["img_b64"] = base64.b64encode(fh.read()).decode()
            ext = Path(ruta).suffix.lower().lstrip(".")
            f["img_mime"] = f"image/{ext if ext in ('png','jpg','jpeg','gif','webp') else 'png'}"
        except Exception:
            f["img_b64"] = None
    return f

@app.get("/api/figuras")
def get_figuras():
    return {"figuras": [_figura_a_dict(f) for f in _estado["figuras_manuales"]]}

@app.put("/api/figuras")
def set_figuras(payload: FigurasPayload):
    _estado["figuras_manuales"] = payload.figuras
    return {"ok": True, "total": len(payload.figuras)}

class RutasImagenesPayload(BaseModel):
    rutas: list[str]   # lista de rutas absolutas (desde PyWebView)

@app.post("/api/figuras/agregar-rutas")
def agregar_figuras_por_rutas(payload: RutasImagenesPayload):
    """PyWebView: agrega imágenes por sus rutas en disco."""
    agregadas = 0
    for ruta in payload.rutas:
        if os.path.isfile(ruta):
            _estado["figuras_manuales"].append({"ruta": ruta, "pie": "", "ancla": ""})
            agregadas += 1
    return {
        "ok": True,
        "agregadas": agregadas,
        "figuras": [_figura_a_dict(f) for f in _estado["figuras_manuales"]],
    }

@app.post("/api/figuras/agregar-upload")
async def agregar_figuras_upload(files: list[UploadFile] = File(...)):
    """Browser fallback: recibe imágenes como upload y las guarda en temp."""
    if _estado["fig_dir"] is None:
        _estado["fig_dir"] = tempfile.mkdtemp(prefix="editor_figs_")
    agregadas = 0
    for file in files:
        ext  = Path(file.filename).suffix or ".png"
        dest = os.path.join(_estado["fig_dir"], f"fig_{len(_estado['figuras_manuales'])+agregadas+1}{ext}")
        with open(dest, "wb") as fh:
            fh.write(await file.read())
        _estado["figuras_manuales"].append({"ruta": dest, "pie": "", "ancla": ""})
        agregadas += 1
    return {
        "ok": True,
        "agregadas": agregadas,
        "figuras": [_figura_a_dict(f) for f in _estado["figuras_manuales"]],
    }

class ActualizarFiguraPayload(BaseModel):
    idx: int
    pie:   str | None = None
    ancla: str | None = None

@app.patch("/api/figuras/{idx}")
def actualizar_figura(idx: int, datos: ActualizarFiguraPayload):
    if idx < 0 or idx >= len(_estado["figuras_manuales"]):
        raise HTTPException(status_code=404, detail="Figura no encontrada")
    f = _estado["figuras_manuales"][idx]
    if datos.pie   is not None: f["pie"]   = datos.pie
    if datos.ancla is not None: f["ancla"] = datos.ancla
    return {"ok": True}

@app.delete("/api/figuras/{idx}")
def eliminar_figura(idx: int):
    if idx < 0 or idx >= len(_estado["figuras_manuales"]):
        raise HTTPException(status_code=404, detail="Figura no encontrada")
    _estado["figuras_manuales"].pop(idx)
    return {"ok": True, "figuras": [_figura_a_dict(f) for f in _estado["figuras_manuales"]]}


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Tablas
# ═════════════════════════════════════════════════════════════════════════════

# La vista previa envía la tabla COMPLETA; el viewport (~8 filas visibles con
# scroll interno) se controla por CSS en la tarjeta.

# ── Carpeta estable de recursos para los .xlsx de tablas ─────────────────────
def _app_data_base() -> Path:
    """Directorio de datos de la app según el SO (persistente, no temporal)."""
    if sys.platform.startswith("win"):
        base = Path(os.environ.get("LOCALAPPDATA") or (Path.home() / "AppData" / "Local"))
    elif sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    else:
        base = Path(os.environ.get("XDG_DATA_HOME") or (Path.home() / ".local" / "share"))
    return base / "EditorSemantico"

# Los .xlsx de tablas (extraídas del PDF, subidas o unidas) se guardan aquí en
# vez de en la carpeta temporal del sistema, para que "Editar en Excel" y las
# ediciones del usuario sobrevivan y no los borre el SO. Contiene solo archivos
# generados por el programa, así que puede limpiarse al cargar un nuevo PDF.
_RECURSOS_TABLAS = _app_data_base() / "recursos" / "tablas"

# RF-04 — guardar/abrir proyecto: carpeta de trabajo donde se extraen los
# recursos del proyecto abierto.
_WORK_DIR = _app_data_base() / "trabajo"

def _leer_filas_xlsx(ruta: str, hoja: str | None,
                     max_filas: int | None = None,
                     max_cols:  int | None = None) -> dict:
    """Lee una hoja de un .xlsx a matriz de strings para vista previa.

    El .xlsx es la fuente de verdad: se relee cada vez, así que refleja lo que
    el usuario haya editado en Excel. Devuelve un dict con la matriz ya capada
    y metadatos de cuántas filas/columnas tiene en total.
    """
    res = {
        "ok": False, "filas": [], "n_filas": 0, "n_cols": 0,
        "truncada_filas": False, "truncada_cols": False, "error": "",
        "mtime": 0.0,
    }
    if not ruta or not os.path.isfile(ruta):
        res["error"] = "archivo no encontrado"
        return res
    try:
        res["mtime"] = os.path.getmtime(ruta)   # para detectar ediciones en Excel
    except OSError:
        pass
    try:
        import openpyxl
    except ImportError:
        res["error"] = "openpyxl no disponible"
        return res
    try:
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
        filas: list[list[str]] = []
        for fila in ws.iter_rows(values_only=True):
            vals = ["" if c is None else str(c) for c in fila]
            if any(v.strip() for v in vals):
                filas.append(vals)
        wb.close()
    except Exception as e:
        res["error"] = str(e)
        return res

    n_cols = max((len(f) for f in filas), default=0)
    filas = [f + [""] * (n_cols - len(f)) for f in filas]   # normaliza ancho
    res["n_filas"], res["n_cols"] = len(filas), n_cols

    if max_filas is not None and len(filas) > max_filas:
        filas = filas[:max_filas]
        res["truncada_filas"] = True
    if max_cols is not None and n_cols > max_cols:
        filas = [f[:max_cols] for f in filas]
        res["truncada_cols"] = True

    res["ok"], res["filas"] = True, filas
    return res

def _sugiere_unir_siguiente(items: list[dict], previews: list[dict], i: int) -> bool:
    """True si la tabla i+1 parece la continuación de i por salto de página (RF-28).

    Heurística: ambas extraídas del PDF (``auto_pdf``), en páginas consecutivas y
    con el mismo número de columnas.
    """
    if i + 1 >= len(items):
        return False
    a, b = items[i], items[i + 1]
    if a.get("origen") != "auto_pdf" or b.get("origen") != "auto_pdf":
        return False
    pa, pb = a.get("pagina"), b.get("pagina")
    if not (isinstance(pa, int) and isinstance(pb, int) and pb == pa + 1):
        return False
    na, nb = previews[i].get("n_cols", 0), previews[i + 1].get("n_cols", 0)
    return na > 0 and na == nb

def _tablas_para_frontend() -> list[dict]:
    """Lista de tablas para la UI: datos limpios + vista previa (matriz capada).

    ``preview`` y ``sugiere_unir_siguiente`` son campos derivados (no se persisten
    en ``_estado``); se recalculan en cada respuesta releyendo el .xlsx.
    """
    items    = _estado["tablas_manuales"]
    previews = [
        _leer_filas_xlsx(t.get("ruta", ""), t.get("hoja"))   # tabla completa
        for t in items
    ]
    salida = []
    for i, t in enumerate(items):
        limpio = {k: v for k, v in t.items()
                  if not k.startswith("_")
                  and k not in ("preview", "sugiere_unir_siguiente")}
        limpio["preview"] = previews[i]
        limpio["sugiere_unir_siguiente"] = _sugiere_unir_siguiente(items, previews, i)
        salida.append(limpio)
    return salida

def _mover_tabla_a_recursos(t: dict) -> str | None:
    """Mueve el .xlsx de una tabla a _RECURSOS_TABLAS y actualiza t['ruta'].

    Devuelve la carpeta de origen (para poder limpiarla si era temporal), o None
    si no había nada que mover. Si el move falla, conserva la ruta original.
    """
    ruta = t.get("ruta", "")
    if not ruta or not os.path.isfile(ruta):
        return None
    _RECURSOS_TABLAS.mkdir(parents=True, exist_ok=True)
    origen_dir = os.path.dirname(os.path.abspath(ruta))
    if origen_dir == os.path.abspath(_RECURSOS_TABLAS):
        return None                                   # ya está en recursos
    dest = _RECURSOS_TABLAS / os.path.basename(ruta)
    if dest.exists():
        dest = _RECURSOS_TABLAS / f"{dest.stem}_{uuid.uuid4().hex[:6]}{dest.suffix}"
    try:
        shutil.move(ruta, str(dest))
        t["ruta"] = str(dest)
        return origen_dir
    except Exception:
        return None

def _limpiar_recursos_huerfanos() -> None:
    """Borra de la carpeta de recursos los .xlsx que ya no referencia ninguna
    tabla del estado (p. ej. sobras de un PDF anterior). Nunca borra archivos
    fuera de la carpeta de recursos, así que los .xlsx propios del usuario están
    a salvo."""
    if not _RECURSOS_TABLAS.is_dir():
        return
    en_uso = {
        os.path.abspath(t["ruta"])
        for t in _estado["tablas_manuales"] if t.get("ruta")
    }
    for f in _RECURSOS_TABLAS.iterdir():
        if f.is_file() and os.path.abspath(str(f)) not in en_uso:
            try:
                f.unlink()
            except OSError:
                pass

def _reubicar_tablas_del_pdf() -> None:
    """Tras cargar un PDF: mueve las tablas extraídas automáticamente
    (origen ``auto_pdf``) desde el temporal del sistema a la carpeta estable de
    recursos, y limpia archivos huérfanos de sesiones previas. Solo mueve
    archivos generados por el PDF; nunca toca los .xlsx propios del usuario.

    Es idempotente: mover primero y limpiar huérfanos después evita borrar una
    tabla que ya viva en recursos.
    """
    _RECURSOS_TABLAS.mkdir(parents=True, exist_ok=True)
    temp_dirs: set[str] = set()
    for t in _estado["tablas_manuales"]:
        if t.get("origen") != "auto_pdf":
            continue
        origen = _mover_tabla_a_recursos(t)
        if origen and os.path.basename(origen).startswith("pm_tab_"):
            temp_dirs.add(origen)
    _limpiar_recursos_huerfanos()
    for d in temp_dirs:                               # limpia los temp ya vaciados
        shutil.rmtree(d, ignore_errors=True)

@app.get("/api/tablas")
def get_tablas():
    return {"tablas": _tablas_para_frontend()}

@app.put("/api/tablas")
def set_tablas(payload: TablasPayload):
    # ``preview``/``sugiere_unir_siguiente`` son derivados: no se persisten.
    _derivados = ("preview", "sugiere_unir_siguiente")
    _estado["tablas_manuales"] = [
        {k: v for k, v in t.items() if k not in _derivados} for t in payload.tablas
    ]
    return {"ok": True, "total": len(_estado["tablas_manuales"])}

class RutasExcelTablasPayload(BaseModel):
    rutas: list[str]

@app.post("/api/tablas/agregar-rutas")
def agregar_tablas_por_rutas(payload: RutasExcelTablasPayload):
    """PyWebView: agrega tablas desde rutas de Excel en disco."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Instala openpyxl")
    agregadas = 0
    for ruta in payload.rutas:
        if not os.path.isfile(ruta):
            continue
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
            for hoja in wb.sheetnames:
                ws = wb[hoja]
                filas = list(ws.iter_rows(values_only=True))
                preview = "\n".join(
                    "\t".join(str(c) if c is not None else "" for c in fila)
                    for fila in filas[:8]
                )
                _estado["tablas_manuales"].append({
                    "ruta":        ruta,
                    "hoja":        hoja,
                    "rotulo":      "",
                    "descripcion": "",
                    "ancla":       "",
                    "origen":      "importada",
                    "contenido":   preview,
                })
                agregadas += 1
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Error leyendo {Path(ruta).name}: {e}")
    return {"ok": True, "agregadas": agregadas, "tablas": _tablas_para_frontend()}

@app.post("/api/tablas/agregar-upload")
async def agregar_tablas_upload(files: list[UploadFile] = File(...)):
    """Browser fallback."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Instala openpyxl")
    agregadas = 0
    _RECURSOS_TABLAS.mkdir(parents=True, exist_ok=True)
    for file in files:
        # Se guarda en la carpeta estable de recursos (no en un temporal del SO)
        # para que el .xlsx sobreviva y sea editable en Excel.
        ruta_dest = str(_RECURSOS_TABLAS / f"subida_{uuid.uuid4().hex[:8]}.xlsx")
        with open(ruta_dest, "wb") as fdst:
            fdst.write(await file.read())
        try:
            wb = openpyxl.load_workbook(ruta_dest, data_only=True)
            for hoja in wb.sheetnames:
                ws = wb[hoja]
                filas = list(ws.iter_rows(values_only=True))
                preview = "\n".join(
                    "\t".join(str(c) if c is not None else "" for c in fila)
                    for fila in filas[:8]
                )
                _estado["tablas_manuales"].append({
                    "ruta":        ruta_dest,
                    "hoja":        hoja,
                    "rotulo":      "",
                    "descripcion": "",
                    "ancla":       "",
                    "origen":      "subida",
                    "contenido":   preview,
                })
                agregadas += 1
            wb.close()
        except Exception as e:
            os.unlink(ruta_dest)
            raise HTTPException(status_code=422, detail=str(e))
    return {"ok": True, "agregadas": agregadas, "tablas": _tablas_para_frontend()}

class ActualizarTablaPayload(BaseModel):
    idx:         int
    rotulo:      str | None = None   # → [label] (ej. "Tabla 1")
    descripcion: str | None = None   # → [caption]
    ancla:       str | None = None
    titulo:      str | None = None   # compat: alias antiguo de descripción

@app.patch("/api/tablas/{idx}")
def actualizar_tabla(idx: int, datos: ActualizarTablaPayload):
    if idx < 0 or idx >= len(_estado["tablas_manuales"]):
        raise HTTPException(status_code=404, detail="Tabla no encontrada")
    t = _estado["tablas_manuales"][idx]
    if datos.rotulo      is not None: t["rotulo"]      = datos.rotulo
    if datos.descripcion is not None: t["descripcion"] = datos.descripcion
    if datos.ancla       is not None: t["ancla"]       = datos.ancla
    if datos.titulo      is not None: t["descripcion"] = datos.titulo  # compat
    return {"ok": True}

@app.delete("/api/tablas/{idx}")
def eliminar_tabla(idx: int):
    if idx < 0 or idx >= len(_estado["tablas_manuales"]):
        raise HTTPException(status_code=404, detail="Tabla no encontrada")
    _estado["tablas_manuales"].pop(idx)
    return {"ok": True, "tablas": _tablas_para_frontend()}

def _abrir_en_sistema(ruta: str) -> None:
    """Abre un archivo con la aplicación predeterminada del sistema operativo.

    El .xlsx sigue siendo la fuente de verdad: el usuario lo edita en Excel y el
    programa lo relee (ver _leer_filas_xlsx). En escritorio (pywebview) el backend
    corre en la máquina del usuario, así que esto abre Excel localmente.
    """
    if sys.platform.startswith("win"):
        os.startfile(ruta)                       # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", ruta])
    else:
        subprocess.Popen(["xdg-open", ruta])

@app.post("/api/tablas/{idx}/abrir-excel")
def abrir_tabla_excel(idx: int):
    """Abre el .xlsx de la tabla en Excel para editarlo (RF-29)."""
    if idx < 0 or idx >= len(_estado["tablas_manuales"]):
        raise HTTPException(status_code=404, detail="Tabla no encontrada")
    ruta = _estado["tablas_manuales"][idx].get("ruta", "")
    if not ruta or not os.path.isfile(ruta):
        raise HTTPException(
            status_code=404,
            detail="El archivo .xlsx de esta tabla ya no existe en disco.",
        )
    try:
        _abrir_en_sistema(ruta)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo abrir el archivo: {e}")
    return {"ok": True, "ruta": ruta}

def _misma_fila(fa: list[str], fb: list[str]) -> bool:
    """Compara dos filas ignorando mayúsculas, espacios y celdas vacías al final."""
    def norm(f):
        s = [str(c).strip().lower() for c in f]
        while s and s[-1] == "":
            s.pop()
        return s
    na = norm(fa)
    return na != [] and na == norm(fb)

def _escribir_xlsx(filas: list[list[str]], titulo_hoja: str = "Tabla") -> str:
    """Escribe una matriz de filas a un .xlsx nuevo (editable en Excel) y
    devuelve su ruta. Usado al unir tablas (RF-28)."""
    import openpyxl
    _RECURSOS_TABLAS.mkdir(parents=True, exist_ok=True)
    ruta = str(_RECURSOS_TABLAS / f"tabla_unida_{uuid.uuid4().hex[:8]}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (titulo_hoja or "Tabla")[:31]
    for fila in filas:
        ws.append(list(fila))
    wb.save(ruta)
    wb.close()
    return ruta

class UnirTablasPayload(BaseModel):
    quitar_encabezado_repetido: bool = True

@app.post("/api/tablas/{idx}/unir-siguiente")
def unir_tabla_siguiente(idx: int, payload: UnirTablasPayload | None = None):
    """Une la tabla idx con la siguiente (idx+1) en un .xlsx combinado (RF-28)."""
    tabs = _estado["tablas_manuales"]
    if idx < 0 or idx + 1 >= len(tabs):
        raise HTTPException(status_code=404, detail="No hay una tabla siguiente para unir")

    a, b = tabs[idx], tabs[idx + 1]
    fa = _leer_filas_xlsx(a.get("ruta", ""), a.get("hoja"))   # sin capar: todo
    fb = _leer_filas_xlsx(b.get("ruta", ""), b.get("hoja"))
    if not fa["ok"] or not fb["ok"]:
        raise HTTPException(
            status_code=422,
            detail="No se pudo leer alguna tabla: " + (fa["error"] or fb["error"]),
        )

    filas_a, filas_b = fa["filas"], fb["filas"]
    quitar = payload.quitar_encabezado_repetido if payload else True
    encabezado_quitado = False
    if quitar and filas_a and filas_b and _misma_fila(filas_a[0], filas_b[0]):
        filas_b = filas_b[1:]
        encabezado_quitado = True

    combinadas = filas_a + filas_b
    ncols = max((len(f) for f in combinadas), default=0)
    combinadas = [f + [""] * (ncols - len(f)) for f in combinadas]

    ruta_nueva = _escribir_xlsx(combinadas, "Tabla_unida")
    nueva = {
        "ruta":        ruta_nueva,
        "hoja":        "Tabla_unida",
        "rotulo":      a.get("rotulo")      or b.get("rotulo")      or "",
        "descripcion": a.get("descripcion") or b.get("descripcion") or "",
        "ancla":       a.get("ancla")       or b.get("ancla")       or "",
        "origen":      "unida",
    }
    tabs[idx:idx + 2] = [nueva]
    return {
        "ok": True,
        "encabezado_quitado": encabezado_quitado,
        "tablas": _tablas_para_frontend(),
    }


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Exportar
# ═════════════════════════════════════════════════════════════════════════════

def _bloques_snapshot() -> list[dict]:
    """Convierte bloques de estado al formato que esperan los exportadores.
    Si hay referencias externas cargadas, corta todo lo que viene después
    del encabezado de sección "Referencias" del PDF — sin importar cómo
    estén clasificados esos bloques (Cuerpo, Referencia, Cómo citar, etc.).
    Así se evitan duplicados al exportar.
    """
    bloques = _estado["bloques"]
    hay_refs_externas = bool(_estado["referencias_externas"])

    if hay_refs_externas:
        # Encontrar el índice del encabezado "Referencias" en el PDF
        idx_corte = None
        for i, b in enumerate(bloques):
            cls  = b.get("clasificacion", "")
            cont = b.get("contenido", "").lower().strip()
            if cls == "Encabezado sección" and re.search(r"referencia|reference", cont):
                idx_corte = i
                break

        # Si encontró el encabezado, cortar ahí (incluye el encabezado,
        # lo excluye el html_exporter al detectar en_refs=True y sustituye
        # con las referencias externas)
        if idx_corte is not None:
            bloques = bloques[:idx_corte + 1]  # +1 para incluir el encabezado

    return [
        {
            "contenido":     b["contenido"],
            "clasificacion": b["clasificacion"],
            "italic":        b.get("italic", False),
        }
        for b in bloques
    ]


@app.post("/api/exportar/html")
def exportar_html(payload: ExportPayload):
    if not _estado["bloques"]:
        raise HTTPException(status_code=400, detail="No hay bloques cargados. Carga un PDF primero.")
    try:
        html_str = build_html(
            bloques              = _bloques_snapshot(),
            referencias_externas = _estado["referencias_externas"],
            autores_orcid        = _estado["autores_orcid"],
            afiliaciones_txt     = _estado["afiliaciones_txt"],
            figuras              = _estado["figuras_manuales"],
            tablas               = _estado["tablas_manuales"],
        )
        with open(payload.ruta_destino, "w", encoding="utf-8") as f:
            f.write(html_str)
        return {"ok": True, "ruta": payload.ruta_destino}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/exportar/xml")
def exportar_xml(payload: ExportPayload):
    if not _estado["bloques"]:
        raise HTTPException(status_code=400, detail="No hay bloques cargados. Carga un PDF primero.")
    try:
        xml_str = build_jats_xml(
            bloques              = _bloques_snapshot(),
            referencias_externas = _estado["referencias_externas"],
            autores_orcid        = _estado["autores_orcid"],
            afiliaciones_txt     = _estado["afiliaciones_txt"],
            figuras              = _estado["figuras_manuales"],
            tablas               = _estado["tablas_manuales"],
        )
        with open(payload.ruta_destino, "w", encoding="utf-8") as f:
            f.write(xml_str)
        return {"ok": True, "ruta": payload.ruta_destino}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/exportar/epub")
def exportar_epub(payload: ExportPayload):
    if not _estado["bloques"]:
        raise HTTPException(status_code=400, detail="No hay bloques cargados. Carga un PDF primero.")
    try:
        snap = _bloques_snapshot()

        html_str = build_html(
            bloques              = snap,
            referencias_externas = _estado["referencias_externas"],
            autores_orcid        = _estado["autores_orcid"],
            afiliaciones_txt     = _estado["afiliaciones_txt"],
            figuras              = _estado["figuras_manuales"],
            tablas               = _estado["tablas_manuales"],
        )

        titulo_art = next(
            (b["contenido"].strip() for b in snap if b["clasificacion"] == "Título principal"),
            "Artículo",
        )
        autores_lista = [
            a["nombre"].strip() for a in _estado["autores_orcid"]
            if a.get("nombre", "").strip()
        ]
        doi_art = ""
        for b in snap:
            m = re.search(r"https?://doi\.org/\S+", b["contenido"])
            if m:
                doi_art = m.group(0).rstrip(".")
                break
        secciones = [
            (re.sub(r"[^a-z0-9]", "-", b["contenido"].strip().lower())[:40].strip("-"),
             b["contenido"].strip())
            for b in snap
            if b["clasificacion"] == "Encabezado sección" and b["contenido"].strip()
        ]

        epub_bytes = build_epub(
            html_str  = html_str,
            titulo    = titulo_art,
            autores   = autores_lista,
            doi       = doi_art,
            secciones = secciones,
        )
        with open(payload.ruta_destino, "wb") as f:
            f.write(epub_bytes)
        return {"ok": True, "ruta": payload.ruta_destino}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/validar/xml")
def validar_xml_endpoint():
    """
    Genera el XML del estado actual y lo valida contra las reglas JATS/SciELO.
    No guarda ningún archivo — solo valida y devuelve el resultado.
    """
    if not _estado["bloques"]:
        raise HTTPException(
            status_code=400,
            detail="No hay bloques cargados. Carga un PDF primero."
        )
    try:
        xml_str = build_jats_xml(
            bloques              = _bloques_snapshot(),
            referencias_externas = _estado["referencias_externas"],
            autores_orcid        = _estado["autores_orcid"],
            afiliaciones_txt     = _estado["afiliaciones_txt"],
            figuras              = _estado["figuras_manuales"],
            tablas               = _estado["tablas_manuales"],
        )
        resultado = validar_jats(xml_str)
        return resultado.to_dict()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Estado general
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/api/estado")
def get_estado():
    """Resumen rápido del estado actual (para el stepper del frontend)."""
    return {
        "tiene_pdf":       bool(_estado["bloques"]),
        "num_bloques":     len(_estado["bloques"]),
        "num_autores":     len(_estado["autores_orcid"]),
        "tiene_afiliaciones": bool(_estado["afiliaciones_txt"].strip()),
        "num_referencias": len(_estado["referencias_externas"]),
        "num_figuras":     len(_estado["figuras_manuales"]),
        "num_tablas":      len(_estado["tablas_manuales"]),
        "tiene_metadatos": bool(_estado["metadatos"]) and any(_estado["metadatos"].values()),
        "pdf_info":        _estado["pdf_info"],
    }


@app.delete("/api/estado")
def resetear_estado():
    """Limpia toda la sesión."""
    if _estado["fig_dir"] and os.path.isdir(_estado["fig_dir"]):
        shutil.rmtree(_estado["fig_dir"], ignore_errors=True)
    shutil.rmtree(_RECURSOS_TABLAS, ignore_errors=True)   # .xlsx generados
    _estado.update({
        "bloques": [], "referencias_externas": [], "figuras_manuales": [],
        "tablas_manuales": [], "autores_orcid": [], "afiliaciones_txt": "",
        "fig_dir": None, "pdf_info": {}, "metadatos": {},
    })
    return {"ok": True}


# ═════════════════════════════════════════════════════════════════════════════
# Endpoints — Proyecto (guardar / abrir)  [RF-04]
# ═════════════════════════════════════════════════════════════════════════════

def _fuente_actual() -> dict:
    info = _estado.get("pdf_info") or {}
    return {"nombre": info.get("nombre", ""), "tipo": info.get("tipo", "pdf")}

def _aplicar_estado_cargado(est: dict, work_dir: str) -> None:
    """Vuelca en _estado un proyecto cargado desde un .pmz."""
    if _estado.get("fig_dir") and os.path.isdir(_estado["fig_dir"]):
        shutil.rmtree(_estado["fig_dir"], ignore_errors=True)
    _estado["bloques"]              = est.get("bloques") or []
    _estado["referencias_externas"] = est.get("referencias_externas") or []
    _estado["autores_orcid"]        = est.get("autores_orcid") or []
    _estado["afiliaciones_txt"]     = est.get("afiliaciones_txt") or ""
    _estado["metadatos"]            = est.get("metadatos") or {}
    _estado["pdf_info"]             = est.get("pdf_info") or {}
    _estado["figuras_manuales"]     = est.get("figuras_manuales") or []
    _estado["tablas_manuales"]      = est.get("tablas_manuales") or []
    _estado["fig_dir"]              = os.path.join(work_dir, "recursos", "figuras")

class RutaProyectoPayload(BaseModel):
    ruta: str

@app.post("/api/proyecto/guardar")
def guardar_proyecto(payload: RutaProyectoPayload):
    """Guarda el proyecto actual como .pmz en la ruta indicada (Guardar como)."""
    if not _estado.get("bloques"):
        raise HTTPException(status_code=400, detail="No hay nada que guardar todavía.")
    ruta = payload.ruta
    if not ruta.lower().endswith(".pmz"):
        ruta += ".pmz"
    try:
        proyecto.empaquetar(_estado, ruta, fuente=_fuente_actual())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo guardar el proyecto: {e}")
    return {"ok": True, "ruta": ruta}

@app.post("/api/proyecto/abrir")
def abrir_proyecto(payload: RutaProyectoPayload):
    """Abre un .pmz y lo carga como proyecto activo."""
    if not os.path.isfile(payload.ruta):
        raise HTTPException(status_code=404, detail="Archivo no encontrado.")
    shutil.rmtree(_WORK_DIR, ignore_errors=True)
    try:
        est = proyecto.cargar(payload.ruta, str(_WORK_DIR))
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    _aplicar_estado_cargado(est, str(_WORK_DIR))
    return {"ok": True, "info": _estado["pdf_info"], "meta": est.get("_meta", {})}


# ─────────────────────────────────────────────────────────────────────────────
# Arranque directo (desarrollo)
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("server:app", host="127.0.0.1", port=8000, reload=False)