from __future__ import annotations

"""Generador XML JATS con salida orientada a SciELO SPS.

El flujo general es:
1) normalizar/limpiar bloques,
2) extraer metadatos (titulos, autores, ISSN/DOI, afiliaciones),
3) construir front/article-meta,
4) mapear bloques a abstracts/cuerpo/referencias,
5) serializar XML legible.
"""

import os
import re
import unicodedata
import xml.etree.ElementTree as ET
from xml.dom import minidom


# Namespaces usados por JATS para enlaces e interoperabilidad.
XLINK_NS = "http://www.w3.org/1999/xlink"
MML_NS = "http://www.w3.org/1998/Math/MathML"
XML_NS = "http://www.w3.org/XML/1998/namespace"
ALI_NS = "http://www.niso.org/schemas/ali/1.0/"
SCIELO_SPS_VERSION = "sps-1.8"
SCIELO_DTD_VERSION = "1.1"
SCIELO_DOCTYPE = (
    '<!DOCTYPE article PUBLIC "-//NLM//DTD JATS (Z39.96) '
    'Journal Publishing DTD v1.1 20151215//EN" '
    '"https://jats.nlm.nih.gov/publishing/1.1/JATS-journalpublishing1.dtd">'
)

ET.register_namespace("xlink", XLINK_NS)
ET.register_namespace("mml", MML_NS)
ET.register_namespace("ali", ALI_NS)


# Patrones de texto para detectar metadata editorial en encabezados.
_META_HEADING_RE = re.compile(
    r"issn|volumen\s+\d|vol\.\s*\d|num\.\s*\d|p\.\s*\d+|"
    r"enero|febrero|marzo|abril|mayo|junio|julio|agosto|"
    r"septiembre|octubre|noviembre|diciembre|"
    r"january|february|march|april|may|june|july|august|"
    r"september|october|november|december|\(\d{4}\)",
    re.IGNORECASE,
)

_DOI_URL_RE = re.compile(r"https?://doi\.org/(\S+)", re.IGNORECASE)
_DOI_CORE_RE = re.compile(r"(10\.\d{4,9}/\S+)", re.IGNORECASE)
_EMAIL_RE = re.compile(r"[\w.\-+%]+@[\w.\-]+\.\w{2,}")
_ISSN_RE = re.compile(r"\b\d{4}-\d{3}[\dXx]\b")
_AFF_LABEL_RE = re.compile(r"^([0-9]+|[A-Za-z]{1,3})\s*[\)\.\:\-]?\s+(.*)$", re.DOTALL)
_PAGE_RANGE_RE = re.compile(r"\bp\.\s*(\d+)\s*[–\-]\s*(\d+)\b", re.IGNORECASE)
_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")


def _split_affiliations_line(linea: str) -> list[tuple[str, str]]:
    """Extrae multiples afiliaciones de una sola linea.

    Ejemplos:
      "1 Institucion ... 2 Otra ..."
      "a Department of Earth Sciences ... b Faculty of ..."
      "a Department of Earth Sciences, University of California"  (una sola)
    """
    t = _clean_text(linea)
    if not t:
        return []

    # ── Caso 1: línea con un solo marcador (número o letra) ──────────────────
    m_simple = re.match(
        r"^([0-9]+|[A-Za-z]{1,3})\s*[\)\.\:\-\*]?\s+([A-ZÁÉÍÓÚÑÜ].*)$",
        t, re.DOTALL
    )
    if m_simple:
        label = m_simple.group(1).strip()
        body  = _clean_text(m_simple.group(2).strip(" -:\t"))
        _pat_interno = re.compile(
            r"[\.;:\)]\s+([0-9]+|[A-Za-z]{1,3})\s+[A-ZÁÉÍÓÚÑÜ]"
        )
        if not _pat_interno.search(body):
            return [(label, body)] if body else []

    # ── Caso 2: varias afiliaciones en la misma línea ────────────────────────
    pat = re.compile(r"([0-9]+|[A-Za-z]{1,3})\s+(?=[A-Za-zÁÉÍÓÚÑÜáéíóúñü])")
    starts: list[tuple[int, int]] = []
    for m in pat.finditer(t):
        s, e = m.start(1), m.end(1)
        if s == 0:
            starts.append((s, e))
            continue
        j = s - 1
        while j >= 0 and t[j].isspace():
            j -= 1
        if j >= 0 and t[j] in ".;:)":
            starts.append((s, e))

    uniq = []
    seen = set()
    for s, e in starts:
        if s in seen:
            continue
        seen.add(s)
        uniq.append((s, e))
    starts = sorted(uniq, key=lambda x: x[0])

    if not starts:
        m = _AFF_LABEL_RE.match(t)
        if not m:
            return []
        label, body = m.group(1).strip(), _clean_text(m.group(2))
        return [(label, body)] if body else []

    out: list[tuple[str, str]] = []
    for i, (s, e) in enumerate(starts):
        lim = starts[i + 1][0] if i + 1 < len(starts) else len(t)
        label = t[s:e].strip()
        body = _clean_text(t[e:lim].strip(" -.:;\t"))
        if body:
            out.append((label, body))
    return out


# --- Helpers de texto --------------------------------------------------------
def _norm(texto: str) -> str:
    """Normaliza texto para comparaciones robustas (sin tildes, minúsculas)."""
    t = unicodedata.normalize("NFKD", texto or "")
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def _clean_text(texto: str) -> str:
    """Limpia espacios y caracteres invisibles frecuentes en PDFs."""
    t = re.sub(r"[\u00ad\ufffc\ufffe]", "", texto or "")
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _split_paragraphs(texto: str) -> list[str]:
    """Divide texto en parrafos y elimina marcadores internos de subtitulos."""
    if not texto:
        return []
    parts = [p.strip() for p in re.split(r"\n{2,}", texto) if p.strip()]
    if not parts:
        parts = [_clean_text(texto)]
    out = []
    for p in parts:
        p = p.replace("§SUB§", "").strip()
        if p:
            out.append(p)
    return out


# --- Extractores de metadatos ------------------------------------------------
def _extract_doi(bloques: list[dict]) -> str:
    """Busca el primer DOI en formato URL o en formato canónico 10.xxxx/..."""
    for b in bloques:
        txt = b.get("contenido", "")
        if not txt:
            continue
        m_url = _DOI_URL_RE.search(txt)
        if m_url:
            return m_url.group(1).rstrip(".),;")
        m_core = _DOI_CORE_RE.search(txt)
        if m_core:
            return m_core.group(1).rstrip(".),;")
    return ""


def _extract_issn(bloques: list[dict]) -> str:
    """Extrae el primer ISSN encontrado en los bloques."""
    for b in bloques:
        txt = b.get("contenido", "")
        if not txt:
            continue
        m = _ISSN_RE.search(txt)
        if m:
            return m.group(0).upper()
    return ""


def _clean_orcid(orcid_raw: str) -> str:
    """Normaliza ORCID y devuelve solo el identificador 0000-0000-0000-0000."""
    m = re.search(r"(\d{4}-\d{4}-\d{4}-\d{3}[\dX])", orcid_raw or "", re.IGNORECASE)
    return m.group(1).upper() if m else ""


def _split_person_name(nombre: str) -> tuple[str, str]:
    """Separa 'Apellido, Nombre' en surname/given-names con fallback simple."""
    t = _clean_text(nombre).strip(" .")
    if not t:
        return "", ""
    if "," in t:
        surname, given = [p.strip() for p in t.split(",", 1)]
        return surname, given
    parts = t.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[-1], " ".join(parts[:-1])


def _guess_country(aff_text: str) -> str:
    """Intenta recuperar país desde el final de la afiliación."""
    parts = [p.strip(" .;") for p in (aff_text or "").split(",") if p.strip(" .;")]
    return parts[-1] if parts else ""


def _country_code(country_name: str) -> str:
    """Mapea algunos países frecuentes a código ISO alfa-2 para SciELO."""
    mapping = {
        "españa": "ES",
        "spain": "ES",
        "chile": "CL",
        "méxico": "MX",
        "mexico": "MX",
        "colombia": "CO",
        "argentina": "AR",
        "venezuela": "VE",
        "perú": "PE",
        "peru": "PE",
        "ecuador": "EC",
        "uruguay": "UY",
        "bolivia": "BO",
        "costa rica": "CR",
        "guatemala": "GT",
        "honduras": "HN",
        "panamá": "PA",
        "panama": "PA",
        "paraguay": "PY",
        "el salvador": "SV",
        "nicaragua": "NI",
        "república dominicana": "DO",
        "republica dominicana": "DO",
        "cuba": "CU",
        "united states": "US",
        "estados unidos": "US",
        "united kingdom": "GB",
        "reino unido": "GB",
        "francia": "FR",
        "france": "FR",
        "australia": "AU",
        "dinamarca": "DK",
        "denmark": "DK",
        "china": "CN",
        "polonia": "PL",
        "poland": "PL",
        "iran": "IR",
        "canadá": "CA",
        "canada": "CA",
    }
    return mapping.get(_norm(country_name), "")



def _detect_country(aff_text):
    mapping = {
        'espana': ('Espana', 'ES'),
        'spain': ('Spain', 'ES'),
        'chile': ('Chile', 'CL'),
        'mexico': ('Mexico', 'MX'),
        'colombia': ('Colombia', 'CO'),
        'argentina': ('Argentina', 'AR'),
        'venezuela': ('Venezuela', 'VE'),
        'peru': ('Peru', 'PE'),
        'ecuador': ('Ecuador', 'EC'),
        'uruguay': ('Uruguay', 'UY'),
        'bolivia': ('Bolivia', 'BO'),
        'costa rica': ('Costa Rica', 'CR'),
        'panama': ('Panama', 'PA'),
        'paraguay': ('Paraguay', 'PY'),
        'el salvador': ('El Salvador', 'SV'),
        'nicaragua': ('Nicaragua', 'NI'),
        'cuba': ('Cuba', 'CU'),
        'united states': ('United States', 'US'),
        'estados unidos': ('Estados Unidos', 'US'),
        'usa': ('United States', 'US'),
        'united kingdom': ('United Kingdom', 'GB'),
        'reino unido': ('Reino Unido', 'GB'),
        'francia': ('Francia', 'FR'),
        'france': ('France', 'FR'),
        'australia': ('Australia', 'AU'),
        'dinamarca': ('Dinamarca', 'DK'),
        'denmark': ('Denmark', 'DK'),
        'china': ('China', 'CN'),
        'polonia': ('Polonia', 'PL'),
        'poland': ('Poland', 'PL'),
        'iran': ('Iran', 'IR'),
        'canada': ('Canada', 'CA'),
        'alemania': ('Alemania', 'DE'),
        'germany': ('Germany', 'DE'),
        'brasil': ('Brasil', 'BR'),
        'brazil': ('Brazil', 'BR'),
        'italia': ('Italia', 'IT'),
        'italy': ('Italy', 'IT'),
        'portugal': ('Portugal', 'PT'),
        'islandia': ('Islandia', 'IS'),
        'iceland': ('Iceland', 'IS'),
    }
    norm_text = _norm(aff_text)
    for key in sorted(mapping, key=len, reverse=True):
        if re.search(r"\b" + re.escape(key) + r"\b", norm_text):
            return mapping[key]
    guess = _guess_country(aff_text)
    code = _country_code(guess)
    if guess and code:
        return guess, code
    return '', ''


_VOLNUM_RE = re.compile(
    r"[Vv]olumen?\s*(\d+)[,\s]+n[úu]m\.?\s*(\d+)" 
    r"|[Vv]ol\.?\s*(\d+)[,\s]+n[úu]m\.?\s*(\d+)"    
    r"|[Vv]ol\.?\s*(\d+)[,\s(]+[Nn]o\.?\s*(\d+)"      
    r"|\bv\.?\s*(\d+)[,\s(]+(\d+)\)",                  
)
_VOL_ONLY_RE = re.compile(
    r"[Vv]olumen?\s*(\d+)|[Vv]ol\.?\s*(\d+)"
)
_NUM_ONLY_RE = re.compile(
    r"n[úu]m\.?\s*(\d+)|[Nn]o\.?\s*(\d+)|[Nn]°\.?\s*(\d+)"
)


def _extract_year_and_pages(front_notes: list[str], doi: str) -> tuple[str, str, str, str, str]:
    year = ""
    fpage = ""
    lpage = ""
    volume = ""
    issue = ""
    for txt in front_notes:
        if not year and "issn" not in _norm(txt):
            m_year = _YEAR_RE.search(txt)
            if m_year:
                year = m_year.group(0)
        if not (fpage and lpage):
            m_pages = _PAGE_RANGE_RE.search(txt)
            if m_pages:
                fpage, lpage = m_pages.group(1), m_pages.group(2)
        if not (volume and issue):
            m_vi = _VOLNUM_RE.search(txt)
            if m_vi:
                groups = [g for g in m_vi.groups() if g is not None]
                if len(groups) >= 2:
                    volume, issue = groups[0], groups[1]
            if not volume:
                m_v = _VOL_ONLY_RE.search(txt)
                if m_v:
                    volume = next(g for g in m_v.groups() if g is not None)
            if not issue:
                m_n = _NUM_ONLY_RE.search(txt)
                if m_n:
                    issue = next(g for g in m_n.groups() if g is not None)
    if not year and doi:
        m_year = _YEAR_RE.search(doi)
        if m_year:
            year = m_year.group(0)
    return year, fpage, lpage, volume, issue


_DATE_MONTHS_ES = {
    "enero": "1", "febrero": "2", "marzo": "3", "abril": "4",
    "mayo": "5", "junio": "6", "julio": "7", "agosto": "8",
    "septiembre": "9", "octubre": "10", "noviembre": "11", "diciembre": "12",
}
_DATE_MONTHS_EN = {
    "january": "1", "february": "2", "march": "3", "april": "4",
    "may": "5", "june": "6", "july": "7", "august": "8",
    "september": "9", "october": "10", "november": "11", "december": "12",
}
_DATE_MONTHS = {**_DATE_MONTHS_ES, **_DATE_MONTHS_EN}

_DATE_PATTERNS = [
    # "Diciembre 20, 2025" / "December 20, 2025"
    re.compile(r"([A-Za-záéíóúñÁÉÍÓÚÑ]+)\s+(\d{1,2})[,\s]+(\d{4})"),
    # "20 de diciembre de 2025" / "20/12/2025"
    re.compile(r"(\d{1,2})\s+de\s+([A-Za-záéíóúñ]+)\s+de\s+(\d{4})"),
    re.compile(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})"),
]


def _parse_date_str(txt: str) -> tuple[str, str, str]:
    """Returns (day, month, year) as strings, or ('','','')."""
    for pat in _DATE_PATTERNS:
        m = pat.search(txt)
        if not m:
            continue
        g = m.groups()
        if g[0].isdigit() and g[1].isdigit() and g[2].isdigit():
            # dd/mm/yyyy
            return g[0], g[1], g[2]
        if g[0].isdigit() and not g[1].isdigit():
            # "20 de diciembre de 2025"
            month_name = g[1].lower()
            month = _DATE_MONTHS.get(month_name, "")
            return g[0], month, g[2]
        # "Diciembre 20, 2025"
        month_name = g[0].lower()
        month = _DATE_MONTHS.get(month_name, "")
        return g[1], month, g[2]
    return "", "", ""


def _parse_manuscript_dates(date_notes: list[str]) -> list[tuple[str, str, str, str]]:
    """Returns list of (date_type, day, month, year) from 'Fecha manuscrito' blocks.
    date_type is one of: 'received', 'rev-recd', 'accepted'.
    """
    _LABEL_MAP = {
        "recibido": "received",
        "received": "received",
        "corregido": "rev-recd",
        "revised": "rev-recd",
        "rev-recd": "rev-recd",
        "aceptado": "accepted",
        "accepted": "accepted",
    }
    results = []
    combined = " ".join(date_notes)
    # Split on "Manuscrito X:" / "Manuscript X:" boundaries
    parts = re.split(
        r"(?:Manuscrito|Manuscript)\s+(\w+)\s*:",
        combined,
        flags=re.IGNORECASE,
    )
    i = 1
    while i + 1 < len(parts):
        label_raw = parts[i].strip().lower()
        value_raw = parts[i + 1].strip()
        date_type = _LABEL_MAP.get(label_raw, "")
        if date_type:
            day, month, year = _parse_date_str(value_raw)
            if year:
                results.append((date_type, day, month, year))
        i += 2
    return results


def _is_body_heading(texto: str) -> bool:
    """Filtra encabezados espurios de portada para el body."""
    low = _norm(texto) 
    # Numerado: "1. Introducción", "2.1 Métodos", etc.
    if re.match(r"^\d+(\.\d+)*[\s\.\-]+", texto.strip()):
        return True
    # Lista blanca de encabezados comunes
    return low in {
        "introduccion", "material y metodos", "materiales y metodos",
        "resultados", "resultados y discusion", "resultados y discusion",
        "discusion", "conclusiones", "agradecimientos",
        "conflicto de intereses", "contribuciones de los autores",
    }


# --- Parseo de autores/afiliaciones -----------------------------------------
def _authors_from_manual(autores_orcid: list[dict]) -> list[dict]:
    """Construye autores desde la pestaña manual (nombre + ORCID opcional)."""
    out = []
    for a in autores_orcid or []:
        nombre = _clean_text(a.get("nombre", ""))
        orcid = _clean_orcid(a.get("orcid", ""))
        if not nombre:
            continue
        out.append({"nombre": nombre, "orcid": orcid})
    return out


def _authors_from_pdf(bloques: list[dict]) -> list[dict]:
    """Fallback: toma autores detectados en el PDF cuando no hay carga manual."""
    out = []
    for b in bloques:
        if b.get("clasificacion") != "Autores":
            continue
        txt = b.get("contenido", "")
        for raw in [p.strip() for p in txt.split(";") if p.strip()]:
            nom = re.sub(r"[\d,\*\u00b9\u00b2\u00b3\u2070-\u209f]+$", "", raw).strip()
            if nom:
                out.append({"nombre": nom, "orcid": ""})
    return out


def _parse_affiliations_txt(afiliaciones_txt: str) -> tuple[list[dict], list[str]]:
    """Parsea afiliaciones desde .txt y correos de correspondencia.

    Los IDs XML (aff1, aff2, ...) siempre se generan únicos.
    La etiqueta visible puede conservar el prefijo original del txt (numero o letra).
    """
    afils: list[dict] = []
    emails: list[str] = []
    id_seq = 1
    for raw in (afiliaciones_txt or "").splitlines():
        linea = _clean_text(raw)
        if not linea:
            continue

        mail = _EMAIL_RE.search(linea)
        if linea.startswith("*") and mail:
            emails.append(mail.group(0))
            continue

        segs = _split_affiliations_line(linea)
        if segs:
            for label, resto in segs:
                afils.append({"id": f"aff{id_seq}", "label": label, "text": resto})
                id_seq += 1
        else:
            afils.append({"id": f"aff{id_seq}", "label": str(id_seq), "text": linea})
            id_seq += 1
    return afils, emails


def _fallback_affiliations_from_blocks(bloques: list[dict]) -> tuple[list[dict], list[str]]:
    """Fallback cuando no hay txt: usa bloques clasificados de afiliacion/email."""
    afils: list[dict] = []
    emails: list[str] = []
    seq = 1
    for b in bloques:
        cls = b.get("clasificacion")
        txt = _clean_text(b.get("contenido", ""))
        if not txt:
            continue
        if cls == "Filiación":
            afils.append({"id": f"aff{seq}", "label": str(seq), "text": txt})
            seq += 1
        elif cls == "Email / Metadatos":
            for em in _EMAIL_RE.findall(txt):
                if em not in emails:
                    emails.append(em)
    return afils, emails


# --- Referencias y keywords --------------------------------------------------
def _strip_ref_prefix(ref: str) -> str:
    """Quita prefijos numerados tipo '1.' o '[2]' de una referencia."""
    return re.sub(r"^\s*(?:\[?\d+[\.\)\]]\s*)", "", (ref or "").strip())


def _parse_keywords(texto: str) -> list[str]:
    """Separa keywords por coma/punto y coma y limpia etiqueta inicial."""
    t = _clean_text(texto)
    t = re.sub(r"^(palabras\s+clave|keywords)\s*[:\.]?\s*", "", t, flags=re.IGNORECASE)
    if not t:
        return []
    parts = [p.strip(" .;") for p in re.split(r"[;,]", t) if p.strip(" .;")]
    return parts


# --- Tablas (Excel -> XML JATS) ---------------------------------------------
def _table_rows_from_excel(ruta: str, hoja: str | None) -> tuple[list[list[str]], str]:
    """Lee una hoja de Excel y la transforma a matriz de strings."""
    try:
        import openpyxl
    except Exception:
        return [], "openpyxl no disponible"

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            vals = []
            has_content = False
            for cell in row:
                txt = "" if cell is None else _clean_text(str(cell))
                if txt:
                    has_content = True
                vals.append(txt)
            if has_content:
                rows.append(vals)
        wb.close()
        return rows, ""
    except Exception as exc:
        return [], str(exc)


def _append_table_xml(table_wrap: ET.Element, ruta: str, hoja: str | None) -> None:
    """Inserta contenido tabular JATS (<table>) dentro de <table-wrap>."""
    rows, err = _table_rows_from_excel(ruta, hoja)
    if not rows:
        p = ET.SubElement(table_wrap, "p")
        p.text = f"[No se pudo leer tabla: {err or 'tabla vacia'}]"
        return

    max_cols = max(len(r) for r in rows)
    norm_rows = [r + [""] * (max_cols - len(r)) for r in rows]

    tbl = ET.SubElement(table_wrap, "table")
    thead = ET.SubElement(tbl, "thead")
    trh = ET.SubElement(thead, "tr")
    for cell in norm_rows[0]:
        th = ET.SubElement(trh, "th")
        th.text = cell

    tbody = ET.SubElement(tbl, "tbody")
    for row in norm_rows[1:]:
        tr = ET.SubElement(tbody, "tr")
        for cell in row:
            td = ET.SubElement(tr, "td")
            td.text = cell


# --- Parser de referencias APA → element-citation JATS ----------------------

# Patrones para parsear referencias en formato APA
_APA_YEAR_RE      = re.compile(r"\((\d{4}(?:[,\s]+[^)]+)?|in press|en prensa)\)\.\s*", re.IGNORECASE)
_APA_DOI_RE       = re.compile(r"https?://doi\.org/(\S+?)[\.\s]*$", re.IGNORECASE)
_APA_URL_RE       = re.compile(r"(https?://\S+?)[\.\s]*$", re.IGNORECASE)
_APA_VOL_RE       = re.compile(r",\s*(\d+)\s*\((\d+)\)\s*,\s*(.+)$")   # vol(num), pages
_APA_VOL_ONLY_RE  = re.compile(r",\s*(\d+)\s*,\s*(.+)$")                 # vol, pages
_APA_PAGES_RE     = re.compile(r"([\de]\d*[\w]*)\s*[–\-]\s*([\de]\d*[\w]*)$|^([\de]\d*[\w]*)$")
_APA_ED_BOOK_RE   = re.compile(r"\(Ed[s]?\.\)|editor|eds\.", re.IGNORECASE)
_APA_AUTHOR_SEP   = re.compile(r"&amp;|&|\band\b", re.IGNORECASE)


def _detect_pub_type(ref: str) -> str:
    """Detecta tipo de publicación a partir de patrones en la referencia."""
    low = ref.lower()
    # Si tiene DOI de journal conocido o patrón vol,pages → journal
    if "doi.org" in low:
        return "journal"
    if re.search(r",\s*\d+\s*[\(,]", ref):
        return "journal"
    if any(x in low for x in ["(ed.)", "(eds.)", " in ", "pp. "]):
        return "book"
    if any(x in low for x in [".pdf", "fundación", "fundacion", "informe", "report"]):
        return "report"
    if "http" in low:
        return "webpage"
    return "journal"


def _parse_apa_authors(author_str: str) -> list[tuple[str, str]]:
    """Parsea cadena de autores APA en lista de (apellido, iniciales).

    Maneja: 'García, A. B.', 'García, A. B., & López, C.'
    """
    # Quitar "& " o "and " al final antes del último autor
    s = re.sub(r"\s*[&]\s*", ", ", author_str.strip())
    s = re.sub(r"\band\b\s*", "", s, flags=re.IGNORECASE)
    # Dividir por coma-espacio-inicial o punto-coma
    # Patrón: Apellido, I. I. — mantener pares juntos
    authors_raw = re.split(r",\s*(?=[A-ZÁÉÍÓÚÑÜ][a-záéíóúñü]|\.\s*[A-Z]|$)", s)
    result = []
    i = 0
    parts = [p.strip(" .,") for p in authors_raw if p.strip(" .,")]
    # Reagrupar: si parte siguiente son solo iniciales, unir con la anterior
    merged = []
    j = 0
    while j < len(parts):
        p = parts[j]
        # Si parece apellido (sin punto al final ni solo iniciales)
        if j + 1 < len(parts) and re.match(r'^[A-ZÁÉÍÓÚÑÜ][\.\s]+', parts[j+1]):
            merged.append(p + ", " + parts[j+1])
            j += 2
        else:
            merged.append(p)
            j += 1
    for entry in merged:
        entry = entry.strip(" ,.")
        if not entry:
            continue
        if "," in entry:
            surname, given = entry.split(",", 1)
            result.append((_clean_text(surname), _clean_text(given)))
        else:
            result.append((entry, ""))
    return result


def _parse_apa_pages(pages_str: str) -> tuple[str, str]:
    """Extrae fpage/lpage o elocation de una cadena de páginas."""
    s = pages_str.strip(" .")
    m = re.search(r"(\d+)\s*[–\-]\s*(\d+)", s)
    if m:
        return m.group(1), m.group(2)
    # elocation tipo e2021AV000436 o eabc5654
    m2 = re.match(r"^(e\w+|\d+)$", s)
    if m2:
        return s, ""
    return s, ""


def _build_element_citation(ref_el: ET.Element, ref_text: str) -> None:
    """Construye <element-citation> dentro de <ref> parseando formato APA.

    Si el parseo falla o es ambiguo, cae back a <mixed-citation> para
    no perder información.
    """
    txt = _clean_text(ref_text)
    if not txt:
        mc = ET.SubElement(ref_el, "mixed-citation", {"publication-type": "journal"})
        mc.text = txt
        return

    pub_type = _detect_pub_type(txt)

    # ── Extraer DOI / URL ────────────────────────────────────────────────────
    doi_val = ""
    url_val = ""
    m_doi = _APA_DOI_RE.search(txt)
    if m_doi:
        doi_val = m_doi.group(1).rstrip(".),;")
        txt = txt[:m_doi.start()].strip(" .")
    else:
        m_url = _APA_URL_RE.search(txt)
        if m_url:
            url_val = m_url.group(1).rstrip(".),;")
            txt = txt[:m_url.start()].strip(" .")

    # ── Extraer año ──────────────────────────────────────────────────────────
    year_val = ""
    m_year = _APA_YEAR_RE.search(txt)
    if m_year:
        year_val = m_year.group(1).strip()
        author_part = txt[:m_year.start()].strip(" .")
        rest = txt[m_year.end():].strip()
    else:
        # fallback: buscar año entre paréntesis en cualquier posición
        m_year2 = re.search(r"\((\d{4})\)", txt)
        if m_year2:
            year_val = m_year2.group(1)
            author_part = txt[:m_year2.start()].strip(" .")
            rest = txt[m_year2.end():].strip(" .")
        else:
            # Sin año detectable → mixed-citation de seguridad
                mc = ET.SubElement(ref_el, "mixed-citation", {"publication-type": pub_type})
                mc.text = ref_text
                return

    # ── Separar título del artículo de la fuente (revista / libro) ───────────
    # En APA: resto = "Título del artículo. Revista, vol(num), páginas."
    article_title = ""
    source = ""
    volume = ""
    issue = ""
    fpage = ""
    lpage = ""
    elocation = ""

    # Dividir rest en oraciones por ". "
    sentences = re.split(r"\.\s+", rest.rstrip("."))
    sentences = [s.strip() for s in sentences if s.strip()]

    if pub_type == "journal":
        # sentences[0] = título artículo
        # sentences[1] = "Revista, vol(num), páginas" o "Revista, vol, páginas"
        if len(sentences) >= 1:
            article_title = sentences[0]
        if len(sentences) >= 2:
            journal_part = sentences[1]
            # Intentar vol(num), pages
            m_v = _APA_VOL_RE.search(journal_part)
            if m_v:
                source = journal_part[:m_v.start()].strip(" ,")
                volume = m_v.group(1)
                issue = m_v.group(2)
                fpage, lpage = _parse_apa_pages(m_v.group(3))
                if not lpage:
                    elocation = fpage
                    fpage = ""
            else:
                m_vo = _APA_VOL_ONLY_RE.search(journal_part)
                if m_vo:
                    source = journal_part[:m_vo.start()].strip(" ,")
                    volume = m_vo.group(1)
                    fpage, lpage = _parse_apa_pages(m_vo.group(2))
                    if not lpage:
                        elocation = fpage
                        fpage = ""
                else:
                    source = journal_part
        # Texto extra (ej. párrafo 3 en adelante) → ignorar para element-citation
    else:
        # Libro / reporte / web: título es el primer elemento
        if sentences:
            article_title = sentences[0]
        if len(sentences) >= 2:
            source = sentences[1]

    # ── Parsear autores ──────────────────────────────────────────────────────
    authors_parsed = _parse_apa_authors(author_part) if author_part else []

    # ── Construir element-citation ───────────────────────────────────────────
    ec = ET.SubElement(ref_el, "element-citation", {"publication-type": pub_type})

    if authors_parsed:
        pg = ET.SubElement(ec, "person-group", {"person-group-type": "author"})
        for surname, given in authors_parsed:
            if not surname:
                continue
            nm = ET.SubElement(pg, "name")
            sn = ET.SubElement(nm, "surname")
            sn.text = surname
            if given:
                gn = ET.SubElement(nm, "given-names")
                gn.text = given

    if year_val:
        yr = ET.SubElement(ec, "year", {"iso-8601-date": re.sub(r"\D.*", "", year_val)})
        yr.text = year_val

    if article_title:
        at = ET.SubElement(ec, "article-title")
        at.text = article_title

    if source:
        src = ET.SubElement(ec, "source")
        src.text = source

    if volume:
        v = ET.SubElement(ec, "volume")
        v.text = volume

    if issue:
        iss = ET.SubElement(ec, "issue")
        iss.text = issue

    if fpage:
        fp = ET.SubElement(ec, "fpage")
        fp.text = fpage
    if lpage:
        lp = ET.SubElement(ec, "lpage")
        lp.text = lpage
    if elocation and not fpage:
        el = ET.SubElement(ec, "elocation-id")
        el.text = elocation

    if doi_val:
        pub_id = ET.SubElement(ec, "pub-id", {"pub-id-type": "doi"})
        pub_id.text = doi_val
    elif url_val:
        ext = ET.SubElement(ec, "ext-link", {
            "ext-link-type": "uri",
            f"{{{XLINK_NS}}}href": url_val,
        })
        ext.text = url_val


# --- Salida XML --------------------------------------------------------------
# Namespaces que SciELO SPS requiere declarados en el elemento raíz <article>
# pero que ElementTree solo incluye cuando los usa activamente.
_EXTRA_NS = {
    "xmlns:ali":  "http://www.niso.org/schemas/ali/1.0/",
    "xmlns:xsi":  "http://www.w3.org/2001/XMLSchema-instance",
    "xmlns:xlink": "http://www.w3.org/1999/xlink",
    "xmlns:mml":   "http://www.w3.org/1998/Math/MathML",
}

def _pretty_xml(root: ET.Element) -> str:
    """Serializa XML UTF-8 e inserta DOCTYPE JATS 1.1 para SciELO."""
    raw = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    try:
        parsed = minidom.parseString(raw)
        pretty = parsed.toprettyxml(indent="  ", encoding="utf-8").decode("utf-8")
        lines = [ln for ln in pretty.splitlines() if ln.strip()]
        xml = "\n".join(lines)
    except Exception:
        xml = raw.decode("utf-8", errors="replace")

    if SCIELO_DOCTYPE not in xml:
        decl_end = xml.find("?>")
        if decl_end != -1:
            xml = xml[:decl_end + 2] + "\n" + SCIELO_DOCTYPE + xml[decl_end + 2:]
        else:
            xml = SCIELO_DOCTYPE + "\n" + xml

    # Inyectar namespaces faltantes en el tag <article ...>
    # Solo se agregan si aún no están presentes (evita duplicados).
    art_start = xml.find("<article")
    art_end   = xml.find(">", art_start)
    if art_start != -1 and art_end != -1:
        tag_str = xml[art_start:art_end + 1]
        extra = ""
        for attr, uri in _EXTRA_NS.items():
            if attr not in tag_str:
                extra += f'\n  {attr}="{uri}"'
        if extra:
            # Insertar justo antes del cierre del tag de apertura
            close = "/>" if tag_str.endswith("/>") else ">"
            insert_at = art_end - (1 if close == "/>" else 0)
            xml = xml[:insert_at] + extra + xml[insert_at:]

    if not xml.endswith("\n"):
        xml += "\n"

    # Workaround: Python ElementTree abrevia el tag name como n
    # Reemplazar <n> -> <name> y </n> -> </name>
    _nop = str.replace
    xml = _nop(xml, '<n>', '<na' + 'me>')
    xml = _nop(xml, '<n ', '<na' + 'me ')
    xml = _nop(xml, '</' + 'n>', '</' + 'na' + 'me>')
    return xml



def build_jats_xml(
    *,
    bloques: list[dict],
    referencias_externas: list[str],
    autores_orcid: list[dict],
    afiliaciones_txt: str,
    figuras: list[dict],
    tablas: list[dict],
) -> str:
    """Genera un documento XML con perfil base compatible con SciELO SPS."""
    # Copia defensiva de bloques limpios
    bks = []
    for b in bloques or []:
        bks.append({
            "contenido": b.get("contenido", ""),
            "clasificacion": b.get("clasificacion", "Cuerpo"),
            "italic": bool(b.get("italic", False)),
        })

    # Extraer metadatos escalares
    title_es = next(
        (_clean_text(b["contenido"]) for b in bks
         if b["clasificacion"] == "Título principal" and _clean_text(b["contenido"])),
        "Articulo",
    )
    title_en = next(
        (_clean_text(b["contenido"]) for b in bks
         if b["clasificacion"] == "Título secundario" and _clean_text(b["contenido"])),
        "",
    )

    authors = _authors_from_manual(autores_orcid)
    if not authors:
        authors = _authors_from_pdf(bks)

    affs, emails = _parse_affiliations_txt(afiliaciones_txt)
    if not affs and not emails:
        affs, emails = _fallback_affiliations_from_blocks(bks)

    # Bloques de encabezado con metadatos editoriales (ISSN, volumen, etc.)
    front_notes_seed = [
        _clean_text(b["contenido"])
        for b in bks
        if b["clasificacion"] == "Encabezado sección"
        and _clean_text(b["contenido"])
        and (
            _META_HEADING_RE.search(_norm(b["contenido"]))
            or _norm(b["contenido"]) == "paleontologia mexicana"
        )
    ]

    doi = _extract_doi(bks)
    pub_year, fpage, lpage, pub_volume, pub_issue = _extract_year_and_pages(
        front_notes_seed, doi
    )

    # Recorrer bloques y clasificar
    abstracts: dict[str, list[str]] = {"es": [], "en": [], "plain_es": [], "plain_en": []}
    kwds: dict[str, list[str]] = {"es": [], "en": [], "general": []}
    refs_from_blocks: list[str] = []
    cite_notes: list[str] = []
    date_notes: list[str] = []
    front_notes: list[str] = []
    body_stream: list[dict] = []

    abs_mode: str | None = None
    in_refs = False
    body_started = False

    for b in bks:
        cls = b["clasificacion"]
        txt = _clean_text(b["contenido"])
        if not txt:
            continue

        # Encabezados de seccion
        if cls == "Encabezado sección":
            low = _norm(txt)

            if low in ("referencias", "references"):
                in_refs = True
                abs_mode = None
                continue

            if _META_HEADING_RE.search(low) or low == "paleontologia mexicana":
                front_notes.append(txt)
                continue

            if low == "resumen":
                abs_mode = "es"; in_refs = False; continue
            if low == "abstract":
                abs_mode = "en"; in_refs = False; continue
            if low == "resumen no tecnico":
                abs_mode = "plain_es"; in_refs = False; continue
            if low == "non-technical abstract":
                abs_mode = "plain_en"; in_refs = False; continue

            # Cualquier otro encabezado cierra abs_mode
            in_refs = False
            abs_mode = None
            if _is_body_heading(txt):
                body_started = True
                body_stream.append({"kind": "h", "text": txt})
            continue

        # Dentro de referencias
        if in_refs:
            if cls in ("Referencia", "Cuerpo", "Normal"):
                refs_from_blocks.append(txt)
            continue

        # Clases de front-matter
        if cls in ("Título principal", "Título secundario", "Autores",
                   "Filiación", "Email / Metadatos"):
            continue
        if cls in ("Ignorar", "Imagen", "Pie de figura", "Título tabla"):
            continue

        # Notas editoriales 
        if cls == "Cómo citar":
            cite_notes.append(txt); continue
        if cls == "Fecha manuscrito":
            date_notes.append(txt); continue

        # Palabras clave
        if cls == "Palabras clave":
            vals = _parse_keywords(txt)
            if abs_mode in ("es", "plain_es"):
                kwds["es"].extend(vals)
            elif abs_mode in ("en", "plain_en"):
                kwds["en"].extend(vals)
            else:
                kwds["general"].extend(vals)
            continue

        # Cuerpo del abstract (clase explícita del usuario)
        if cls == "Cuerpo del abstract":
            target = abs_mode or "es"
            abstracts[target].extend(_split_paragraphs(txt))
            continue

        # Resumen / Abstract
        if abs_mode and cls == "Resumen / Abstract":
            abstracts[abs_mode].extend(_split_paragraphs(txt))
            continue

        # ── Subencabezados
        if cls in ("Subencabezado", "Subencabezado-bajo"):
            abs_mode = None
            body_started = True
            body_stream.append({"kind": "h", "text": txt})
            continue

        # ── Cuerpo / Normal
        if cls in ("Cuerpo", "Normal", "Resumen / Abstract"):
            abs_mode = None
            body_started = True
            body_stream.append({"kind": "p", "text": txt})
            continue

    hist_dates = _parse_manuscript_dates(date_notes)

    # Periodo de publicación
    _pub_period = ""
    _period_re = re.compile(
        r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto"
        r"|septiembre|octubre|noviembre|diciembre"
        r"|january|february|march|april|may|june|july|august"
        r"|september|october|november|december)"
        r".{1,30}(\d{4})",
        re.IGNORECASE,
    )
    for fn in front_notes:
        m_p = _period_re.search(fn)
        if m_p:
            _pub_period = fn.strip()
            break

    # Construir el árbol XML en el orden correcto del DTD
    root = ET.Element(
        "article",
        {
            "article-type": "research-article",
            "dtd-version": SCIELO_DTD_VERSION,
            "specific-use": SCIELO_SPS_VERSION,
            f"{{{XML_NS}}}lang": "es",
        },
    )

    front = ET.SubElement(root, "front")

    # journal-meta
    journal_meta = ET.SubElement(front, "journal-meta")
    jid = ET.SubElement(journal_meta, "journal-id", {"journal-id-type": "publisher-id"})
    jid.text = "Paleontologia Mexicana"
    jtg = ET.SubElement(journal_meta, "journal-title-group")
    jt = ET.SubElement(jtg, "journal-title")
    jt.text = "Paleontología Mexicana"
    jab = ET.SubElement(jtg, "abbrev-journal-title", {"abbrev-type": "publisher"})
    jab.text = "Paleontol. Mex."
    issn_val = _extract_issn(bks)
    issn_node = ET.SubElement(journal_meta, "issn", {"pub-type": "epub"})
    issn_node.text = issn_val or "0000-0000"
    pub = ET.SubElement(journal_meta, "publisher")
    ET.SubElement(pub, "publisher-name").text = "Universidad Nacional Autónoma de México"

    # article-meta — se construye con SubElement en el orden exacto del DTD
    am = ET.SubElement(front, "article-meta")

    # 1. article-id
    if doi:
        ET.SubElement(am, "article-id", {"pub-id-type": "doi"}).text = doi

    # 2. article-categories
    ac = ET.SubElement(am, "article-categories")
    sg = ET.SubElement(ac, "subj-group", {"subj-group-type": "heading"})
    ET.SubElement(sg, "subject").text = "Research Article"

    # 3. title-group
    tg = ET.SubElement(am, "title-group")
    ET.SubElement(tg, "article-title").text = title_es
    if title_en:
        ttg = ET.SubElement(tg, "trans-title-group", {f"{{{XML_NS}}}lang": "en"})
        ET.SubElement(ttg, "trans-title").text = title_en

    # 4. contrib-group (solo si hay autores)
    if authors:
        cg = ET.SubElement(am, "contrib-group")
        for idx, a in enumerate(authors):
            contrib = ET.SubElement(cg, "contrib", {"contrib-type": "author"})
            if a.get("orcid"):
                cid = ET.SubElement(contrib, "contrib-id",
                                    {"contrib-id-type": "orcid", "authenticated": "false"})
                cid.text = f"https://orcid.org/{a['orcid']}"
            name_el = ET.SubElement(contrib, "name")
            surname_txt, given_txt = _split_person_name(a["nombre"])
            ET.SubElement(name_el, "surname").text = surname_txt or a["nombre"]
            if given_txt:
                ET.SubElement(name_el, "given-names").text = given_txt
            # xref a afiliaciones: solo 1 xref si hay exactamente 1 afiliación,
            # o una por afiliación si son varias (el usuario decide cuál corresponde)
            for af in affs:
                xr = ET.SubElement(contrib, "xref", {"ref-type": "aff", "rid": af["id"]})
                xr.text = af.get("label", "")
            if emails and idx == 0:
                ET.SubElement(contrib, "xref", {"ref-type": "corresp", "rid": "cor1"})

    # 5. aff (una por afiliación)
    for af in affs:
        aff_el = ET.SubElement(am, "aff", {"id": af["id"]})
        ET.SubElement(aff_el, "label").text = af["label"]
        ET.SubElement(aff_el, "institution",
                      {"content-type": "original"}).text = af["text"]
        country_txt, country_code = _detect_country(af["text"])
        cel = ET.SubElement(aff_el, "country")
        if country_code:
            cel.set("country", country_code)
        cel.text = country_txt or ""

    # 6. author-notes (solo si hay email de correspondencia)
    if emails:
        an = ET.SubElement(am, "author-notes")
        corresp = ET.SubElement(an, "corresp", {"id": "cor1"})
        ET.SubElement(corresp, "label").text = "*"
        for em in emails:
            ET.SubElement(corresp, "email").text = em

    # 7. pub-date
    pd = ET.SubElement(am, "pub-date", {"pub-type": "epub"})
    ET.SubElement(pd, "year").text = pub_year or "2026"

    # 8. volume, issue
    if pub_volume:
        ET.SubElement(am, "volume").text = pub_volume
    if pub_issue:
        ET.SubElement(am, "issue").text = pub_issue

    # 9. fpage / lpage / elocation-id
    if fpage:
        ET.SubElement(am, "fpage").text = fpage
    if lpage:
        ET.SubElement(am, "lpage").text = lpage
    elif doi and not fpage:
        ET.SubElement(am, "elocation-id").text = doi.rsplit(".", 1)[-1]

    # 10. history
    if hist_dates:
        hist = ET.SubElement(am, "history")
        for date_type, day, month, yr in hist_dates:
            de = ET.SubElement(hist, "date", {"date-type": date_type})
            if day:
                ET.SubElement(de, "day").text = day
            if month:
                ET.SubElement(de, "month").text = month
            ET.SubElement(de, "year").text = yr

    # 11. permissions
    perms = ET.SubElement(am, "permissions")
    ET.SubElement(perms, "copyright-statement").text = \
        f"© {pub_year or '2026'} Paleontología Mexicana"
    ET.SubElement(perms, "copyright-year").text = pub_year or "2026"
    ET.SubElement(perms, "copyright-holder").text = "Paleontología Mexicana"
    ET.SubElement(perms, f"{{{ALI_NS}}}free_to_read")
    lic = ET.SubElement(perms, "license", {
        "license-type": "open-access",
        f"{{{XLINK_NS}}}href": "https://creativecommons.org/licenses/by/4.0/",
        f"{{{XML_NS}}}lang": "es",
    })
    ET.SubElement(lic, "license-p").text = (
        "Distribuido bajo una licencia Creative Commons Attribution 4.0 "
        "International (CC BY 4.0)."
    )

    # 12. abstract (español)
    if abstracts["es"]:
        abs_es = ET.SubElement(am, "abstract")
        ET.SubElement(abs_es, "title").text = "Resumen"
        for ptxt in abstracts["es"]:
            ET.SubElement(abs_es, "p").text = ptxt

    if abstracts["plain_es"]:
        abs_pes = ET.SubElement(am, "abstract",
                                {"abstract-type": "plain-language-summary"})
        ET.SubElement(abs_pes, "title").text = "Resumen no técnico"
        for ptxt in abstracts["plain_es"]:
            ET.SubElement(abs_pes, "p").text = ptxt

    # 13. trans-abstract
    if abstracts["en"]:
        abs_en = ET.SubElement(am, "trans-abstract", {f"{{{XML_NS}}}lang": "en"})
        ET.SubElement(abs_en, "title").text = "Abstract"
        for ptxt in abstracts["en"]:
            ET.SubElement(abs_en, "p").text = ptxt

    if abstracts["plain_en"]:
        abs_pen = ET.SubElement(am, "trans-abstract", {
            "abstract-type": "plain-language-summary",
            f"{{{XML_NS}}}lang": "en",
        })
        ET.SubElement(abs_pen, "title").text = "Non-technical Abstract"
        for ptxt in abstracts["plain_en"]:
            ET.SubElement(abs_pen, "p").text = ptxt

    # 14. kwd-group
    _KWD_TITLES = {"es": "Palabras clave", "en": "Keywords"}

    def _append_kwd_group(lang: str, items: list[str]) -> None:
        seen: set[str] = set()
        uniques = []
        for k in items:
            kk = _clean_text(k).strip(" .;")
            if kk and kk.lower() not in seen:
                seen.add(kk.lower())
                uniques.append(kk)
        if not uniques:
            return
        kg = ET.SubElement(am, "kwd-group", {
            "kwd-group-type": "author-generated",
            f"{{{XML_NS}}}lang": lang,
        })
        ET.SubElement(kg, "title").text = _KWD_TITLES.get(lang, "Palabras clave")
        for k in uniques:
            ET.SubElement(kg, "kwd").text = k

    _append_kwd_group("es", kwds["es"])
    _append_kwd_group("en", kwds["en"])
    if kwds["general"]:
        _append_kwd_group("es", kwds["general"])

    # 15. custom-meta-group (solo publication-period y how-to-cite)
    _cmg_items: list[tuple[str, str]] = []
    if _pub_period:
        _cmg_items.append(("publication-period", _pub_period))
    for txt_c in cite_notes:
        _cmg_items.append(("how-to-cite", txt_c))

    if _cmg_items:
        cmg = ET.SubElement(am, "custom-meta-group")
        for meta_name, meta_val in _cmg_items:
            cm = ET.SubElement(cmg, "custom-meta")
            ET.SubElement(cm, "meta-name").text = meta_name
            ET.SubElement(cm, "meta-value").text = meta_val

    # body
    body = ET.SubElement(root, "body")
    has_sections = any(item["kind"] == "h" for item in body_stream)
    current_sec: ET.Element | None = None

    def _ensure_sec(title_fallback: str = "") -> ET.Element:
        """Devuelve el <sec> activo. Si no hay uno, crea uno CON <title>
        para evitar el error DTD 'Missing element title'."""
        nonlocal current_sec
        if current_sec is None:
            current_sec = ET.SubElement(body, "sec")
            t = ET.SubElement(current_sec, "title")
            t.text = title_fallback or ""
        return current_sec

    for item in body_stream:
        if item["kind"] == "h":
            current_sec = ET.SubElement(body, "sec")
            ET.SubElement(current_sec, "title").text = item["text"]
            continue
        for ptxt in _split_paragraphs(item["text"]):
            if has_sections:
                p = ET.SubElement(_ensure_sec(), "p")
            else:
                p = ET.SubElement(body, "p")
            p.text = ptxt

    # Tablas al final del body
    if tablas:
        sec_tables = ET.SubElement(body, "sec", {"sec-type": "tables"})
        ET.SubElement(sec_tables, "title").text = "Tablas"
        for i, tab in enumerate(tablas, 1):
            tw = ET.SubElement(sec_tables, "table-wrap", {"id": f"tbl{i}"})
            ET.SubElement(tw, "label").text = f"Tabla {i}"
            cap = ET.SubElement(tw, "caption")
            ET.SubElement(cap, "title").text = f"Tabla {i}"
            ET.SubElement(cap, "p").text = \
                _clean_text(tab.get("titulo", "")) or f"Tabla {i}"
            _append_table_xml(tw, tab.get("ruta", ""), tab.get("hoja"))

    # Figuras al final del body
    if figuras:
        sec_figs = ET.SubElement(body, "sec", {"sec-type": "figures"})
        ET.SubElement(sec_figs, "title").text = "Figuras"
        for i, fig in enumerate(figuras, 1):
            fg = ET.SubElement(sec_figs, "fig", {"id": f"fig{i}"})
            ET.SubElement(fg, "label").text = f"Figura {i}"
            cap = ET.SubElement(fg, "caption")
            ET.SubElement(cap, "title").text = f"Figura {i}"
            ET.SubElement(cap, "p").text = \
                _clean_text(fig.get("pie", "")) or f"Figura {i}"
            g = ET.SubElement(fg, "graphic")
            g.set(f"{{{XLINK_NS}}}href",
                  os.path.basename(fig.get("ruta", f"fig{i}.png")))

    # Back
    back = ET.SubElement(root, "back")
    refs = referencias_externas if referencias_externas else refs_from_blocks
    refs = [_strip_ref_prefix(r) for r in refs if _strip_ref_prefix(r)]
    if refs:
        ref_list = ET.SubElement(back, "ref-list")
        ET.SubElement(ref_list, "title").text = "Referencias"
        for i, ref in enumerate(refs, 1):
            r_el = ET.SubElement(ref_list, "ref", {"id": f"R{i}"})
            ET.SubElement(r_el, "mixed-citation").text = ref
            _build_element_citation(r_el, ref)

    return _pretty_xml(root)
