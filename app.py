import customtkinter as ctk
from tkinter import filedialog
from PIL import Image as PILImage
import fitz
import re
import os
import tempfile
import unicodedata
import ebooklib
from ebooklib import epub
from collections import Counter
import base64
import shutil

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# ─── CSS exportado ────────────────────────────────────────────────────────────
HTML_CSS = """
<style>
  @import url('https://fonts.googleapis.com/css2?family=Source+Serif+4:ital,wght@0,400;0,700;1,400;1,700&display=swap');
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --serif: 'Source Serif 4', 'Georgia', serif;
    --sans:  'Helvetica Neue', Arial, sans-serif;
    --tnr:   'Times New Roman', Times, serif;
    --texto: #1a1a1a; --gris: #555; --linea: #bbb;
  }
  body {
    font-family: var(--serif); font-size: 12pt; line-height: 1.65;
    color: var(--texto); background: #fff;
    max-width: 880px; margin: 0 auto; padding: 32px 52px 64px; hyphens: auto;
  }
  h1.titulo-principal  { font-family:var(--serif); font-size:14pt; font-weight:700; font-style:normal;  line-height:1.3; margin:20px 0 6px; }
  h2.titulo-secundario { font-family:var(--serif); font-size:14pt; font-weight:700; font-style:italic;  line-height:1.3; color:#222; margin-bottom:14px; }
  .autores     { font-family:var(--serif); font-size:13pt; font-weight:400; font-style:normal; margin:12px 0 4px; line-height:1.6; }
  .autores a.orcid-autor { color:#1a3a5c; text-decoration:underline; text-underline-offset:3px; text-decoration-color:#A6CE39; text-decoration-thickness:2px; }
  .autores a.orcid-autor:hover { text-decoration-color:#1a3a5c; }
  .autores .orcid-icon img { width:16px; height:16px; vertical-align:middle; margin-left:2px; }
  .filiaciones { font-family:var(--serif); font-size:9pt; font-weight:400; font-style:normal; color:#666; margin:2px 0; line-height:1.5; }
  .email       { font-family:var(--serif); font-size:9pt;  font-style:italic; color:#1a5276; margin-bottom:6px; }
  h2.seccion { font-family:var(--serif); font-size:13pt; font-weight:700; text-align:center; margin:24px 0 10px; }
  h2.seccion.meta { font-size:9pt; font-weight:400; margin:3px 0; }
  h2.seccion.con-linea { border-top:1px solid var(--linea); padding-top:18px; margin-top:32px; }
  /* Solo Abstract y Non-technical Abstract en gris */
  h2.seccion.gris { color:#666; }
  h3.subseccion { font-family:var(--serif); font-size:12pt; font-weight:700; margin:20px 0 8px; }
  h3.subseccion.primer-nivel1 { border-top:1px solid var(--linea); padding-top:18px; margin-top:32px; }
  h3.subseccion-bajo{ font-family:var(--serif); font-size:12pt; font-weight:700; font-style:italic; margin:16px 0 6px; }
  p.resumen  { font-family:var(--serif); font-size:9pt; text-align:justify; text-indent:1.2em; margin-bottom:7px; }
  p.abstract { font-family:var(--tnr);   font-size:12pt; color:#666; font-style:italic; text-align:justify; text-indent:1.2em; margin-bottom:7px; }
  p { font-family:var(--serif); font-size:10pt; text-align:justify; text-indent:1.4em; margin-bottom:8px; font-style:normal; }
  p.sin-sangria { text-indent:0; }
  p.cuerpo { font-family:var(--serif); font-size:12pt; font-style:normal !important; font-weight:normal; text-align:justify; text-indent:1.4em; margin-bottom:9px; color:var(--texto); }
  .keywords { font-family:var(--serif); font-size:9pt; margin:4px 0 16px; text-indent:0; }
  .keywords strong { font-weight:700; }
  ol.referencias { padding-left:2em; margin:8px 0 16px; }
  ol.referencias li { font-family:var(--serif); font-size:10pt; margin-bottom:6px; line-height:1.5; }
  .post-referencias { margin-top:28px; border-top:1px solid var(--linea); padding-top:14px; }
  .como-citar { font-family:var(--serif); font-size:9pt; margin-bottom:10px; line-height:1.5; }
  .fechas-manuscrito ul { list-style:disc; padding-left:1.6em; margin:6px 0 10px; }
  .fechas-manuscrito li { font-family:var(--serif); font-size:9pt; margin-bottom:4px; }
  .fechas-manuscrito a { color:#1a5276; }
  .figuras-finales { margin-top:28px; border-top:1px solid var(--linea); padding-top:14px; }
  .figuras-finales h2 { font-family:var(--serif); font-size:10pt; font-weight:700; text-align:center; margin-bottom:14px; text-transform:uppercase; letter-spacing:0.05em; }
  figure { margin:18px auto; text-align:center; }
  figure img { max-width:100%; border:1px solid var(--linea); }
  figcaption { font-family:var(--serif); font-size:9pt; color:#1a1a1a; margin-top:5px; text-align:left; line-height:1.4; }
  /* ── Tablas ── */
  .tabla-wrapper { margin:20px auto 24px; overflow-x:auto; }
  .tabla-titulo { font-family:var(--serif); font-size:9pt; font-weight:400; color:#666; margin-bottom:6px; }
  .tabla-titulo strong { font-weight:700; }
  table.pm-tabla { border-collapse:collapse; width:100%; font-family:var(--serif); font-size:9pt; }
  table.pm-tabla thead tr th { background:#1b5e9a; color:#fff; font-weight:700; padding:6px 10px; border:1px solid #155080; text-align:center; }
  table.pm-tabla tbody tr td { background:#cbeefb; color:#1a1a1a; padding:4px 10px; border:1px solid #9fd8f0; vertical-align:top; }
  table.pm-tabla tbody tr:nth-child(even) td { background:#b8e6f8; }
</style>
"""

# ─── Clasificaciones ──────────────────────────────────────────────────────────
OPCIONES = [
    "Cuerpo",
    "Título principal", "Título secundario",
    "Encabezado sección",
    "Subencabezado", "Subencabezado-bajo",
    "Palabras clave",
    "Referencia",
    "Cómo citar", "Fecha manuscrito",
    "Título tabla", "Pie de figura",
    "Filiación", "Email / Metadatos",
    "Ignorar",
]

# Mapeo de clases antiguas que ya no existen → clase equivalente
_CLASE_COMPAT = {
    "Normal":            "Cuerpo",
    "Autores":           "Cuerpo",        # autores vienen de la pestaña ORCID
    "Resumen / Abstract":"Cuerpo",        # el estilo lo da el Encabezado sección
    "Imagen":            "Ignorar",
}

COLORES_UI = [
    ("Título principal",   "#1a237e", "Azul marino"),
    ("Título secundario",  "#283593", "Azul índigo"),
    ("Encabezado sección", "#0277bd", "Azul"),
    ("Subencabezado",      "#00695c", "Verde azulado"),
    ("Subencabezado-bajo", "#00796b", "Verde azulado claro"),
    ("Cuerpo",             "#212121", "Negro"),
    ("Palabras clave",     "#6a1b9a", "Morado"),
    ("Referencia",         "#424242", "Gris"),
    ("Cómo citar",         "#e65100", "Naranja"),
    ("Fecha manuscrito",   "#bf360c", "Rojo ladrillo"),
    ("Título tabla",       "#1565c0", "Azul tabla"),
    ("Pie de figura",      "#558b2f", "Verde oliva"),
    ("Filiación",          "#2e7d32", "Verde"),
    ("Email / Metadatos",  "#388e3c", "Verde medio"),
    ("Ignorar",            "#c62828", "Rojo"),
]
COLOR_POR_CLASE = {c: col for c, col, _ in COLORES_UI}

# ─── ORCID SVG embebido ───────────────────────────────────────────────────────
_ORCID_SVG = (
    "data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmci"
    "IHZpZXdCb3g9IjAgMCAyNCAyNCI+PGNpcmNsZSBjeD0iMTIiIGN5PSIxMiIgcj0iMTIiIGZpbGw9"
    "IiNBNkNFMzkiLz48cGF0aCBkPSJNNy41IDVoMXY3LjVoLTF6TTkuMyA3LjhDOS4zIDYgMTAuNCA1"
    "IDEyLjEgNWMxLjggMCAyLjkgMSAyLjkgMi44VjEyaDEuMlY3LjhjMC0yLjQtMS40LTMuOC0zLjct"
    "My44QzkuOCA0IDguMSA1IDguMSA3LjhWMTJoMS4yVjcuOHoiIGZpbGw9IiNmZmYiLz48L3N2Zz4="
)


# ─── Utilidades ───────────────────────────────────────────────────────────────

def esc(t: str) -> str:
    return t.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")


def _insertar_orcid(texto: str, autores_orcid: list | None = None) -> str:
    """
    Construye el HTML de autores con links ORCID.
    Si autores_orcid está disponible, lo usa directamente (nombres + IDs exactos).
    Si no, fallback a búsqueda por nombre en orcid.org.
    """
    _ORCID_BASE = "https://orcid.org/"
    _ORCID_SEARCH = "https://orcid.org/orcid-search/search?searchQuery="

    if autores_orcid:
        # Usar la lista manual — nombres y ORCIDs exactos
        partes = []
        for a in autores_orcid:
            nombre = a["nombre"].strip()
            orcid  = a["orcid"].strip()
            if not nombre: continue
            nombre_esc = esc(nombre)
            if orcid:
                url = _ORCID_BASE + orcid
                ico = (f'<a class="orcid-icon" href="{url}" target="_blank" '
                       f'title="ORCID: {orcid}">'
                       f'<img src="{_ORCID_SVG}" alt="ORCID"></a>')
                partes.append(
                    f'<a class="orcid-autor" href="{url}" target="_blank" '
                    f'title="ORCID: {orcid}">{nombre_esc}</a>{ico}')
            else:
                # Sin ORCID: mostrar nombre sin link pero con búsqueda
                q   = re.sub(r"\s+", "+", nombre)
                url = _ORCID_SEARCH + q
                partes.append(
                    f'<a class="orcid-autor" href="{url}" target="_blank" '
                    f'title="Buscar en ORCID">{nombre_esc}</a>')
        return "; ".join(partes)

    # Fallback: parsear el texto crudo del PDF
    partes = [p.strip() for p in texto.split(";") if p.strip()]
    result = []
    for parte in partes:
        nombre_raw = re.sub(r"[\d,\*\u00b9\u00b2\u00b3\u2070-\u209f]+$", "", parte).strip()
        if not nombre_raw: continue
        nombre_esc = esc(nombre_raw)
        q   = re.sub(r"\s+", "+", nombre_raw)
        url = _ORCID_SEARCH + q
        ico = (f'<a class="orcid-icon" href="{url}" target="_blank" title="Buscar en ORCID">'
               f'<img src="{_ORCID_SVG}" alt="ORCID"></a>')
        result.append(
            f'<a class="orcid-autor" href="{url}" target="_blank" '
            f'title="Buscar {nombre_esc} en ORCID">{nombre_esc}</a>{ico}')
    return "; ".join(result)


def _parsear_referencias(texto: str) -> list[str]:
    patron = re.compile(r'^\s*(?:\[?\d+[\.\)\]]\s*)', re.MULTILINE)
    partes = patron.split(texto)
    refs   = [p.strip() for p in partes if p.strip()]
    if refs:
        return refs
    return [l.strip() for l in texto.splitlines() if l.strip()]


def _es_como_citar(t: str) -> bool:
    return bool(re.match(r"(cómo citar|how to cite)", t.strip().lower()))


def _es_encabezado_resumen_legacy_no_usar(t: str) -> bool:
    """Coincidencia estricta para encabezados de resumen.
    Evita falsos positivos como "Abstracto"."""
    return bool(re.match(
        r"^\s*(resumen|abstract|resumen no t[Ãe]cnico|non-technical abstract)\s*[:\.]?\s*$",
        t, re.IGNORECASE
    ))


def _es_encabezado_resumen(t: str) -> bool:
    """Coincidencia estricta para encabezados de resumen.
    Evita falsos positivos como "Abstracto"."""
    norm = unicodedata.normalize("NFKD", t)
    norm = "".join(ch for ch in norm if not unicodedata.combining(ch))
    norm = re.sub(r"\s+", " ", norm).strip().lower()
    return bool(re.match(
        r"^(resumen|abstract|resumen no tecnico|non-technical abstract)\s*[:\.]?$",
        norm
    ))


def _es_fecha_mss(t: str) -> bool:
    return bool(re.search(
        r"(manuscrito\s+recibido|manuscrito\s+corregido|manuscrito\s+aceptado"
        r"|manuscript\s+received|manuscript\s+revised|manuscript\s+accepted)",
        t, re.IGNORECASE))


def _es_doi(t: str) -> bool:
    """Solo marca como fecha/DOI si es un bloque pequeño de metadatos,
    NO si es una referencia bibliográfica (bloques largos con autores y títulos)."""
    if not re.search(r"https?://doi\.org/", t):
        return False
    # Referencia bibliográfica: tiene patrón Apellido, Inicial. o más de 120 chars
    # sin palabras clave de fecha → NO es fecha manuscrito
    if len(t) > 200:
        return False
    if re.search(r"\.\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]", t) and not _es_fecha_mss(t):
        return False
    return True


def _img_to_base64(path: str) -> str:
    """Devuelve data-URI base64 de la imagen."""
    ext  = os.path.splitext(path)[1].lower().lstrip(".")
    mime = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png",
            "gif": "gif",  "webp": "webp"}.get(ext, "jpeg")
    with open(path, "rb") as f:
        data = base64.b64encode(f.read()).decode()
    return f"data:image/{mime};base64,{data}"


def _limpiar_prefijo_pie_figura(texto: str) -> str:
    """Quita 'Figura 1.'/'Fig. 2:' del inicio del pie."""
    t = re.sub(r"\s+", " ", texto).strip()
    t = re.sub(
        r"^\s*(figura|fig\.?|figure)\s*\d+[a-z]?(?:\s*-\s*[a-z])?[\.\):]?\s*",
        "",
        t,
        flags=re.IGNORECASE,
    )
    return t.strip(" -:\t")


def _limpiar_prefijo_titulo_tabla(texto: str) -> str:
    """Quita 'Tabla 1.'/'Table 2:' del inicio del titulo."""
    t = re.sub(r"\s+", " ", texto).strip()
    t = re.sub(
        r"^\s*(tabla|table)\s*\d+[a-z]?(?:\s*-\s*[a-z])?[\.\):]?\s*",
        "",
        t,
        flags=re.IGNORECASE,
    )
    return t.strip(" -:\t")


def _excel_a_html_tabla(ruta: str, hoja: str = None) -> str:
    """
    Convierte una hoja de un .xlsx en tabla HTML con estilos PM.
    Si hoja=None usa la hoja activa (primera).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            return "<p><em>[Tabla vacía]</em></p>"

        html = ['<table class="pm-tabla">']
        html.append("<thead><tr>")
        for celda in filas[0]:
            val = "" if celda is None else str(celda)
            html.append(f"<th>{esc(val)}</th>")
        html.append("</tr></thead><tbody>")
        for fila in filas[1:]:
            if all(c is None for c in fila):
                continue
            html.append("<tr>")
            for celda in fila:
                val = "" if celda is None else str(celda)
                html.append(f"<td>{esc(val)}</td>")
            html.append("</tr>")
        html.append("</tbody></table>")
        return "\n".join(html)
    except ImportError:
        return "<p><em>[Instala openpyxl: pip install openpyxl]</em></p>"
    except Exception as e:
        return f"<p><em>[Error al leer tabla: {esc(str(e))}]</em></p>"


# ─────────────────────────────────────────────────────────────────────────────
class LimpiadorEditorialApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Paleontología Mexicana — Editor Semántico")
        self.geometry("1280x940")
        self.minsize(1024, 700)
        self.configure(fg_color="#0f1117")

        self.datos_bloques:        list[dict] = []
        self.referencias_externas: list[str]  = []
        self.figuras_manuales:     list[dict] = []
        self.tablas_manuales:      list[dict] = []
        self.autores_orcid:        list[dict] = []
        self.afiliaciones_txt:     str        = ""
        self._vista_estructura:    bool       = True
        self._auto_fig_dir:        str | None = None
        self._auto_tab_dir:        str | None = None

        # ══════════════════════════════════════════════════════════
        # BARRA SUPERIOR — logo + título + exportar
        # ══════════════════════════════════════════════════════════
        topbar = ctk.CTkFrame(self, fg_color="#16213e", corner_radius=0, height=64)
        topbar.pack(fill="x", side="top")
        topbar.pack_propagate(False)

        # Logo / nombre
        brand = ctk.CTkFrame(topbar, fg_color="transparent")
        brand.pack(side="left", padx=20, pady=10)

        ctk.CTkLabel(brand,
                     text="◈",
                     font=ctk.CTkFont(size=28, weight="bold"),
                     text_color="#3b82f6").pack(side="left", padx=(0, 8))

        title_block = ctk.CTkFrame(brand, fg_color="transparent")
        title_block.pack(side="left")
        ctk.CTkLabel(title_block,
                     text="Editor Semántico",
                     font=ctk.CTkFont(size=17, weight="bold"),
                     text_color="#f1f5f9").pack(anchor="w")
        ctk.CTkLabel(title_block,
                     text="Paleontología Mexicana · UNAM",
                     font=ctk.CTkFont(size=15),
                     text_color="#64748b").pack(anchor="w")

        # Botones exportar — derecha de la topbar
        export_bar = ctk.CTkFrame(topbar, fg_color="transparent")
        export_bar.pack(side="right", padx=20, pady=12)

        ctk.CTkLabel(export_bar,
                     text="EXPORTAR",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="#64748b").pack(side="left", padx=(0, 10))

        for texto, cmd, fg, hv, w in [
            ("⬡  HTML",  self.evento_exportar_html,  "#16a34a", "#15803d", 110),
            ("◻  XML",   self.evento_exportar_xml,   "#334155", "#475569", 90),
            ("◻  EPUB",  self.evento_exportar_epub,  "#334155", "#475569", 90),
        ]:
            ctk.CTkButton(export_bar, text=texto, command=cmd,
                          fg_color=fg, hover_color=hv,
                          width=w, height=34,
                          corner_radius=6,
                          font=ctk.CTkFont(size=15, weight="bold")
                          ).pack(side="left", padx=3)

        # ══════════════════════════════════════════════════════════
        # CONTENIDO PRINCIPAL — Pestañas
        # ══════════════════════════════════════════════════════════
        main = ctk.CTkFrame(self, fg_color="#0f1117", corner_radius=0)
        main.pack(fill="both", expand=True, padx=0, pady=0)

        self.tabs = ctk.CTkTabview(main,
                                   fg_color="#161b27",
                                   segmented_button_fg_color="#1e2535",
                                   segmented_button_selected_color="#3b82f6",
                                   segmented_button_selected_hover_color="#2563eb",
                                   segmented_button_unselected_color="#1e2535",
                                   segmented_button_unselected_hover_color="#2a3347",
                                   text_color="#94a3b8",
                                   text_color_disabled="#475569",
                                   corner_radius=10)
        self.tabs.pack(fill="both", expand=True, padx=14, pady=(8, 0))

        for nombre in ["📄  PDF", "👥  Autores", "🏛️  Afiliaciones",
                        "📋  Referencias", "🖼️  Figuras"]:
            self.tabs.add(nombre)

        self._construir_tab_pdf()
        self._construir_tab_autores()
        self._construir_tab_afiliaciones()
        self._construir_tab_referencias()
        self._construir_tab_figuras()

        # ══════════════════════════════════════════════════════════
        # STATUS BAR
        # ══════════════════════════════════════════════════════════
        statusbar = ctk.CTkFrame(self, fg_color="#0d1117", corner_radius=0, height=34)
        statusbar.pack(fill="x", side="bottom")
        statusbar.pack_propagate(False)

        self._status = ctk.CTkLabel(
            statusbar, text="Listo  ·  Carga un PDF para comenzar",
            anchor="w", font=ctk.CTkFont(size=15), text_color="#475569")
        self._status.pack(side="left", padx=14, pady=4)

        ctk.CTkLabel(statusbar,
                     text="v2.0",
                     font=ctk.CTkFont(size=15),
                     text_color="#1e293b").pack(side="right", padx=12)

    # ═════════════════════════════════════════════════════════════
    # TAB 1 — PDF
    # ═════════════════════════════════════════════════════════════

    def _construir_tab_pdf(self):
        tab = self.tabs.tab("📄  PDF")

        # ── Toolbar ───────────────────────────────────────────────
        toolbar = ctk.CTkFrame(tab, fg_color="#1e2535", corner_radius=8, height=48)
        toolbar.pack(fill="x", padx=4, pady=(6, 4))
        toolbar.pack_propagate(False)

        self._btn_cargar = ctk.CTkButton(
            toolbar, text="📂  Cargar PDF",
            command=self.evento_cargar_archivo,
            fg_color="#3b82f6", hover_color="#2563eb",
            width=150, height=34, corner_radius=6,
            font=ctk.CTkFont(size=15, weight="bold"))
        self._btn_cargar.pack(side="left", padx=(8, 4), pady=7)

        ctk.CTkButton(toolbar, text="Leyenda",
                      command=self._toggle_leyenda,
                      fg_color="#334155", hover_color="#475569",
                      width=80, height=34, corner_radius=6,
                      font=ctk.CTkFont(size=15)
                      ).pack(side="left", padx=4, pady=7)

        # Separador visual
        ctk.CTkFrame(toolbar, width=1, height=34, fg_color="#334155").pack(
            side="left", padx=8, pady=10)

        ctk.CTkLabel(toolbar, text="Filtrar:",
                     font=ctk.CTkFont(size=15),
                     text_color="#94a3b8").pack(side="left", padx=(0, 4), pady=7)
        self._filtro_menu = ctk.CTkOptionMenu(
            toolbar, values=["Todos"] + OPCIONES,
            command=self._aplicar_filtro,
            fg_color="#334155", button_color="#3b82f6",
            button_hover_color="#2563eb",
            width=190, height=34, corner_radius=6,
            font=ctk.CTkFont(size=15))
        self._filtro_menu.set("Todos")
        self._filtro_menu.pack(side="left", padx=4, pady=7)

        self._stats_lbl = ctk.CTkLabel(
            toolbar, text="", font=ctk.CTkFont(size=15), text_color="#64748b")
        self._stats_lbl.pack(side="left", padx=12)

        # ── Leyenda (oculta) ──────────────────────────────────────
        self._leyenda_visible = False
        self._leyenda_panel   = ctk.CTkFrame(tab, fg_color="#1e2535", corner_radius=8)
        self._construir_leyenda(tab)

        # ── Banner de completado ──────────────────────────────────
        self._banner = ctk.CTkLabel(
            tab, text="",
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color="#f0fdf4", fg_color="#166534",
            corner_radius=6, height=36, anchor="center")

        # ── Scroll de bloques ─────────────────────────────────────
        self.frame_scroll = ctk.CTkScrollableFrame(
            tab, fg_color="#161b27", label_text="",
            scrollbar_button_color="#334155",
            scrollbar_button_hover_color="#475569")
        self.frame_scroll.pack(fill="both", expand=True, padx=4, pady=(0, 4))

    # ═════════════════════════════════════════════════════════════
    # TAB 2 — AUTORES / ORCID
    # ═════════════════════════════════════════════════════════════

    def _construir_tab_autores(self):
        tab = self.tabs.tab("👥  Autores")

        ctk.CTkLabel(
            tab,
            text="Agrega autores uno a uno  ó  carga un Excel con columnas: Autor | ORCID\n"
                 "El ORCID puede ser el link completo (https://orcid.org/0000-...) o solo los números.",
            font=ctk.CTkFont(size=15), justify="left", text_color="#aaa"
        ).pack(anchor="w", padx=14, pady=(10, 4))

        bf = ctk.CTkFrame(tab, fg_color="transparent")
        bf.pack(fill="x", padx=10, pady=(0, 4))
        ctk.CTkButton(bf, text="➕  Agregar autor",
                      command=self._agregar_autor,
                      fg_color="#1565c0", hover_color="#0d47a1",
                      width=170, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        ctk.CTkButton(bf, text="📂  Cargar Excel",
                      command=self._cargar_autores_excel,
                      fg_color="#1b5e20", hover_color="#2e7d32",
                      width=150, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        ctk.CTkButton(bf, text="🗑  Limpiar todo",
                      command=self._limpiar_autores,
                      fg_color="#c62828", hover_color="#8b0000",
                      width=130, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        self._autores_lbl = ctk.CTkLabel(
            bf, text="Sin autores cargados",
            font=ctk.CTkFont(size=15), text_color="#888")
        self._autores_lbl.pack(side="left", padx=10)

        # Encabezado columnas
        hdr = ctk.CTkFrame(tab, fg_color="#1a2a3a", corner_radius=4)
        hdr.pack(fill="x", padx=10, pady=(2, 0))
        hdr.columnconfigure(1, weight=2)
        hdr.columnconfigure(2, weight=1)
        ctk.CTkLabel(hdr, text="#", width=36,
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="#aaa").grid(row=0, column=0, padx=6, pady=4)
        ctk.CTkLabel(hdr, text="Apellido, Nombre",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="#aaa", anchor="w").grid(row=0, column=1, padx=4, pady=4, sticky="ew")
        ctk.CTkLabel(hdr, text="ORCID  (solo números: 0000-0001-2345-6789)",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="#aaa", anchor="w").grid(row=0, column=2, padx=4, pady=4, sticky="ew")
        ctk.CTkLabel(hdr, text="", width=36,
                     font=ctk.CTkFont(size=15)).grid(row=0, column=3, padx=6)

        self._autores_scroll = ctk.CTkScrollableFrame(tab)
        self._autores_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 6))

    def _agregar_autor(self):
        self._sync_autores()
        self.autores_orcid.append({"nombre": "", "orcid": ""})
        self._refrescar_lista_autores()

    def _sync_autores(self):
        for a in self.autores_orcid:
            if "_var_nom" in a:
                a["nombre"] = a["_var_nom"].get().strip()
            if "_var_orc" in a:
                raw = a["_var_orc"].get().strip()
                # Si pegaron el link completo, extraer solo los números
                m = re.search(r"(\d{4}-\d{4}-\d{4}-\d{3}[\dX])", raw, re.IGNORECASE)
                a["orcid"] = m.group(1) if m else raw

    def _refrescar_lista_autores(self):
        for w in self._autores_scroll.winfo_children():
            w.destroy()
        for i, autor in enumerate(self.autores_orcid):
            row = ctk.CTkFrame(self._autores_scroll, fg_color="#1e2a3a", corner_radius=4)
            row.pack(fill="x", padx=2, pady=2)
            row.columnconfigure(1, weight=2)
            row.columnconfigure(2, weight=1)
            ctk.CTkLabel(row, text=f"{i+1}", width=36,
                         font=ctk.CTkFont(size=15, weight="bold"),
                         text_color="#7986cb").grid(row=0, column=0, padx=6, pady=6)
            var_nom = ctk.StringVar(value=autor.get("nombre", ""))
            ctk.CTkEntry(row, textvariable=var_nom,
                         placeholder_text="Apellido, Nombre",
                         font=ctk.CTkFont(size=15), height=38).grid(
                row=0, column=1, padx=4, pady=6, sticky="ew")
            autor["_var_nom"] = var_nom
            var_orc = ctk.StringVar(value=autor.get("orcid", ""))
            ctk.CTkEntry(row, textvariable=var_orc,
                         placeholder_text="0000-0001-2345-6789",
                         font=ctk.CTkFont(size=15), height=38).grid(
                row=0, column=2, padx=4, pady=6, sticky="ew")
            autor["_var_orc"] = var_orc
            def _borrar(idx=i):
                self._sync_autores()
                self.autores_orcid.pop(idx)
                self._refrescar_lista_autores()
            ctk.CTkButton(row, text="✕", width=28, height=34,
                          fg_color="#c62828", hover_color="#8b0000",
                          command=_borrar,
                          font=ctk.CTkFont(size=15)).grid(row=0, column=3, padx=(0,6), pady=6)
        n = len(self.autores_orcid)
        self._autores_lbl.configure(
            text=f"{n} autor{'es' if n!=1 else ''}" if n else "Sin autores cargados",
            text_color="#90caf9" if n else "#888")

    def _aplicar_autores(self):
        self._sync_autores()

    # ═════════════════════════════════════════════════════════════
    # TAB 3 — AFILIACIONES
    # ═════════════════════════════════════════════════════════════

    def _construir_tab_afiliaciones(self):
        tab = self.tabs.tab("🏛️  Afiliaciones")

        ctk.CTkLabel(
            tab,
            text=(
                "Carga un .txt con las afiliaciones, una por línea:\n"
                "    1 Colección de Paleontología, Facultad...\n"
                "    2 Departamento de Paleontología, Instituto...\n"
                "    * correo@unam.mx\n"
                "Los correos se vinculan automáticamente."
            ),
            font=ctk.CTkFont(size=15), justify="left", text_color="#aaa"
        ).pack(anchor="w", padx=14, pady=(12, 6))

        bf = ctk.CTkFrame(tab, fg_color="transparent")
        bf.pack(fill="x", padx=10, pady=(0, 6))

        ctk.CTkButton(bf, text="📂  Cargar .txt",
                      command=self._cargar_afiliaciones,
                      fg_color="#1b5e20", hover_color="#2e7d32",
                      width=170, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        ctk.CTkButton(bf, text="🗑  Limpiar",
                      command=self._limpiar_afiliaciones,
                      fg_color="#c62828", hover_color="#8b0000",
                      width=110, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        self._afil_lbl = ctk.CTkLabel(
            bf, text="Sin afiliaciones cargadas",
            font=ctk.CTkFont(size=15), text_color="#888")
        self._afil_lbl.pack(side="left", padx=10)

        self._afil_scroll = ctk.CTkScrollableFrame(tab, label_text="Afiliaciones cargadas")
        self._afil_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 6))

    def _cargar_afiliaciones(self):
        ruta = filedialog.askopenfilename(
            title="Selecciona .txt de afiliaciones",
            filetypes=[("Texto", "*.txt"), ("Todos", "*.*")])
        if not ruta: return
        with open(ruta, encoding="utf-8", errors="replace") as f:
            self.afiliaciones_txt = f.read()
        self._refrescar_afiliaciones()

    def _limpiar_afiliaciones(self):
        self.afiliaciones_txt = ""
        for w in self._afil_scroll.winfo_children():
            w.destroy()
        self._afil_lbl.configure(text="Sin afiliaciones cargadas", text_color="#888")

    def _refrescar_afiliaciones(self):
        for w in self._afil_scroll.winfo_children():
            w.destroy()
        lineas = [l for l in self.afiliaciones_txt.splitlines() if l.strip()]
        for linea in lineas:
            frame = ctk.CTkFrame(self._afil_scroll, fg_color="#1a2a1a", corner_radius=4)
            frame.pack(fill="x", padx=2, pady=2)
            ctk.CTkLabel(frame, text=linea, font=ctk.CTkFont(size=15),
                         justify="left", anchor="w", wraplength=700).pack(
                padx=10, pady=5, fill="x")
        n = len(lineas)
        self._afil_lbl.configure(
            text=f"{n} afiliación{'es' if n != 1 else ''}" if n else "Sin afiliaciones",
            text_color="#a5d6a7" if n else "#888")

    def _limpiar_autores(self):
        self.autores_orcid = []
        self._refrescar_lista_autores()

    def _cargar_autores_excel(self):
        """Carga autores y ORCIDs desde un Excel con columnas Autor | ORCID."""
        ruta = filedialog.askopenfilename(
            title="Selecciona Excel de autores",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if not ruta: return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta, data_only=True)
            ws = wb.active
            filas = list(ws.iter_rows(values_only=True))
            wb.close()
        except ImportError:
            self._set_status("❌ Instala openpyxl: pip install openpyxl"); return
        except Exception as e:
            self._set_status(f"❌ Error leyendo Excel: {e}"); return

        self._sync_autores()
        nuevos = 0
        for fila in filas:
            if not fila or all(c is None for c in fila):
                continue
            nombre = str(fila[0]).strip() if fila[0] else ""
            orcid  = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""

            # Saltar fila de encabezado
            if nombre.lower() in ("autor", "nombre", "author", "name"):
                continue
            if not nombre:
                continue

            # Extraer solo los números del ORCID si viene como URL completa
            m = re.search(r"(\d{4}-\d{4}-\d{4}-\d{3}[\dX])", orcid)
            orcid_limpio = m.group(1) if m else orcid

            self.autores_orcid.append({"nombre": nombre, "orcid": orcid_limpio})
            nuevos += 1

        self._refrescar_lista_autores()
        self._set_status(f"✓ {nuevos} autor(es) importados desde Excel.")

    # ═════════════════════════════════════════════════════════════
    # TAB 3 — REFERENCIAS
    # ═════════════════════════════════════════════════════════════

    def _construir_tab_referencias(self):
        tab = self.tabs.tab("📋  Referencias")

        info = ctk.CTkLabel(
            tab,
            text=(
                "Carga un .txt con las referencias numeradas.\n"
                "Formatos aceptados:   1. Texto...   |   1) Texto...   |   [1] Texto...\n"
            ),
            font=ctk.CTkFont(size=15), justify="left", text_color="#aaa")
        info.pack(anchor="w", padx=12, pady=(12, 6))

        btn_f = ctk.CTkFrame(tab, fg_color="transparent")
        btn_f.pack(fill="x", padx=10, pady=(0, 8))

        ctk.CTkButton(btn_f, text="📂  Cargar .txt de referencias",
                      command=self.evento_cargar_referencias,
                      fg_color="#c62828", hover_color="#8b0000",
                      width=230, font=ctk.CTkFont(size=15)
                      ).pack(side="left", padx=5)

        ctk.CTkButton(btn_f, text="🗑  Limpiar",
                      command=self._limpiar_referencias,
                      fg_color="#c62828", hover_color="#8b0000",
                      width=110, font=ctk.CTkFont(size=15)
                      ).pack(side="left", padx=5)

        self._refs_count_lbl = ctk.CTkLabel(
            btn_f, text="Sin referencias cargadas",
            font=ctk.CTkFont(size=15), text_color="#888")
        self._refs_count_lbl.pack(side="left", padx=10)

        # Lista previa
        self._refs_scroll = ctk.CTkScrollableFrame(
            tab, label_text="Referencias cargadas")
        self._refs_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 6))

    # ═════════════════════════════════════════════════════════════
    # TAB 3 — FIGURAS + TABLAS
    # ═════════════════════════════════════════════════════════════

    def _construir_tab_figuras(self):
        tab = self.tabs.tab("🖼️  Figuras")

        # ── Segmented: Figuras / Tablas ───────────────────────────
        seg = ctk.CTkSegmentedButton(
            tab,
            values=["🖼️  Figuras", "📊  Tablas"],
            command=self._cambiar_panel_media,
            font=ctk.CTkFont(size=15), width=280)
        seg.set("🖼️  Figuras")
        seg.pack(pady=(10, 6))

        # ── Panel Figuras ─────────────────────────────────────────
        self._panel_figs = ctk.CTkFrame(tab, fg_color="transparent")
        self._panel_figs.pack(fill="both", expand=True)

        info_f = ctk.CTkLabel(
            self._panel_figs,
            text="Agrega imágenes con pie de figura. Numeración automática (Figura 1, 2…).\n",
            font=ctk.CTkFont(size=15), justify="left", text_color="#aaa")
        info_f.pack(anchor="w", padx=12, pady=(4, 6))

        bf = ctk.CTkFrame(self._panel_figs, fg_color="transparent")
        bf.pack(fill="x", padx=10, pady=(0, 6))
        ctk.CTkButton(bf, text="➕  Agregar figura",
                      command=self._agregar_figura,
                      fg_color="#1565c0", hover_color="#0d47a1",
                      width=170, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        ctk.CTkButton(bf, text="🗑  Limpiar",
                      command=self._limpiar_figuras,
                      fg_color="#c62828", hover_color="#8b0000",
                      width=110, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        self._figs_count_lbl = ctk.CTkLabel(bf, text="Sin figuras",
                                             font=ctk.CTkFont(size=15), text_color="#888")
        self._figs_count_lbl.pack(side="left", padx=10)

        self._figs_scroll = ctk.CTkScrollableFrame(
            self._panel_figs, label_text="Figuras cargadas")
        self._figs_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        # ── Panel Tablas ──────────────────────────────────────────
        self._panel_tabs = ctk.CTkFrame(tab, fg_color="transparent")
        # (oculto hasta que el usuario pulse "Tablas")

        info_t = ctk.CTkLabel(
            self._panel_tabs,
            text="Importa archivos Excel (.xlsx). Numeración automática (Tabla 1, 2…).\n"
                 "Escribe el pie de la tabla y el texto del cuerpo correspondiente a insertar.",
            font=ctk.CTkFont(size=15), justify="left", text_color="#aaa")
        info_t.pack(anchor="w", padx=12, pady=(4, 6))

        bt = ctk.CTkFrame(self._panel_tabs, fg_color="transparent")
        bt.pack(fill="x", padx=10, pady=(0, 6))
        ctk.CTkButton(bt, text="➕  Agregar tabla (.xlsx)",
                      command=self._agregar_tabla,
                      fg_color="#1565c0", hover_color="#0d47a1",
                      width=190, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        ctk.CTkButton(bt, text="🗑  Limpiar",
                      command=self._limpiar_tablas,
                      fg_color="#c62828", hover_color="#8b0000",
                      width=110, font=ctk.CTkFont(size=15)).pack(side="left", padx=5)
        self._tabs_count_lbl = ctk.CTkLabel(bt, text="Sin tablas",
                                             font=ctk.CTkFont(size=15), text_color="#888")
        self._tabs_count_lbl.pack(side="left", padx=10)

        self._tabs_scroll = ctk.CTkScrollableFrame(
            self._panel_tabs, label_text="Tablas cargadas")
        self._tabs_scroll.pack(fill="both", expand=True, padx=10, pady=(0, 6))

    def _cambiar_panel_media(self, valor: str):
        if valor == "🖼️  Figuras":
            self._panel_tabs.pack_forget()
            self._panel_figs.pack(fill="both", expand=True)
        else:
            self._panel_figs.pack_forget()
            self._panel_tabs.pack(fill="both", expand=True)

    # ── Helpers generales ────────────────────────────────────────

    def _set_status(self, msg: str):
        self._status.configure(text=msg)
        self.update_idletasks()

    # Clases que pertenecen a cada zona del segmented button
    _ZONA_CLASES = {
        "📋 Portada": {
            "Título principal", "Título secundario",
            "Filiación", "Email / Metadatos",
            "Cómo citar", "Fecha manuscrito",
            "Encabezado sección", "Palabras clave",
        },
        "📖 Cuerpo": {
            "Subencabezado", "Subencabezado-bajo", "Cuerpo",
            "Encabezado sección", "Referencia", "Título tabla",
        },
        "🖼 Medios": {
            "Pie de figura", "Ignorar",
        },
    }

    def _mostrar_banner(self, msg: str, color_bg="#c8e6c9", color_txt="#1b5e20"):
        """Muestra el banner verde de completado y lo oculta tras 4 segundos."""
        self._banner.configure(text=msg, fg_color=color_bg, text_color=color_txt)
        # Ocultar el scroll temporalmente, insertar banner, volver a mostrar scroll
        self.frame_scroll.pack_forget()
        self._banner.pack(fill="x", padx=8, pady=(0, 4))
        self.frame_scroll.pack(fill="both", expand=True, padx=0, pady=(0, 2))
        self.after(4000, self._ocultar_banner)

    def _ocultar_banner(self):
        self._banner.pack_forget()

    def _actualizar_stats(self, *_):
        if not self.datos_bloques:
            self._stats_lbl.configure(text=""); return
        c = Counter(b["menu"].get() for b in self.datos_bloques)
        self._stats_lbl.configure(
            text="  ".join(f"{k[:5]}:{v}" for k, v in c.most_common(6)))

    def _aplicar_filtro(self, valor: str):
        for b in self.datos_bloques:
            cls = b["menu"].get()
            if valor == "Todos" or cls == valor:
                b["frame"].pack(fill="x", padx=8, pady=2)
            else:
                b["frame"].pack_forget()

    # ── Leyenda ───────────────────────────────────────────────────

    def _construir_leyenda(self, parent):
        cols = 4
        for i, (cls, color, etiqueta) in enumerate(COLORES_UI):
            row, col = divmod(i, cols)
            celda = ctk.CTkFrame(self._leyenda_panel, fg_color="transparent")
            celda.grid(row=row, column=col, padx=10, pady=4, sticky="w")
            sw = ctk.CTkFrame(celda, width=18, height=18,
                               fg_color=color, corner_radius=3)
            sw.pack(side="left", padx=(0, 5)); sw.pack_propagate(False)
            ctk.CTkLabel(celda, text=f"{cls}  ({etiqueta})",
                         font=ctk.CTkFont(size=15),
                         text_color="#ddd").pack(side="left")

    def _toggle_leyenda(self):
        self._leyenda_visible = not self._leyenda_visible
        if self._leyenda_visible:
            self._leyenda_panel.pack(fill="x", padx=0, pady=(0, 6))
        else:
            self._leyenda_panel.pack_forget()

    # ═════════════════════════════════════════════════════════════
    # REFERENCIAS externas
    # ═════════════════════════════════════════════════════════════

    def evento_cargar_referencias(self):
        ruta = filedialog.askopenfilename(
            title="Referencias (.txt)",
            filetypes=[("Texto", "*.txt"), ("Todos", "*.*")])
        if not ruta: return
        try:
            with open(ruta, encoding="utf-8", errors="replace") as f:
                contenido = f.read()
            self.referencias_externas = _parsear_referencias(contenido)
            self._refrescar_lista_refs()
            n = len(self.referencias_externas)
            self._refs_count_lbl.configure(
                text=f"✓ {n} referencias  ({os.path.basename(ruta)})",
                text_color="#a5d6a7")
            self._set_status(f"✓ {n} referencias cargadas desde '{os.path.basename(ruta)}'")
        except Exception as e:
            self._set_status(f"❌ Error: {e}")

    def _limpiar_referencias(self):
        self.referencias_externas = []
        self._refrescar_lista_refs()
        self._refs_count_lbl.configure(
            text="Sin referencias cargadas", text_color="#888")
        self._set_status("Referencias limpiadas.")

    def _refrescar_lista_refs(self):
        for w in self._refs_scroll.winfo_children():
            w.destroy()
        for i, ref in enumerate(self.referencias_externas, 1):
            frame = ctk.CTkFrame(self._refs_scroll, fg_color="#252525", corner_radius=4)
            frame.pack(fill="x", padx=4, pady=2)
            ctk.CTkLabel(frame, text=f"{i}.", width=32,
                         font=ctk.CTkFont(size=15, weight="bold"),
                         text_color="#aaa").pack(side="left", padx=(6, 2), pady=4)
            ctk.CTkLabel(frame, text=ref[:160] + ("…" if len(ref) > 160 else ""),
                         wraplength=780, justify="left",
                         font=ctk.CTkFont(size=15)).pack(
                             side="left", padx=4, pady=4, fill="x", expand=True)

    # ═════════════════════════════════════════════════════════════
    # FIGURAS manuales
    # ═════════════════════════════════════════════════════════════

    def _limpiar_cache_figuras_auto(self):
        if self._auto_fig_dir and os.path.isdir(self._auto_fig_dir):
            try:
                shutil.rmtree(self._auto_fig_dir)
            except Exception:
                pass
        self._auto_fig_dir = None

    def _extraer_figuras_desde_pdf(self, doc, ruta_pdf: str) -> list[dict]:
        """Extrae imagenes del PDF y propone pie por cercania, se puede editar de manera manual."""
        self._limpiar_cache_figuras_auto()

        base = os.path.splitext(os.path.basename(ruta_pdf))[0]
        base = re.sub(r"[^A-Za-z0-9_-]+", "_", base).strip("_") or "pdf"
        base = base[:24]
        try:
            self._auto_fig_dir = tempfile.mkdtemp(prefix=f"pm_fig_{base}_")
        except Exception:
            self._auto_fig_dir = os.path.join(os.getcwd(), f"_pm_fig_{base}")
            os.makedirs(self._auto_fig_dir, exist_ok=True)

        pat_caption = re.compile(
            r"^\s*(figura|fig\.?|figure)\s*\d+[a-z]?(?:\s*-\s*[a-z])?[\.\):]?\s*",
            re.IGNORECASE,
        )

        encontrados = []
        for pnum in range(len(doc)):
            page = doc.load_page(pnum)
            blocks = page.get_text("dict").get("blocks", [])

            text_blocks = []
            for block in blocks:
                if block.get("type") != 0:
                    continue
                txt = self._texto_bloque(block).strip()
                if len(txt) < 3:
                    continue
                x0, y0, x1, y1 = block.get("bbox", (0, 0, 0, 0))
                text_blocks.append({"texto": txt, "rect": fitz.Rect(x0, y0, x1, y1)})

            vistos = set()
            for img_idx, img in enumerate(page.get_images(full=True), 1):
                xref = img[0]
                try:
                    img_data = doc.extract_image(xref)
                except Exception:
                    continue
                if not img_data or "image" not in img_data:
                    continue

                ext = (img_data.get("ext") or "png").lower()
                ext = re.sub(r"[^a-z0-9]", "", ext) or "png"
                if ext == "jpx":
                    ext = "jpg"

                rects = page.get_image_rects(xref) or []
                for occ_idx, rect in enumerate(rects, 1):
                    if rect.width < 32 or rect.height < 32:
                        continue

                    key = (xref, round(rect.x0, 1), round(rect.y0, 1),
                           round(rect.x1, 1), round(rect.y1, 1))
                    if key in vistos:
                        continue
                    vistos.add(key)

                    nom = f"p{pnum+1:03d}_img{img_idx:02d}_{occ_idx:02d}.{ext}"
                    ruta_img = os.path.join(self._auto_fig_dir, nom)
                    try:
                        with open(ruta_img, "wb") as f:
                            f.write(img_data["image"])
                    except Exception:
                        continue

                    pie_auto = ""
                    candidatos = []
                    for tb in text_blocks:
                        tb_rect = tb["rect"]
                        overlap = max(0.0, min(rect.x1, tb_rect.x1) - max(rect.x0, tb_rect.x0))
                        ancho_ref = max(1.0, min(rect.width, tb_rect.width))
                        if overlap / ancho_ref < 0.15:
                            continue

                        dy = tb_rect.y0 - rect.y1
                        if dy < -8 or dy > 180:
                            continue

                        txt = tb["texto"].strip()
                        if not pat_caption.match(txt):
                            continue
                        candidatos.append((abs(dy), txt))

                    if candidatos:
                        candidatos.sort(key=lambda c: c[0])
                        pie_auto = _limpiar_prefijo_pie_figura(candidatos[0][1])

                    encontrados.append({
                        "ruta": ruta_img,
                        "pie": pie_auto,
                        "ancla": "",
                        "pagina": pnum + 1,
                        "origen": "auto_pdf",
                        "_y": rect.y0,
                        "_x": rect.x0,
                    })

        encontrados.sort(key=lambda f: (f.get("pagina", 0), f.get("_y", 0), f.get("_x", 0)))
        for fig in encontrados:
            fig.pop("_y", None)
            fig.pop("_x", None)
        return encontrados

    def _agregar_figura(self):
        self._sync_pies()   # guardar pie y ancla actuales antes de reconstruir
        rutas = filedialog.askopenfilenames(
            title="Selecciona imagen(es)",
            filetypes=[
                ("Imágenes", "*.jpg *.jpeg *.png *.gif *.webp *.bmp *.tiff"),
                ("Todos", "*.*"),
            ])
        if not rutas: return
        for ruta in rutas:
            self.figuras_manuales.append({"ruta": ruta, "pie": "", "ancla": ""})
        self._refrescar_lista_figuras()
        self._set_status(f"✓ {len(self.figuras_manuales)} figura(s) en total.")

    def _limpiar_figuras(self):
        self._limpiar_cache_figuras_auto()
        self.figuras_manuales = []
        self._refrescar_lista_figuras()
        self._figs_count_lbl.configure(text="Sin figuras", text_color="#888")
        self._set_status("Figuras limpiadas.")

    def _refrescar_lista_figuras(self):
        for w in self._figs_scroll.winfo_children():
            w.destroy()

        for i, fig in enumerate(self.figuras_manuales):
            num = i + 1
            frame = ctk.CTkFrame(self._figs_scroll, fg_color="#1e2a1e", corner_radius=6)
            frame.pack(fill="x", padx=4, pady=6)
            frame.columnconfigure(2, weight=1)

            # Miniatura
            try:
                img_pil = PILImage.open(fig["ruta"])
                img_pil.thumbnail((72, 72))
                thumb   = ctk.CTkImage(img_pil, size=img_pil.size)
                ctk.CTkLabel(frame, image=thumb, text="").grid(
                    row=0, column=0, rowspan=3, padx=(8, 6), pady=8)
            except Exception:
                ctk.CTkLabel(frame, text="🖼️", font=ctk.CTkFont(size=28),
                             width=72).grid(row=0, column=0, rowspan=3,
                                            padx=(8, 6), pady=8)

            ctk.CTkLabel(frame, text=f"Figura {num}",
                         font=ctk.CTkFont(size=15, weight="bold"),
                         text_color="#a5d6a7").grid(
                row=0, column=1, padx=(0, 8), pady=(8, 0), sticky="w")

            subt = os.path.basename(fig["ruta"])
            if fig.get("origen") == "auto_pdf" and fig.get("pagina"):
                subt = f'{subt} - pagina {fig["pagina"]} (auto)'
            ctk.CTkLabel(frame, text=subt,
                         font=ctk.CTkFont(size=15), text_color="#666").grid(
                row=1, column=1, padx=(0, 8), sticky="w")

            # Campo pie de figura
            pie_var = ctk.StringVar(value=fig.get("pie", ""))
            ctk.CTkEntry(frame,
                         placeholder_text=f"Pie de la Figura {num}…",
                         textvariable=pie_var,
                         font=ctk.CTkFont(size=15), height=36).grid(
                row=0, column=2, columnspan=2,
                padx=(0, 8), pady=(8, 2), sticky="ew")
            fig["_var"] = pie_var

            # Campo ancla
            ctk.CTkLabel(frame,
                         text="📍 Pega aquí el texto del párrafo donde va la figura:",
                         font=ctk.CTkFont(size=15), text_color="#a5d6a7").grid(
                row=2, column=1, columnspan=3, padx=(0, 8), pady=(4, 0), sticky="w")

            anc_var = ctk.StringVar(value=fig.get("ancla", ""))
            ctk.CTkEntry(frame,
                         placeholder_text='Ej: "Se muestra en la Figura 1A-B."',
                         textvariable=anc_var,
                         font=ctk.CTkFont(size=15), height=36).grid(
                row=3, column=1, columnspan=3,
                padx=(0, 8), pady=(0, 8), sticky="ew")
            fig["_var_anc"] = anc_var

            def _borrar(idx=i):
                self._sync_pies()
                self.figuras_manuales.pop(idx)
                self._refrescar_lista_figuras()
                self._set_status(f"Figura {idx+1} eliminada.")

            ctk.CTkButton(frame, text="✕", width=28, height=34,
                          fg_color="#c62828", hover_color="#8b0000",
                          command=_borrar,
                          font=ctk.CTkFont(size=15)).grid(
                row=0, column=4, padx=(0, 6), pady=8)

        n = len(self.figuras_manuales)
        self._figs_count_lbl.configure(
            text=f"{n} figura{'s' if n != 1 else ''}" if n else "Sin figuras",
            text_color="#a5d6a7" if n else "#888")

    def _sync_pies(self):
        """Sincroniza pie y ancla con self.figuras_manuales."""
        for fig in self.figuras_manuales:
            if "_var" in fig:
                fig["pie"] = fig["_var"].get()
            if "_var_anc" in fig:
                fig["ancla"] = fig["_var_anc"].get()

    def _sync_titulos_tablas(self):
        for tab in self.tablas_manuales:
            if "_var_tit" in tab:
                tab["titulo"] = tab["_var_tit"].get()
            if "_var_anc" in tab:
                tab["ancla"]  = tab["_var_anc"].get()

    # ─── Tablas ───────────────────────────────────────────────────

    def _limpiar_cache_tablas_auto(self):
        if self._auto_tab_dir and os.path.isdir(self._auto_tab_dir):
            try:
                shutil.rmtree(self._auto_tab_dir)
            except Exception:
                pass
        self._auto_tab_dir = None

    def _remover_tablas_auto(self):
        self.tablas_manuales = [
            t for t in self.tablas_manuales if t.get("origen") != "auto_pdf"
        ]

    def _extraer_tablas_desde_pdf(self, doc, ruta_pdf: str):
        """Devuelve (tablas_auto, rects_por_pagina) para tablas detectadas."""
        self._limpiar_cache_tablas_auto()

        try:
            import openpyxl
        except Exception:
            return [], {}

        base = os.path.splitext(os.path.basename(ruta_pdf))[0]
        base = re.sub(r"[^A-Za-z0-9_-]+", "_", base).strip("_") or "pdf"
        base = base[:24]
        try:
            self._auto_tab_dir = tempfile.mkdtemp(prefix=f"pm_tab_{base}_")
        except Exception:
            self._auto_tab_dir = os.path.join(os.getcwd(), f"_pm_tab_{base}")
            os.makedirs(self._auto_tab_dir, exist_ok=True)

        pat_title = re.compile(
            r"^\s*(tabla|table)\s*\d+[a-z]?(?:\s*-\s*[a-z])?[\.\):]?\s*",
            re.IGNORECASE,
        )

        tablas_auto = []
        rects_por_pagina = {}

        for pnum in range(len(doc)):
            page = doc.load_page(pnum)
            if not hasattr(page, "find_tables"):
                continue

            try:
                finder = page.find_tables()
            except Exception:
                continue
            tables = getattr(finder, "tables", []) if finder else []
            if not tables:
                continue

            blocks = page.get_text("dict").get("blocks", [])
            text_blocks = []
            for block in blocks:
                if block.get("type") != 0:
                    continue
                txt = self._texto_bloque(block).strip()
                if len(txt) < 3:
                    continue
                x0, y0, x1, y1 = block.get("bbox", (0, 0, 0, 0))
                text_blocks.append((fitz.Rect(x0, y0, x1, y1), txt))

            vistos = set()
            for idx_tab, tb in enumerate(tables, 1):
                bbox = getattr(tb, "bbox", None)
                if not bbox:
                    continue
                rect = fitz.Rect(bbox)
                key = (round(rect.x0, 1), round(rect.y0, 1), round(rect.x1, 1), round(rect.y1, 1))
                if key in vistos:
                    continue
                vistos.add(key)

                try:
                    rows = tb.extract()
                except Exception:
                    continue
                if not rows or len(rows) < 2:
                    continue

                cleaned = []
                max_cols = 0
                for row in rows:
                    if row is None:
                        continue
                    vals = []
                    has_content = False
                    for cell in row:
                        txt = "" if cell is None else str(cell)
                        txt = re.sub(r"\s+", " ", txt).strip()
                        if txt:
                            has_content = True
                        vals.append(txt)
                    if has_content:
                        cleaned.append(vals)
                        max_cols = max(max_cols, len(vals))

                if len(cleaned) < 2 or max_cols < 2:
                    continue

                for row in cleaned:
                    if len(row) < max_cols:
                        row.extend([""] * (max_cols - len(row)))

                titulo = ""
                candidatos_titulo = []
                for txt_rect, txt in text_blocks:
                    overlap = max(0.0, min(rect.x1, txt_rect.x1) - max(rect.x0, txt_rect.x0))
                    ancho_ref = max(1.0, min(rect.width, txt_rect.width))
                    if overlap / ancho_ref < 0.15:
                        continue

                    dy_up = rect.y0 - txt_rect.y1
                    dy_down = txt_rect.y0 - rect.y1
                    if 0 <= dy_up <= 140 and pat_title.match(txt):
                        candidatos_titulo.append((dy_up, txt))
                    elif 0 <= dy_down <= 80 and pat_title.match(txt):
                        candidatos_titulo.append((dy_down + 200, txt))

                if candidatos_titulo:
                    candidatos_titulo.sort(key=lambda c: c[0])
                    titulo = _limpiar_prefijo_titulo_tabla(candidatos_titulo[0][1])

                xlsx_path = os.path.join(
                    self._auto_tab_dir,
                    f"p{pnum+1:03d}_tabla{idx_tab:02d}.xlsx",
                )
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Tabla"
                    for row in cleaned:
                        ws.append(row)
                    wb.save(xlsx_path)
                    wb.close()
                except Exception:
                    continue

                rects_por_pagina.setdefault(pnum, []).append(rect)
                tablas_auto.append({
                    "ruta": xlsx_path,
                    "hoja": "Tabla",
                    "titulo": titulo,
                    "ancla": "",
                    "pagina": pnum + 1,
                    "origen": "auto_pdf",
                    "_y": rect.y0,
                    "_x": rect.x0,
                })

        tablas_auto.sort(key=lambda t: (t.get("pagina", 0), t.get("_y", 0), t.get("_x", 0)))
        for tab in tablas_auto:
            tab.pop("_y", None)
            tab.pop("_x", None)
        return tablas_auto, rects_por_pagina

    def _agregar_tabla(self):
        self._sync_titulos_tablas()
        rutas = filedialog.askopenfilenames(
            title="Selecciona archivo(s) Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if not rutas: return
        try:
            import openpyxl
        except ImportError:
            self._set_status("❌ Instala openpyxl: pip install openpyxl"); return

        for ruta in rutas:
            try:
                wb = openpyxl.load_workbook(ruta, data_only=True)
                hojas = wb.sheetnames
                wb.close()
                for hoja in hojas:
                    self.tablas_manuales.append({
                        "ruta":   ruta,
                        "hoja":   hoja,
                        "titulo": "",
                        "ancla":  ""
                    })
            except Exception as e:
                self._set_status(f"❌ Error leyendo {os.path.basename(ruta)}: {e}")
                return

        self._refrescar_lista_tablas()
        self._set_status(f"✓ {len(self.tablas_manuales)} tabla(s) detectadas.")

    def _limpiar_tablas(self):
        self._limpiar_cache_tablas_auto()
        self.tablas_manuales = []
        self._refrescar_lista_tablas()
        self._tabs_count_lbl.configure(text="Sin tablas", text_color="#888")
        self._set_status("Tablas limpiadas.")

    def _refrescar_lista_tablas(self):
        for w in self._tabs_scroll.winfo_children():
            w.destroy()

        for i, tab_item in enumerate(self.tablas_manuales):
            num   = i + 1
            frame = ctk.CTkFrame(self._tabs_scroll, fg_color="#1a1a2e", corner_radius=6)
            frame.pack(fill="x", padx=4, pady=6)
            frame.columnconfigure(2, weight=1)

            # ── Columna izquierda: icono + etiqueta ──
            ctk.CTkLabel(frame, text="📊",
                         font=ctk.CTkFont(size=26), width=48).grid(
                row=0, column=0, rowspan=3, padx=(8, 4), pady=8, sticky="n")

            ctk.CTkLabel(frame, text=f"Tabla {num}",
                         font=ctk.CTkFont(size=15, weight="bold"),
                         text_color="#ce93d8").grid(
                row=0, column=1, padx=(0, 6), pady=(8, 0), sticky="w")

            # Nombre del archivo + hoja
            hoja = tab_item.get("hoja", "")
            archivo = os.path.basename(tab_item["ruta"])
            subtitulo = f"{archivo}  >  {hoja}" if hoja else archivo
            if tab_item.get("origen") == "auto_pdf" and tab_item.get("pagina"):
                subtitulo += f'  -  pagina {tab_item["pagina"]} (auto)'
            ctk.CTkLabel(frame, text=subtitulo,
                         font=ctk.CTkFont(size=15), text_color="#888").grid(
                row=1, column=1, padx=(0, 6), pady=0, sticky="w")

            # ── Campo 1: Título ──
            tit_var = ctk.StringVar(value=tab_item.get("titulo", ""))
            ctk.CTkEntry(
                frame,
                placeholder_text=f"Título de la Tabla {num}…",
                textvariable=tit_var,
                font=ctk.CTkFont(size=15), height=36
            ).grid(row=0, column=2, columnspan=2,
                   padx=(0, 8), pady=(8, 2), sticky="ew")
            tab_item["_var_tit"] = tit_var

            # ── Campo 2: Texto ancla ──
            ctk.CTkLabel(frame,
                         text="📍 Pega aquí el texto del párrafo donde va la tabla:",
                         font=ctk.CTkFont(size=15), text_color="#ce93d8").grid(
                row=2, column=1, columnspan=3, padx=(0, 8), pady=(4, 0), sticky="w")

            anc_var = ctk.StringVar(value=tab_item.get("ancla", ""))
            ctk.CTkEntry(
                frame,
                placeholder_text='Ej: "la Dra. Elena Centeno (Tabla 1)."',
                textvariable=anc_var,
                font=ctk.CTkFont(size=15), height=36
            ).grid(row=3, column=1, columnspan=3,
                   padx=(0, 8), pady=(0, 8), sticky="ew")
            tab_item["_var_anc"] = anc_var

            # ── Botón eliminar ──
            def _borrar_t(idx=i):
                self._sync_titulos_tablas()
                self.tablas_manuales.pop(idx)
                self._refrescar_lista_tablas()
                self._set_status(f"Tabla {idx+1} eliminada.")

            ctk.CTkButton(frame, text="✕", width=28, height=34,
                          fg_color="#c62828", hover_color="#8b0000",
                          command=_borrar_t,
                          font=ctk.CTkFont(size=15)).grid(
                row=0, column=4, padx=(0, 6), pady=8)

        n = len(self.tablas_manuales)
        self._tabs_count_lbl.configure(
            text=f"{n} tabla{'s' if n != 1 else ''}" if n else "Sin tablas",
            text_color="#ce93d8" if n else "#888")

    # ═════════════════════════════════════════════════════════════
    # EXTRACCIÓN PDF
    # ═════════════════════════════════════════════════════════════

    def _info_fuente(self, block):
        sizes, bold, italic, fuentes = [], False, False, []
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                sizes.append(span["size"])
                fuentes.append(span.get("font", "").lower())
                if span["flags"] & (1 << 4): bold   = True
                if span["flags"] & (1 << 1): italic = True
        avg = sum(sizes)/len(sizes) if sizes else 10.0
        dom = Counter(fuentes).most_common(1)[0][0] if fuentes else ""
        return avg, bold, italic, dom

    def _texto_bloque(self, block):
        """Extrae el texto del bloque reconectando palabras cortadas por guión."""
        lineas = []
        for line in block.get("lines", []):
            lineas.append("".join(s["text"] for s in line.get("spans", [])))

        # Unir líneas: si una termina en guión (- o ­), pegar directamente
        resultado = ""
        for i, linea in enumerate(lineas):
            if i == 0:
                resultado = linea
            else:
                if resultado.endswith("-") or resultado.endswith("\u00ad"):
                    # Guión de corte: quitar guión y pegar sin espacio
                    resultado = resultado.rstrip("-\u00ad") + linea.lstrip()
                else:
                    resultado = resultado + " " + linea

        # Limpiar espacios múltiples y soft-hyphens residuales
        resultado = resultado.replace("\u00ad", "")       # soft hyphens
        resultado = re.sub(r"\s+", " ", resultado).strip()
        # Guiones de corte que quedaron con espacio: "pos­ teriormente" → "posteriormente"
        resultado = re.sub(r"(\w)-\s+(\w)", r"\1\2", resultado)
        return resultado

    def _clasificar_auto(self, texto, size, bold, italic, font, body_size):
        t, t_low = texto.strip(), texto.strip().lower()

        secciones = {
            "resumen","abstract","resumen no técnico","non-technical abstract",
            "palabras clave","keywords","introducción","introduction",
            "conclusiones","conclusions","referencias","references",
            "agradecimientos","acknowledgements","discusión","discussion",
            "metodología","methods","resultados","results",
            "contribuciones de los autores",
            "paleontología sistemática","systematic palaeontology",
        }
        if t_low in secciones: return "Encabezado sección"
        if _es_como_citar(t):  return "Cómo citar"
        if _es_fecha_mss(t) or _es_doi(t): return "Fecha manuscrito"
        if re.match(r"^(palabras\s+clave|keywords)\s*[:\.]", t_low):
            return "Palabras clave"
        if re.search(r"@[\w\-\.]+\.\w{2,}", t) and len(t) < 120:
            return "Email / Metadatos"

        s = round(size)
        # ── Subtítulos numerados ─────────────────────────────────────────────
        # Nivel 1:  "4. Texto"      (sin punto decimal) → bold
        # Nivel 2+: "4.1 Texto"    (con punto decimal) → sin bold
        if re.match(r"^\d+\.\s+\S", t) and not re.match(r"^\d+\.\d", t) and s <= 12:
            return "Subencabezado"
        if re.match(r"^\d+\.\d+", t) and s <= 12:
            return "Subencabezado-bajo"
        if s >= 13 and bold and not italic: return "Título principal"
        if s >= 13 and bold and italic:     return "Título secundario"
        if s == 13 and not bold and not italic: return "Autores"
        if s == 12 and italic:  return "Email / Metadatos"
        if s == 12 and not bold: return "Normal"
        if s == 10 and bold:    return "Encabezado sección"
        if s == 10 and not bold: return "Subencabezado"
        if s == 9:
            is_tnr = "times" in font or "roman" in font
            if is_tnr or len(t) > 100: return "Resumen / Abstract"
            if len(t) < 50: return "Filiación"
            return "Resumen / Abstract"
        if len(t) < 4: return "Ignorar"
        return "Normal"

    def evento_cargar_archivo(self):
        ruta = filedialog.askopenfilename(
            title="Selecciona el PDF",
            filetypes=[("Archivos PDF", "*.pdf")])
        if not ruta: return

        # ── Bloquear UI mientras carga ────────────────────────────
        self._btn_cargar.configure(
            state="disabled",
            text="⏳  Analizando…",
            fg_color="#37474f"
        )
        self._status.configure(text_color="#888")   # resetear color
        self._set_status("⏳  Analizando PDF, por favor espera…")
        self.update()   # forzar redibujado antes de bloquear el hilo

        for w in self.frame_scroll.winfo_children():
            w.destroy()
        self.datos_bloques.clear()
        self._limpiar_cache_figuras_auto()
        self.figuras_manuales = []
        self._refrescar_lista_figuras()
        self._limpiar_cache_tablas_auto()
        self._remover_tablas_auto()
        self._refrescar_lista_tablas()

        try:
            doc = fitz.open(ruta)

            all_sizes = []
            for pnum in range(len(doc)):
                for block in doc.load_page(pnum).get_text("dict")["blocks"]:
                    if block["type"] == 0:
                        for line in block.get("lines", []):
                            for span in line.get("spans", []):
                                all_sizes.append(round(span["size"]))
            body_size = Counter(all_sizes).most_common(1)[0][0] if all_sizes else 12

            tablas_auto, rects_tablas_por_pagina = self._extraer_tablas_desde_pdf(doc, ruta)

            def _en_bloque_tabla(pnum: int, bbox) -> bool:
                rects = rects_tablas_por_pagina.get(pnum, [])
                if not rects:
                    return False
                brect = fitz.Rect(bbox)
                area_b = max(1.0, brect.width * brect.height)
                for trect in rects:
                    inter = brect & trect
                    if inter.is_empty:
                        continue
                    if (inter.width * inter.height) / area_b >= 0.40:
                        return True
                return False

            # ── Filtro de cornisas: posición + patrón de texto ────────────
            # Umbral conservador: solo top 5% / bottom 5%.
            # DOIs y números de página están en el 2-4%, nunca más abajo.
            # Encabezados de sección legítimos (Abstract, 1. Intro…) pueden
            # estar al 6-8% de la página y NO deben filtrarse.
            _pat_cornisa_txt = re.compile(
                r"^(https?://doi\.org/|doi\.org/|\d{1,3}$"
                r"|paleontolog[íi]a mexicana\s+vol\."
                r"|velasco-de le[oó]n et al"
                r"|ra[íi]ces de la paleobotánica)",
                re.IGNORECASE
            )
            def _es_cornisa(by0, by1, page_h, pnum):
                if pnum == 0:
                    return False
                # Filtrar por posición muy estricta (top/bottom 5%)
                return by1 < page_h * 0.05 or by0 > page_h * 0.95

            # ─── PASO 1: extraer todos los bloques de texto crudos ────────
            raw = []   # list of {"texto", "size", "bold", "italic", "font", "pnum"}
            for pnum in range(len(doc)):
                page   = doc.load_page(pnum)
                page_h = page.rect.height
                page_w = page.rect.width
                all_blocks = page.get_text("dict")["blocks"]

                # ── Detectar si la página tiene dos columnas ───────────────
                # Tomamos el centroide X de cada bloque de texto y vemos si
                # se agrupan claramente en dos mitades.
                text_blocks = [b for b in all_blocks if b["type"] == 0]
                cx_list = [(b["bbox"][0] + b["bbox"][2]) / 2 for b in text_blocks]

                # Si hay bloques a ambos lados del 55% de la página → dos columnas
                mid = page_w * 0.55
                left_cx  = [cx for cx in cx_list if cx < mid]
                right_cx = [cx for cx in cx_list if cx >= mid]
                two_cols = len(left_cx) >= 2 and len(right_cx) >= 2

                if two_cols:
                    # Separar en columna izquierda y derecha, cada una ordenada por Y
                    left_blocks  = sorted(
                        [b for b in all_blocks if (b["bbox"][0]+b["bbox"][2])/2 < mid],
                        key=lambda b: b["bbox"][1])
                    right_blocks = sorted(
                        [b for b in all_blocks if (b["bbox"][0]+b["bbox"][2])/2 >= mid],
                        key=lambda b: b["bbox"][1])
                    ordered_blocks = left_blocks + right_blocks
                else:
                    ordered_blocks = sorted(all_blocks, key=lambda b: b["bbox"][1])

                for block in ordered_blocks:
                    if block["type"] == 1:
                        x0, y0, x1, y1 = block["bbox"]
                        w_px, h_px = abs(x1-x0), abs(y1-y0)
                        ignorar = (w_px < 32 and h_px < 32) or \
                                  _es_cornisa(y0, y1, page_h, pnum)
                        raw.append({
                            "texto": f"[IMAGEN {w_px:.0f}×{h_px:.0f}px]",
                            "clasificacion": "Ignorar" if ignorar else "Imagen",
                            "size": 0, "bold": False, "italic": False,
                            "imagen": True, "pnum": pnum,
                        })
                        continue
                    if block["type"] != 0:
                        continue
                    bx0, by0, bx1, by1 = block["bbox"]
                    if _es_cornisa(by0, by1, page_h, pnum):
                        continue
                    if _en_bloque_tabla(pnum, (bx0, by0, bx1, by1)):
                        continue
                    size, bold, italic, font = self._info_fuente(block)
                    texto = self._texto_bloque(block)
                    if not texto or len(texto) < 3:
                        continue
                    # Zona 5-12%: filtrar solo si el texto parece cornisa
                    if pnum > 0 and by1 < page_h * 0.12:
                        if _pat_cornisa_txt.search(texto.strip()[:100]):
                            continue
                    raw.append({
                        "texto": texto, "size": size,
                        "bold": bold, "italic": italic, "font": font,
                        "imagen": False, "clasificacion": None,
                        "pnum": pnum,
                    })

            # ─── PASO 2: detectar zonas ───────────────────────────────────
            # Zona A (portada/frente): antes del primer encabezado numerado
            # Zona B (cuerpo):         desde primer "N. Texto" hasta "Referencias"
            # Zona C (post-refs):      después de "Referencias"
            #
            # Patrón de encabezado numerado nivel-1: "4.Texto" o "4. Texto"
            _pat_nivel1 = re.compile(r"^\d+\.\s*\S")
            _pat_nivel2 = re.compile(r"^\d+\.\d+")
            _SECCIONES_EXACTAS = {
                "resumen", "abstract", "resumen no técnico",
                "non-technical abstract", "palabras clave", "keywords",
                "referencias", "references", "conclusiones", "conclusions",
                "agradecimientos", "acknowledgements", "discusión",
                "resultados", "introducción", "metodología",
                "contribuciones de los autores",
                "conflicto de intereses", "conflict of interest",
            }

            zona_b_inicio = None   # índice en raw donde empieza el cuerpo
            zona_b_fin    = None   # índice en raw donde empieza "Referencias"

            for i, r in enumerate(raw):
                if r["imagen"] or r["clasificacion"] == "Ignorar":
                    continue
                t_low = r["texto"].strip().lower()
                s     = round(r["size"])

                # Primer encabezado numerado nivel-1 → inicio zona B
                if zona_b_inicio is None:
                    if _pat_nivel1.match(r["texto"].strip()) and \
                       not _pat_nivel2.match(r["texto"].strip()) and \
                       s <= 12:
                        zona_b_inicio = i

                # "Referencias" o "References" exacto → fin zona B
                if zona_b_inicio is not None and zona_b_fin is None:
                    if t_low in ("referencias", "references"):
                        zona_b_fin = i

            # ─── PASO 3: clasificar cada bloque según su zona ─────────────
            bloques_raw = []
            for i, r in enumerate(raw):
                if r["imagen"]:
                    bloques_raw.append({
                        "contenido": r["texto"],
                        "clasificacion": r["clasificacion"],
                        "size": 0, "bold": False, "italic": False,
                    })
                    continue

                texto = r["texto"].strip()
                t_low = texto.lower()
                s     = round(r["size"])

                # ── Zona B: cuerpo del artículo ──────────────────────────
                en_b = zona_b_inicio is not None and \
                       i >= zona_b_inicio and \
                       (zona_b_fin is None or i < zona_b_fin)

                if en_b:
                    if _pat_nivel1.match(texto) and not _pat_nivel2.match(texto):
                        cls = "Subencabezado"
                    elif _pat_nivel2.match(texto):
                        cls = "Subencabezado-bajo"
                    elif _es_como_citar(texto):
                        cls = "Cómo citar"
                    elif _es_fecha_mss(texto) or _es_doi(texto):
                        cls = "Fecha manuscrito"
                    elif t_low in _SECCIONES_EXACTAS:
                        cls = "Encabezado sección"
                    elif re.search(r"\btabla\s+\d+[\.\:\s]", t_low):
                        cls = "Título tabla"
                    elif re.match(r"^figura\s+\d+[\.\:\s]", t_low):
                        cls = "Pie de figura"   # pie de foto del PDF → ignorar en HTML
                    else:
                        cls = "Cuerpo"
                else:
                    # ── Zona A / C: clasificación normal ─────────────────
                    if i == zona_b_fin:
                        cls = "Encabezado sección"   # la palabra "Referencias"
                    else:
                        cls = self._clasificar_auto(
                            texto, r["size"], r["bold"],
                            r["italic"], r["font"], body_size)

                bloques_raw.append({
                    "contenido": texto,
                    "clasificacion": _CLASE_COMPAT.get(cls, cls),
                    "size": r["size"],
                    "bold": r["bold"],
                    "italic": r["italic"],
                    "pnum": r.get("pnum", 0),
                })

            # ─── PASO 3b: suprimir filas de tabla del PDF ─────────────────
            # Regla universal para todas las revistas:
            # 1. Al detectar "Título tabla" → activar modo tabla
            # 2. También detectar filas-de-tabla por patrón (año + autor, celdas cortas)
            #    incluso sin haber visto el título todavía
            # 3. Salir del modo cuando aparece un párrafo real (>200 chars + punto final)
            #    o un encabezado/subencabezado

            # Patrón universal de fila de tabla académica:
            # SOLO activar proactivamente si empieza con año (1800-2099).
            # La supresión genérica de bloques cortos causaba falsos positivos
            # con párrafos partidos por columnas del PDF.
            _pat_fila_tabla = re.compile(
                r"^(1[89]\d{2}|20\d{2})\s+\S"   # año AAAA + contenido
            )

            def _parece_fila_tabla(item):
                t   = item["contenido"].strip()
                cls = item["clasificacion"]
                if cls != "Cuerpo": return False
                # Solo suprimir proactivamente si empieza con año de 4 dígitos
                if _pat_fila_tabla.match(t): return True
                return False

            bloques_sin_tabla_pdf = []
            en_modo_tabla = False
            for item in bloques_raw:
                cls_i = item["clasificacion"]

                # Activar por título explícito
                if cls_i == "Título tabla":
                    en_modo_tabla = True
                    bloques_sin_tabla_pdf.append(item)
                    continue

                # Activar por patrón de fila (sin necesitar el título)
                if not en_modo_tabla and _parece_fila_tabla(item):
                    en_modo_tabla = True
                    # No agregar esta fila — es contenido de tabla del PDF
                    continue

                if en_modo_tabla:
                    # ¿Salimos del modo tabla?
                    es_fin = cls_i in (
                        "Subencabezado", "Subencabezado-bajo",
                        "Encabezado sección", "Cómo citar",
                        "Fecha manuscrito",
                    ) or (
                        cls_i == "Cuerpo" and
                        len(item["contenido"]) > 200 and
                        item["contenido"].rstrip()[-1] in ".?!"
                    )
                    if es_fin:
                        en_modo_tabla = False
                        bloques_sin_tabla_pdf.append(item)
                    # else: fila de tabla del PDF → descartar
                else:
                    bloques_sin_tabla_pdf.append(item)
            bloques_raw = bloques_sin_tabla_pdf
            # Encabezados que pueden llegar pegados al texto anterior
            HEADERS_EMBEBIDOS = [
                "Non-technical Abstract", "Non-Technical Abstract",
                "Resumen no técnico", "Resumen no Técnico",
                # "Abstract" se maneja por separado con regex (ver abajo)
            ]
            # Detectar "Abstract" sola sin cortar "Non-technical Abstract"
            _pat_abstract_solo = re.compile(
                r'(?<!\w)(Abstract)(?!\w)',   # palabra completa: no parte de "Abstracto"
                re.IGNORECASE
            )
            # Patrón que identifica títulos de tabla embebidos en un bloque mayor.
            # Captura todo desde "Tabla N." hasta el fin de esa "oración" (hasta \n
            # o hasta el final del texto).
            _pat_tabla_embebida = re.compile(
                r'(?<!\w)(Tabla\s+\d+[\.\:\s][^\n]{0,200})',
                re.IGNORECASE
            )

            bloques_clean = []
            for item in bloques_raw:
                txt    = item["contenido"]
                partido = False

                # 1. Separar encabezados de sección embebidos
                for hdr in HEADERS_EMBEBIDOS:
                    idx = txt.find(hdr)
                    if idx > 0:
                        antes = txt[:idx].strip()
                        if antes:
                            b1 = dict(item); b1["contenido"] = antes
                            b1["clasificacion"] = self._clasificar_auto(
                                antes, item["size"], item["bold"],
                                item["italic"], "", body_size)
                            bloques_clean.append(b1)
                        bloques_clean.append({
                            "contenido": hdr, "clasificacion": "Encabezado sección",
                            "size": item["size"], "bold": item["bold"],
                            "italic": item["italic"], "pnum": item.get("pnum", 0),
                        })
                        despues = txt[idx+len(hdr):].strip()
                        if despues:
                            b3 = dict(item); b3["contenido"] = despues
                            bloques_clean.append(b3)
                        partido = True
                        break

                if partido:
                    continue

                # 1b. "Abstract" sola pegada al texto anterior
                if not partido:
                    m_abs = _pat_abstract_solo.search(txt)
                    # Solo aplicar si no es parte de "Non-technical Abstract"
                    if m_abs and m_abs.start() > 0:
                        ctx_antes = txt[max(0, m_abs.start()-20):m_abs.start()].lower()
                        es_non_technical = "technical" in ctx_antes or "técnico" in ctx_antes
                        if not es_non_technical:
                            antes = txt[:m_abs.start()].strip()
                            despues = txt[m_abs.end():].strip()
                            if antes:
                                b1 = dict(item); b1["contenido"] = antes
                                b1["clasificacion"] = self._clasificar_auto(
                                    antes, item["size"], item["bold"],
                                    item["italic"], "", body_size)
                                bloques_clean.append(b1)
                            bloques_clean.append({
                                "contenido": "Abstract",
                                "clasificacion": "Encabezado sección",
                                "size": item["size"], "bold": item["bold"],
                                "italic": item["italic"], "pnum": item.get("pnum", 0),
                            })
                            if despues:
                                b3 = dict(item); b3["contenido"] = despues
                                bloques_clean.append(b3)
                            partido = True

                # 2. Separar títulos de tabla embebidos en bloques Cuerpo
                #    Ej: "...texto previo. Tabla 1. Lista de trabajos…\nTexto sig."
                if item["clasificacion"] == "Cuerpo":
                    m = _pat_tabla_embebida.search(txt)
                    if m and m.start() > 0:
                        # hay texto ANTES del título de tabla
                        antes = txt[:m.start()].strip()
                        titulo_tab = m.group(1).strip()
                        despues = txt[m.end():].strip()

                        if antes:
                            b1 = dict(item); b1["contenido"] = antes
                            bloques_clean.append(b1)

                        bloques_clean.append({
                            "contenido": titulo_tab,
                            "clasificacion": "Título tabla",
                            "size": item["size"],
                            "bold": item["bold"],
                            "italic": item["italic"],
                            "pnum": item.get("pnum", 0),
                        })

                        if despues:
                            b3 = dict(item); b3["contenido"] = despues
                            bloques_clean.append(b3)
                        continue   # no agregar el bloque original
                    elif m and m.start() == 0:
                        # el bloque empieza directamente con "Tabla N." → ya
                        # debería haberse clasificado arriba, pero por si acaso:
                        item["clasificacion"] = "Título tabla"

                bloques_clean.append(item)

            # ─── PASO 5: fusionar bloques Cuerpo/Normal consecutivos ──────
            NO_FUSIONAR = {
                "Título principal", "Título secundario", "Autores",
                "Filiación", "Email / Metadatos", "Cómo citar",
                "Fecha manuscrito", "Encabezado sección",
                "Subencabezado",
                "Palabras clave", "Referencia",
                # "Título tabla" se maneja abajo sin flush para no cortar el texto
                # "Pie de figura", "Imagen", "Ignorar" se saltan sin flush
            }

            def _es_continuacion(anterior: str, siguiente: str,
                                  pnum_ant: int, pnum_sig: int) -> bool:
                """Une bloques de la misma oración cortada por salto de página/columna."""
                # Hasta 5 páginas de diferencia (cubre tablas de varias páginas)
                if abs(pnum_sig - pnum_ant) > 5:
                    return False
                ant = anterior.rstrip()
                if not ant:
                    return False
                if ant[-1] in ".?!:":
                    return False
                return True

            fusionados = []
            buf_parrafos: list[tuple] = []  # (texto, pnum)
            buf_item = None

            def _vaciar():
                nonlocal buf_item
                if buf_parrafos and buf_item is not None:
                    merged = dict(buf_item)
                    merged["contenido"]     = "\n\n".join(t for t, _ in buf_parrafos)
                    merged["clasificacion"] = "Cuerpo"
                    fusionados.append(merged)
                buf_parrafos.clear()
                buf_item = None

            for item in bloques_clean:
                cls  = item["clasificacion"]
                pnum = item.get("pnum", 0)

                # Imágenes, pies de figura, ignorados y TABLAS del PDF: añadir
                # a fusionados sin tocar el buffer de texto. Así el texto antes
                # y después de una tabla/imagen queda unido en el mismo párrafo.
                if cls in ("Imagen", "Ignorar", "Pie de figura", "Título tabla"):
                    fusionados.append(item)   # la tabla aparece en su posición
                    continue                  # pero NO interrumpe el flujo de texto

                if cls in NO_FUSIONAR:
                    _vaciar()
                    fusionados.append(item)
                elif cls == "Subencabezado-bajo":
                    if buf_item is None:
                        buf_item = item
                    buf_parrafos.append(("§SUB§" + item["contenido"], pnum))
                else:
                    if buf_item is None:
                        buf_item = item
                    txt_nuevo = item["contenido"]
                    if buf_parrafos:
                        ultimo_txt, ultimo_pnum = buf_parrafos[-1]
                        # Nunca fusionar con un §SUB§ anterior (subtítulo 1.1, 4.2…)
                        if ultimo_txt.startswith("§SUB§"):
                            buf_parrafos.append((txt_nuevo, pnum))
                        elif _es_continuacion(ultimo_txt, txt_nuevo, ultimo_pnum, pnum):
                            t = ultimo_txt.rstrip()
                            if t.endswith("-"):
                                buf_parrafos[-1] = (t[:-1] + txt_nuevo.lstrip(), pnum)
                            else:
                                buf_parrafos[-1] = (t + " " + txt_nuevo.lstrip(), pnum)
                        else:
                            buf_parrafos.append((txt_nuevo, pnum))
                    else:
                        buf_parrafos.append((txt_nuevo, pnum))
            _vaciar()
            bloques_utiles = fusionados

            for item in bloques_utiles:
                self._crear_bloque_ui(item)

            figuras_auto = self._extraer_figuras_desde_pdf(doc, ruta)
            self.figuras_manuales = figuras_auto
            self._refrescar_lista_figuras()
            self.tablas_manuales.extend(tablas_auto)
            self._refrescar_lista_tablas()

            conteo  = Counter(b["clasificacion"] for b in bloques_utiles)
            resumen = "  |  ".join(f"{k}: {v}" for k,v in conteo.most_common(6))
            resumen += f"  |  Figuras auto: {len(figuras_auto)}"
            resumen += f"  |  Tablas auto: {len(tablas_auto)}"
            self._status.configure(text_color="#66bb6a")
            self._set_status(
                f"✅  Análisis completo — {len(bloques_utiles)} bloques  |  base: {body_size}pt  |  {resumen}")
            self._mostrar_banner(
                f"✅  Análisis completo — {len(bloques_utiles)} bloques extraídos")
            self._actualizar_stats()
            self._aplicar_filtro("Todos")

        except Exception as e:
            self._set_status(f"❌ Error: {e}")
            import traceback; traceback.print_exc()

        finally:
            # ── Restaurar botón siempre, haya error o no ─────────
            self._btn_cargar.configure(
                state="normal",
                text="📂  Cargar PDF",
                fg_color="#1565c0"
            )

    def _crear_bloque_ui(self, item):
        cls  = item["clasificacion"]
        cont = item["contenido"]
        size = item.get("size", 10)
        bold = item.get("bold", False)
        ital = item.get("italic", False)

        color   = COLOR_POR_CLASE.get(cls, "#2b2b2b")
        preview = (cont[:150]+"…") if len(cont)>150 else cont

        frame = ctk.CTkFrame(self.frame_scroll, fg_color=color, corner_radius=5)
        frame.pack(fill="x", padx=8, pady=2)
        frame.columnconfigure(1, weight=1)

        badge = f"{size:.0f}pt"+(" B" if bold else "")+(" I" if ital else "")
        ctk.CTkLabel(frame, text=badge, width=62,
                     font=ctk.CTkFont(size=15), text_color="#bbb"
                     ).grid(row=0, column=0, padx=(6,0), pady=5, sticky="w")
        ctk.CTkLabel(frame, text=preview, wraplength=720,
                     justify="left", anchor="w",
                     font=ctk.CTkFont(size=15)
                     ).grid(row=0, column=1, padx=6, pady=5, sticky="ew")

        menu = ctk.CTkOptionMenu(frame, values=OPCIONES, width=175,
                                  font=ctk.CTkFont(size=15),
                                  command=self._actualizar_stats)
        menu.set(cls)
        menu.grid(row=0, column=2, padx=(0,8), pady=5)

        self.datos_bloques.append({
            "contenido": cont, "menu": menu,
            "italic": ital,    "frame": frame,
        })

    # ═════════════════════════════════════════════════════════════
    # EXPORTAR HTML
    # ═════════════════════════════════════════════════════════════

    def evento_exportar_html(self):
        if not self.datos_bloques:
            self._set_status("⚠ Primero carga un PDF."); return
        ruta = filedialog.asksaveasfilename(
            defaultextension=".html",
            filetypes=[("Archivo HTML", "*.html")])
        if not ruta: return

        self._sync_pies()
        self._sync_titulos_tablas()
        self._sync_autores()   # sincronizar autores/ORCID antes de exportar

        # Si hay autores manuales, el bloque de autores del PDF se ignora
        # y se inyecta el bloque manual justo después del título secundario.
        # ── Zona de autores del PDF: ignorar TODO entre titulo-secundario y Resumen ──
        idx_tit_sec = next((i for i, b in enumerate(self.datos_bloques)
                            if b["menu"].get() == "Título secundario"), None)
        idx_resumen = next((i for i, b in enumerate(self.datos_bloques)
                            if b["menu"].get() == "Encabezado sección"
                            and _es_encabezado_resumen(b["contenido"])
                            and (idx_tit_sec is None or i > idx_tit_sec)), None)
        zona_autores_pdf = set()
        if idx_tit_sec is not None and idx_resumen is not None:
            zona_autores_pdf = set(range(idx_tit_sec + 1, idx_resumen))

        # ── Construir HTML de afiliaciones desde .txt ──────────────
        def _afil_a_html(txt: str) -> str:
            """Convierte líneas de afiliaciones a HTML con superíndices."""
            html_lineas = []
            for linea in txt.splitlines():
                linea = linea.strip()
                if not linea: continue
                # Email: * correo@dominio → link mailto
                if re.match(r"^\*\s*[\w\.\-]+@[\w\-\.]+\.\w{2,}", linea):
                    email = re.search(r"[\w\.\-]+@[\w\-\.]+\.\w{2,}", linea)
                    if email:
                        e = email.group(0)
                        html_lineas.append(
                            f'<p class="email sin-sangria">'
                            f'* <a href="mailto:{e}">{esc(e)}</a></p>')
                    continue
                # Número inicial → superíndice
                m = re.match(r"^(\d+)\s+(.*)", linea, re.DOTALL)
                if m:
                    num, resto = m.group(1), m.group(2).strip()
                    html_lineas.append(
                        f'<p class="filiaciones sin-sangria">'
                        f'<sup>{num}</sup> {esc(resto)}</p>')
                else:
                    html_lineas.append(
                        f'<p class="filiaciones sin-sangria">{esc(linea)}</p>')
            return "\n".join(html_lineas)

        afil_html = _afil_a_html(self.afiliaciones_txt) if self.afiliaciones_txt else ""
        afil_inyectado = False

        autores_html_manual = ""
        autores_inyectado   = False
        primer_nivel1_emitido = False
        en_abstract_ingles  = False   # True dentro de Abstract / Non-technical Abstract
        if self.autores_orcid:
            autores_html_manual = (
                f'<p class="autores sin-sangria">'
                f'{_insertar_orcid("", self.autores_orcid)}</p>'
            )

        # Decidir qué referencias usar
        # Si hay externas → usar esas; los bloques "Referencia" del PDF → ignorar
        # Referencias: SOLO las del .txt externo.
        # Las referencias del PDF se ignoran siempre — el usuario las carga aparte.
        usar_refs_externas = bool(self.referencias_externas)
        refs_a_usar = self.referencias_externas if usar_refs_externas else []

        # ── Separar bloques (saltando zona de autores del PDF) ──────
        idx_refs_start = None
        for i, b in enumerate(self.datos_bloques):
            if b["menu"].get() == "Encabezado sección" and re.search(
                    r"referencia|reference", b["contenido"], re.I):
                idx_refs_start = i
                break

        cuerpo_bloques = []
        como_citar_lst = []
        fechas_mss_lst = []
        pies_pendientes= []

        for i, b in enumerate(self.datos_bloques):
            cls = b["menu"].get()
            # Saltar zona entre titulo-secundario y Resumen (autores rotos del PDF)
            if i in zona_autores_pdf:
                continue
            dentro_de_refs = (idx_refs_start is not None and i > idx_refs_start)

            if cls == "Cómo citar":
                if not dentro_de_refs:      # ignorar si viene de dentro de refs
                    como_citar_lst.append(b)
            elif cls == "Fecha manuscrito":
                if not dentro_de_refs:      # idem – evita DOIs de las citas
                    fechas_mss_lst.append(b)
            elif cls == "Imagen":
                pass                        # imágenes del PDF siempre ignoradas
            elif cls == "Referencia" and usar_refs_externas:
                pass                        # ignorar duplicados cuando hay .txt
            else:
                cuerpo_bloques.append(b)

        # Deduplicar por primeros 80 chars (por si el PDF los repite)
        def _dedup(lst):
            seen, out = set(), []
            for b in lst:
                key = b["contenido"].strip()[:80]
                if key not in seen:
                    seen.add(key); out.append(b)
            return out
        como_citar_lst = _dedup(como_citar_lst)
        fechas_mss_lst = _dedup(fechas_mss_lst)

        # Generar HTML
        lineas = [
            "<!DOCTYPE html>", '<html lang="es">',
            "<head>",
            '  <meta charset="UTF-8">',
            '  <meta name="viewport" content="width=device-width, initial-scale=1.0">',
            "  <title>Artículo</title>",
            HTML_CSS,
            "</head>", "<body><article>",
        ]

        en_refs = False

        for b in cuerpo_bloques:
            cls   = b["menu"].get()
            texto = esc(b["contenido"])
            ital  = b.get("italic", False)
            if cls == "Ignorar": continue

            # ── Si hay referencias externas y estamos dentro de la sección
            #    de referencias del PDF, suprimir TODO excepto nuevos
            #    encabezados de sección (que resetean en_refs).
            if usar_refs_externas and en_refs and cls != "Encabezado sección":
                continue

            if cls == "Título principal":
                lineas.append(f'<h1 class="titulo-principal">{texto}</h1>')
            elif cls == "Título secundario":
                lineas.append(f'<h2 class="titulo-secundario">{texto}</h2>')
                if autores_html_manual and not autores_inyectado:
                    lineas.append(autores_html_manual)
                    autores_inyectado = True
                # Inyectar afiliaciones del txt justo después de los autores
                if afil_html and not afil_inyectado:
                    lineas.append(afil_html)
                    afil_inyectado = True
            elif cls == "Autores":
                if not autores_html_manual:
                    lineas.append(f'<p class="autores sin-sangria">{_insertar_orcid(b["contenido"], None)}</p>')
            elif cls == "Filiación":
                lineas.append(f'<p class="filiaciones sin-sangria">{texto}</p>')
            elif cls == "Email / Metadatos":
                txt_link = re.sub(r"([\w\.\-]+@[\w\-\.]+\.\w{2,})",
                                  r'<a href="mailto:\1">\1</a>', texto)
                lineas.append(f'<p class="email sin-sangria">{txt_link}</p>')
            elif cls == "Encabezado sección":
                en_refs = bool(re.search(r"referencia|reference", texto, re.I))
                _SECCIONES_CON_LINEA = {
                    "resumen", "abstract", "resumen no técnico", "non-technical abstract",
                    "referencias", "references", "contribuciones de los autores",
                    "agradecimientos", "acknowledgements",
                }
                # Metadatos de cabecera: ISSN, volumen, fechas → letra pequeña
                _es_meta = bool(re.search(
                    r"issn|volumen\s+\d|vol\.\s*\d|núm\.\s*\d|p\.\s*\d{2,}"
                    r"|enero|febrero|marzo|abril|mayo|junio|julio|agosto"
                    r"|septiembre|octubre|noviembre|diciembre"
                    r"|january|february|march|april|june|july|august"
                    r"|september|october|november|december"
                    r"|\(\d{4}\)",
                    texto, re.I
                )) or re.match(r"^paleontolog[íi]a mexicana$", texto.strip(), re.I)
                # Solo Abstract y Non-technical Abstract en gris (inglés)
                _SECCIONES_GRISES = {
                    "abstract", "non-technical abstract",
                    "keywords",
                }
                # Secciones con línea divisora
                _SECCIONES_CON_LINEA = {
                    "resumen", "abstract", "resumen no técnico", "non-technical abstract",
                    "referencias", "references", "contribuciones de los autores",
                    "agradecimientos", "acknowledgements",
                    "conflicto de intereses", "conflict of interest",
                }
                txt_low = texto.strip().lower()
                es_gris   = txt_low in _SECCIONES_GRISES
                con_linea = txt_low in _SECCIONES_CON_LINEA
                # Detectar si entramos/salimos de Abstract en inglés
                _ABSTRACT_INGLES = {"abstract", "non-technical abstract"}
                if txt_low in _ABSTRACT_INGLES:
                    en_abstract_ingles = True
                elif txt_low not in _SECCIONES_GRISES:
                    en_abstract_ingles = False
                if _es_meta:
                    clase_h2 = 'seccion meta'
                elif con_linea and es_gris:
                    clase_h2 = 'seccion con-linea gris'
                elif con_linea:
                    clase_h2 = 'seccion con-linea'
                elif es_gris:
                    clase_h2 = 'seccion gris'
                else:
                    clase_h2 = 'seccion'
                lineas.append(f'<h2 class="{clase_h2}">{texto}</h2>')
                if en_refs and refs_a_usar:
                    lineas.append('<ol class="referencias">')
                    for ref in refs_a_usar:
                        lineas.append(f"  <li>{esc(ref)}</li>")
                    lineas.append("</ol>")
            elif cls == "Subencabezado":
                en_abstract_ingles = False
                if not primer_nivel1_emitido:
                    lineas.append(f'<h3 class="subseccion primer-nivel1">{texto}</h3>')
                    primer_nivel1_emitido = True
                else:
                    lineas.append(f'<h3 class="subseccion">{texto}</h3>')
            elif cls == "Subencabezado-bajo":
                lineas.append(f'<h3 class="subseccion-bajo">{texto}</h3>')
            elif cls == "Resumen / Abstract":
                tag = "abstract" if ital else "resumen"
                lineas.append(f'<p class="{tag} sin-sangria">{texto}</p>')
            elif cls == "Palabras clave":
                # "Palabras clave:" / "Keywords:" en negrita, resto normal
                t_kw = esc(b["contenido"])
                t_kw = re.sub(
                    r"^(Palabras\s+clave|Keywords)\s*[:\.]?\s*",
                    lambda m: f"<strong>{m.group(0).rstrip()}</strong> ",
                    t_kw, count=1, flags=re.IGNORECASE)
                lineas.append(f'<p class="keywords sin-sangria">{t_kw}</p>')
            elif cls in ("Normal", "Cuerpo"):
                if en_refs:
                    # Dentro de la sección Referencias del PDF → ignorar siempre
                    # (las refs externas ya se inyectaron con el ol.referencias)
                    continue
                else:
                    # Re-fusionar partes que siguen siendo continuación
                    # (§SUB§ siempre se trata aparte)
                    partes_raw = b["contenido"].split("\n\n")
                    partes_unidas = []
                    for parte in partes_raw:
                        parte = parte.strip()
                        if not parte: continue
                        if parte.startswith("§SUB§"):
                            partes_unidas.append(parte)
                        elif partes_unidas and not partes_unidas[-1].startswith("§SUB§"):
                            prev = partes_unidas[-1].rstrip()
                            if prev and prev[-1] not in ".?!:":
                                # Continuación — unir
                                if prev.endswith("-"):
                                    partes_unidas[-1] = prev[:-1] + parte
                                else:
                                    partes_unidas[-1] = prev + " " + parte
                            else:
                                partes_unidas.append(parte)
                        else:
                            partes_unidas.append(parte)

                    for parte in partes_unidas:
                        if parte.startswith("§SUB§"):
                            lineas.append(f'<h3 class="subseccion-bajo">{esc(parte[5:])}</h3>')
                        else:
                            tag_p = 'class="cuerpo"' if cls == "Cuerpo" else ""
                            # Cuerpo dentro de Abstract/Non-technical Abstract → cursiva gris
                            if en_abstract_ingles:
                                lineas.append(f'<p class="abstract sin-sangria">{esc(parte)}</p>')
                            else:
                                lineas.append(f'<p {tag_p}>{esc(parte)}</p>'.replace("  >", ">"))
            elif cls == "Referencia":
                lineas.append(f'<p style="padding-left:1.5em;text-indent:-1.5em;font-size:10pt;">{texto}</p>')
            elif cls == "Título tabla":
                pass   # suprimido del flujo inline — las tablas van por ancla
            elif cls == "Pie de figura":
                pies_pendientes.append(texto)

        # Post-referencias
        if como_citar_lst or fechas_mss_lst:
            lineas.append('<div class="post-referencias">')
            for b in como_citar_lst:
                lineas.append(f'<p class="como-citar">{esc(b["contenido"])}</p>')
            if fechas_mss_lst:
                lineas.append('<div class="fechas-manuscrito"><ul>')
                doi_items = []
                doi_seen  = set()
                for b in fechas_mss_lst:
                    t      = b["contenido"].strip()
                    doi_m  = re.search(r"(https?://doi\.org/\S+)", t)
                    if doi_m:
                        doi_clean = re.sub(r"\s+", "", doi_m.group(1))
                        if doi_clean not in doi_seen:
                            doi_seen.add(doi_clean)
                            doi_items.append(doi_clean)
                        t = t.replace(doi_m.group(1), "").strip(" .")
                    for parte in re.split(
                            r'(?<=[.!?])\s+(?=Manuscrito|Manuscript)', t):
                        parte = parte.strip()
                        if parte:
                            lineas.append(f"  <li>{esc(parte)}</li>")
                for doi in doi_items:
                    lineas.append(
                        f'  <li><a href="{doi}" target="_blank">{esc(doi)}</a></li>')
                lineas.append("</ul></div>")
            lineas.append("</div>")

        # ── Tablas: posicionadas por ancla, o al final si no tienen ancla ──
        lineas.append("</article></body></html>")
        html_body = "\n".join(lineas)

        if self.tablas_manuales:
            tablas_ordenadas = []
            for idx_t, t_item in enumerate(self.tablas_manuales, 1):
                ancla  = t_item.get("ancla", "").strip()
                titulo = t_item.get("titulo", "") or f"Tabla {idx_t}"
                thtml  = _excel_a_html_tabla(t_item["ruta"], t_item.get("hoja"))
                bloque = (
                    f'\n<div class="tabla-wrapper">\n'
                    f'<p class="tabla-titulo"><strong>Tabla {idx_t}.</strong> {esc(titulo)}</p>\n'
                    f'{thtml}\n</div>\n'
                )
                pos = -1
                if ancla:
                    ancla_norm = re.sub(r"[\u00ad\ufffc\ufffe]", "", ancla)
                    ancla_norm = re.sub(r"-\s+", "", ancla_norm)
                    ancla_norm = re.sub(r"\s+", " ", ancla_norm).strip()
                    muestra = ancla_norm[-60:].strip()
                    escaped = re.sub(r"([.+*?()\[\]{}\\|^$])", r"\\\1", muestra)
                    _spacer = "(?:[\N{SOFT HYPHEN}\ufffc]?\\s*(?:<[^>]+>)?\\s*)+"
                    pattern = escaped.replace("\\ ", _spacer)
                    m = re.search(pattern, html_body, re.IGNORECASE | re.DOTALL)
                    if m:
                        cierre = html_body.find("</p>", m.end())
                        pos = cierre + 4 if cierre != -1 else -1
                tablas_ordenadas.append((pos, idx_t, bloque))

            tablas_ordenadas.sort(key=lambda x: (x[0] == -1, x[0], x[1]))
            tablas_inline   = [(p, b) for p, _, b in tablas_ordenadas if p != -1]
            tablas_al_final = [b for p, _, b in tablas_ordenadas if p == -1]

            for pos, bloque in sorted(tablas_inline, key=lambda x: -x[0]):
                html_body = html_body[:pos] + bloque + html_body[pos:]

            if tablas_al_final:
                seccion = (
                    '\n<div class="figuras-finales">\n'
                    '<h2 style="text-align:center;font-size:10pt;font-weight:700;'
                    'margin-bottom:14px;">Tablas</h2>\n'
                    + "".join(tablas_al_final)
                    + "</div>\n"
                )
                html_body = html_body.replace("</article>", seccion + "</article>", 1)
        # Figuras: inline por ancla, o al final las que no tienen
        figs = self.figuras_manuales
        if figs:
            def _fig_html(i, fig):
                pie_txt = fig.get("pie", "")
                try:
                    src = _img_to_base64(fig["ruta"])
                except Exception:
                    src = "imagen.jpg"
                cap = (f"<strong>Figura {i}.</strong> {esc(pie_txt)}"
                       if pie_txt else f"<strong>Figura {i}.</strong>")
                return (f'<figure id="fig{i}" style="margin:18px auto;text-align:center;">\n'
                        f'  <img src="{src}" alt="Figura {i}" '
                        f'style="max-width:60%;max-height:420px;border:1px solid #bbb;">\n'
                        f'  <figcaption style="font-size:9pt;color:#1a1a1a;'
                        f'margin-top:5px;text-align:left;">{cap}</figcaption>\n'
                        f'</figure>')

            # (pos, fig_index, bloque)
            figs_inline  = []
            figs_al_final = []

            def _buscar_ancla_html(ancla, html):
                if not ancla: return -1
                # Normalizar ancla: quitar soft-hyphens (U+00AD, U+FFFC), guiones de corte
                ancla_norm = re.sub(r"[\u00ad\ufffc\ufffe]", "", ancla)
                ancla_norm = re.sub(r"-\s+", "", ancla_norm)
                ancla_norm = re.sub(r"\s+", " ", ancla_norm).strip()
                muestra = ancla_norm[-80:].strip()
                # Escapar y hacer tolerante a tags HTML y soft-hyphens en el HTML
                escaped = re.sub(r"([.+*?()\[\]{}\\|^$])", r"\\\1", muestra)
                # Espacios → tolerantes a soft-hyphens, tags, saltos
                spacer  = "(?:[\N{SOFT HYPHEN}\ufffc]?\\s*(?:<[^>]+>)?\\s*)+"
                pattern = escaped.replace("\\ ", spacer)
                m = re.search(pattern, html, re.IGNORECASE | re.DOTALL)
                if m:
                    cierre = html.find("</p>", m.end())
                    return cierre + 4 if cierre != -1 else -1
                return -1

            for i, fig in enumerate(figs, 1):
                ancla = fig.get("ancla", "").strip()
                bloque = "\n" + _fig_html(i, fig) + "\n"
                pos = _buscar_ancla_html(ancla, html_body)
                if pos != -1:
                    figs_inline.append((pos, i, bloque))
                else:
                    figs_al_final.append((i, fig))

            # Agrupar por posición: misma posición → concatenar en orden de figura
            # Insertar grupos de mayor a menor posición (para no desplazar)
            from collections import defaultdict
            grupos = defaultdict(list)
            for pos, idx, bloque in sorted(figs_inline, key=lambda x: (x[0], x[1])):
                grupos[pos].append(bloque)

            for pos in sorted(grupos.keys(), reverse=True):
                bloque_conjunto = "".join(grupos[pos])
                html_body = html_body[:pos] + bloque_conjunto + html_body[pos:]

            # Las sin ancla van al final (sin encabezado de sección)
            if figs_al_final:
                figs_html = '\n'
                for i, fig in figs_al_final:
                    figs_html += _fig_html(i, fig) + "\n"
                html_body = html_body.replace("</article>", figs_html + "</article>", 1)

        with open(ruta, "w", encoding="utf-8") as f:
            f.write(html_body)
        self._set_status(f"✓ HTML guardado en: {ruta}")


    # ═════════════════════════════════════════════════════════════
    # EXPORTAR XML  (próximamente)
    # ═════════════════════════════════════════════════════════════

    def evento_exportar_xml(self):
        self._set_status("⚠ Exportación XML en desarrollo.")

    # ═════════════════════════════════════════════════════════════
    # EXPORTAR EPUB  (próximamente)
    # ═════════════════════════════════════════════════════════════

    def evento_exportar_epub(self):
        self._set_status("⚠ Exportación EPUB en desarrollo.")

if __name__ == "__main__":
    app = LimpiadorEditorialApp()
    app.mainloop()
